from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    load_workbook = None


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
DELIVERY_DIR = INPUT_DIR / "deliveries"
ADJUSTMENT_DIR = INPUT_DIR / "adjustments"
SQUARE_EXPORT_DIR = INPUT_DIR / "square_exports"
OUTPUT_DIR = BASE_DIR / "outputs"

MASTER_PATH = OUTPUT_DIR / "square_master_inventory.csv"
LEGACY_MASTER_PATH = BASE_DIR / "square_master_inventory.csv"
RECEIVING_UPDATE_PATH = OUTPUT_DIR / "square_receiving_update.csv"
RECEIVING_UPDATE_XLSX_PATH = OUTPUT_DIR / "square_receiving_update.xlsx"
RECEIVING_AUDIT_PATH = OUTPUT_DIR / "receiving_update_audit.csv"
RECEIVING_ISSUES_PATH = OUTPUT_DIR / "receiving_update_issues.csv"
RECEIVING_SUMMARY_PATH = OUTPUT_DIR / "receiving_update_summary.txt"

PREFERRED_LOCATION = "AZCS"

DELIVERY_FIELD_ALIASES = {
    "Transaction Date": ["Transaction Date", "Date", "Delivery Date", "Received Date"],
    "Vendor": ["Vendor", "Supplier"],
    "SKU": ["SKU", "Square SKU", "Item SKU"],
    "Vendor Code": ["Vendor Code", "Default Vendor Code", "Item Number", "Part Number", "Product Code", "Vendor SKU"],
    "GTIN": ["GTIN", "UPC", "Barcode"],
    "Item Name": ["Item Name", "Product Name", "Description", "Title"],
    "Quantity Received": ["Quantity Received", "Qty Received", "Quantity", "Qty", "Received Qty", "Received"],
    "Unit Cost": ["Unit Cost", "Cost", "Unit Price", "Price"],
    "Reference": ["Reference", "PO Number", "Order Number", "Packing Slip", "Invoice", "Document Number"],
    "Notes": ["Notes", "Memo", "Comment"],
}

ADJUSTMENT_FIELD_ALIASES = {
    "Transaction Date": ["Transaction Date", "Date", "Adjustment Date"],
    "SKU": ["SKU", "Square SKU", "Item SKU"],
    "Vendor Code": ["Vendor Code", "Default Vendor Code", "Item Number", "Part Number", "Product Code", "Vendor SKU"],
    "GTIN": ["GTIN", "UPC", "Barcode"],
    "Item Name": ["Item Name", "Product Name", "Description", "Title"],
    "Quantity Change": ["Quantity Change", "Adjustment", "Adjustment Qty", "Qty Change", "Quantity"],
    "Reference": ["Reference", "Reason", "Adjustment Reason", "Document Number"],
    "Notes": ["Notes", "Memo", "Comment"],
}


@dataclass
class MasterLookup:
    sku_set: set[str]
    vendor_code_map: dict[str, set[str]]
    gtin_map: dict[str, set[str]]
    name_map: dict[str, set[str]]


@dataclass
class StockChange:
    received_qty: Decimal = Decimal("0")
    adjusted_qty: Decimal = Decimal("0")
    source_files: set[str] = field(default_factory=set)
    references: set[str] = field(default_factory=set)
    match_types: set[str] = field(default_factory=set)
    last_transaction_date: str = ""
    last_unit_cost: str = ""


def normalize_cell(value: object) -> str:
    text = str(value or "")
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_name(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_cell(value).upper())


def parse_decimal(value: object, field_name: str, path: Path, row_number: int, issues: list[dict[str, str]]) -> Decimal | None:
    text = normalize_cell(value).replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        issues.append(
            {
                "source_file": path.name,
                "row_number": str(row_number),
                "issue_type": "invalid_number",
                "sku": "",
                "details": f"Could not parse {field_name} value '{text}'.",
            }
        )
        return None


def parse_date(value: object) -> str:
    text = normalize_cell(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def format_quantity(value: Decimal) -> str:
    normalized = value.normalize()
    rendered = format(normalized, "f")
    return rendered.rstrip("0").rstrip(".") if "." in rendered else rendered


def read_csv_matrix(path: Path) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                for index, row in enumerate(csv.reader(handle), start=1):
                    rows.append((index, [normalize_cell(cell) for cell in row]))
            return rows
        except UnicodeDecodeError as exc:
            rows = []
            last_error = exc
    if last_error:
        raise last_error
    raise RuntimeError(f"Could not read {path}")


def read_xlsx_matrix(path: Path) -> list[tuple[int, list[str]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsx files.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[tuple[int, list[str]]] = []
    for sheet in workbook.worksheets:
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cleaned = [normalize_cell(cell) for cell in row]
            rows.append((index, cleaned))
    return rows


def read_tabular_matrix(path: Path) -> list[tuple[int, list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_matrix(path)
    if suffix == ".xlsx":
        return read_xlsx_matrix(path)
    raise ValueError(f"Unsupported file type: {path.suffix}")


def locate_header_row(rows: list[tuple[int, list[str]]], predicate) -> tuple[int, list[str]]:
    for row_number, row in rows:
        if predicate(row):
            last_non_empty = max((index for index, value in enumerate(row) if value), default=-1)
            return row_number, row[: last_non_empty + 1]
    raise ValueError("Could not find the expected header row.")


def row_to_dict(headers: list[str], row: list[str]) -> dict[str, str]:
    values = row[: len(headers)] + [""] * max(0, len(headers) - len(row))
    return {headers[index]: values[index] for index in range(len(headers))}


def read_square_export_rows(path: Path) -> list[dict[str, str]]:
    matrix = read_tabular_matrix(path)

    def is_header(row: list[str]) -> bool:
        return "SKU" in row and any(header.startswith("Current Quantity ") for header in row)

    header_row_number, headers = locate_header_row(matrix, is_header)
    rows: list[dict[str, str]] = []
    for row_number, row in matrix:
        if row_number <= header_row_number:
            continue
        if not any(row):
            continue
        record = row_to_dict(headers, row)
        record["__row_number"] = str(row_number)
        rows.append(record)
    return rows


def standardize_row(row: dict[str, str], aliases: dict[str, list[str]]) -> dict[str, str]:
    standardized: dict[str, str] = {}
    for canonical_name, options in aliases.items():
        value = ""
        for option in options:
            if option in row and normalize_cell(row[option]):
                value = normalize_cell(row[option])
                break
        standardized[canonical_name] = value
    return standardized


def read_delivery_rows(path: Path) -> list[dict[str, str]]:
    matrix = read_tabular_matrix(path)

    def is_header(row: list[str]) -> bool:
        return "Quantity Received" in row or (
            any(option in row for option in DELIVERY_FIELD_ALIASES["SKU"] + DELIVERY_FIELD_ALIASES["Vendor Code"] + DELIVERY_FIELD_ALIASES["GTIN"] + DELIVERY_FIELD_ALIASES["Item Name"])
            and any(option in row for option in DELIVERY_FIELD_ALIASES["Quantity Received"])
        )

    header_row_number, headers = locate_header_row(matrix, is_header)
    rows: list[dict[str, str]] = []
    for row_number, row in matrix:
        if row_number <= header_row_number:
            continue
        if not any(row):
            continue
        record = standardize_row(row_to_dict(headers, row), DELIVERY_FIELD_ALIASES)
        record["__row_number"] = str(row_number)
        rows.append(record)
    return rows


def read_adjustment_rows(path: Path) -> list[dict[str, str]]:
    matrix = read_tabular_matrix(path)

    def is_header(row: list[str]) -> bool:
        return "Quantity Change" in row or (
            any(option in row for option in ADJUSTMENT_FIELD_ALIASES["SKU"] + ADJUSTMENT_FIELD_ALIASES["Vendor Code"] + ADJUSTMENT_FIELD_ALIASES["GTIN"] + ADJUSTMENT_FIELD_ALIASES["Item Name"])
            and any(option in row for option in ADJUSTMENT_FIELD_ALIASES["Quantity Change"])
        )

    header_row_number, headers = locate_header_row(matrix, is_header)
    rows: list[dict[str, str]] = []
    for row_number, row in matrix:
        if row_number <= header_row_number:
            continue
        if not any(row):
            continue
        record = standardize_row(row_to_dict(headers, row), ADJUSTMENT_FIELD_ALIASES)
        record["__row_number"] = str(row_number)
        rows.append(record)
    return rows


def resolve_master_path() -> Path | None:
    if MASTER_PATH.exists():
        return MASTER_PATH
    if LEGACY_MASTER_PATH.exists():
        return LEGACY_MASTER_PATH
    return None


def read_master_rows(path: Path | None) -> list[dict[str, str]]:
    if path is None:
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def build_master_lookup(master_rows: list[dict[str, str]]) -> MasterLookup:
    sku_set: set[str] = set()
    vendor_code_map: dict[str, set[str]] = defaultdict(set)
    gtin_map: dict[str, set[str]] = defaultdict(set)
    name_map: dict[str, set[str]] = defaultdict(set)

    for row in master_rows:
        sku = normalize_cell(row.get("SKU", ""))
        if not sku:
            continue
        sku_set.add(sku)

        vendor_code = normalize_cell(row.get("Default Vendor Code", ""))
        if vendor_code:
            vendor_code_map[vendor_code].add(sku)

        gtin = normalize_cell(row.get("GTIN", ""))
        if gtin:
            gtin_map[gtin].add(sku)

        for name_field in ("Item Name", "Customer-facing Name"):
            name_key = normalize_name(row.get(name_field, ""))
            if name_key:
                name_map[name_key].add(sku)

    return MasterLookup(
        sku_set=sku_set,
        vendor_code_map=dict(vendor_code_map),
        gtin_map=dict(gtin_map),
        name_map=dict(name_map),
    )


def resolve_latest_square_export() -> Path:
    candidates = sorted(
        [
            path
            for path in SQUARE_EXPORT_DIR.glob("*")
            if path.is_file() and path.suffix.lower() in {".csv", ".xlsx"}
        ],
        key=lambda path: path.stat().st_mtime,
    )
    if not candidates:
        raise FileNotFoundError(f"No Square export file found in {SQUARE_EXPORT_DIR}")
    return candidates[-1]


def choose_location_columns(fieldnames: list[str], location_name: str) -> tuple[str, str, str | None]:
    current = f"Current Quantity {location_name}"
    new = f"New Quantity {location_name}"
    enabled = f"Enabled {location_name}"

    if current in fieldnames and new in fieldnames:
        return current, new, enabled if enabled in fieldnames else None

    fallback_current = next((field for field in fieldnames if field.startswith("Current Quantity ")), "")
    fallback_new = next((field for field in fieldnames if field.startswith("New Quantity ")), "")
    if not fallback_current or not fallback_new:
        raise ValueError(f"Could not find quantity columns for location {location_name}.")
    fallback_enabled = fallback_new.replace("New Quantity ", "Enabled ")
    return fallback_current, fallback_new, fallback_enabled if fallback_enabled in fieldnames else None


def resolve_lookup_match(candidates: set[str], export_skus: set[str]) -> tuple[str | None, str]:
    in_export = sorted(sku for sku in candidates if sku in export_skus)
    if len(in_export) == 1:
        return in_export[0], "master_lookup"
    if len(in_export) > 1:
        return None, "ambiguous_export"
    if candidates:
        return None, "missing_from_export"
    return None, "no_match"


def resolve_sku_from_row(row: dict[str, str], export_skus: set[str], master_lookup: MasterLookup) -> tuple[str | None, str]:
    sku = normalize_cell(row.get("SKU", ""))
    if sku:
        if sku in export_skus:
            return sku, "sku"
        if sku in master_lookup.sku_set:
            return None, "sku_missing_from_export"

    vendor_code = normalize_cell(row.get("Vendor Code", ""))
    if vendor_code:
        resolved_sku, status = resolve_lookup_match(master_lookup.vendor_code_map.get(vendor_code, set()), export_skus)
        if resolved_sku:
            return resolved_sku, "vendor_code"
        if status != "no_match":
            return None, f"vendor_code_{status}"

    gtin = normalize_cell(row.get("GTIN", ""))
    if gtin:
        resolved_sku, status = resolve_lookup_match(master_lookup.gtin_map.get(gtin, set()), export_skus)
        if resolved_sku:
            return resolved_sku, "gtin"
        if status != "no_match":
            return None, f"gtin_{status}"

    item_name = normalize_name(row.get("Item Name", ""))
    if item_name:
        resolved_sku, status = resolve_lookup_match(master_lookup.name_map.get(item_name, set()), export_skus)
        if resolved_sku:
            return resolved_sku, "item_name"
        if status != "no_match":
            return None, f"item_name_{status}"

    return None, "no_match"


def merge_stock_change(change: StockChange, qty: Decimal, reference: str, source_file: str, match_type: str, tx_date: str, unit_cost: str = "") -> None:
    change.source_files.add(source_file)
    if reference:
        change.references.add(reference)
    if match_type:
        change.match_types.add(match_type)
    if tx_date and tx_date > change.last_transaction_date:
        change.last_transaction_date = tx_date
    if unit_cost:
        change.last_unit_cost = unit_cost
    change.received_qty += qty


def merge_adjustment(change: StockChange, qty: Decimal, reference: str, source_file: str, match_type: str, tx_date: str) -> None:
    change.source_files.add(source_file)
    if reference:
        change.references.add(reference)
    if match_type:
        change.match_types.add(match_type)
    if tx_date and tx_date > change.last_transaction_date:
        change.last_transaction_date = tx_date
    change.adjusted_qty += qty


def load_stock_changes(export_rows: list[dict[str, str]], master_lookup: MasterLookup) -> tuple[dict[str, StockChange], list[dict[str, str]], list[str], list[str]]:
    export_skus = {normalize_cell(row.get("SKU", "")) for row in export_rows if normalize_cell(row.get("SKU", ""))}
    issues: list[dict[str, str]] = []
    changes: dict[str, StockChange] = defaultdict(StockChange)
    delivery_files: list[str] = []
    adjustment_files: list[str] = []

    for path in sorted(DELIVERY_DIR.glob("*")):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx"}:
            continue
        delivery_files.append(path.name)
        for row in read_delivery_rows(path):
            row_number = int(row.get("__row_number", "0") or "0")
            qty = parse_decimal(row.get("Quantity Received", ""), "Quantity Received", path, row_number, issues)
            if qty is None:
                continue
            resolved_sku, match_type = resolve_sku_from_row(row, export_skus, master_lookup)
            if not resolved_sku:
                issues.append(
                    {
                        "source_file": path.name,
                        "row_number": str(row_number),
                        "issue_type": "unmatched_delivery_row",
                        "sku": normalize_cell(row.get("SKU", "")),
                        "details": f"Could not resolve delivery row by SKU/vendor code/GTIN/item name. Match status: {match_type}.",
                    }
                )
                continue
            merge_stock_change(
                changes[resolved_sku],
                qty,
                normalize_cell(row.get("Reference", "")),
                path.name,
                match_type,
                parse_date(row.get("Transaction Date", "")),
                normalize_cell(row.get("Unit Cost", "")),
            )

    for path in sorted(ADJUSTMENT_DIR.glob("*")):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx"}:
            continue
        adjustment_files.append(path.name)
        for row in read_adjustment_rows(path):
            row_number = int(row.get("__row_number", "0") or "0")
            qty = parse_decimal(row.get("Quantity Change", ""), "Quantity Change", path, row_number, issues)
            if qty is None:
                continue
            resolved_sku, match_type = resolve_sku_from_row(row, export_skus, master_lookup)
            if not resolved_sku:
                issues.append(
                    {
                        "source_file": path.name,
                        "row_number": str(row_number),
                        "issue_type": "unmatched_adjustment_row",
                        "sku": normalize_cell(row.get("SKU", "")),
                        "details": f"Could not resolve adjustment row by SKU/vendor code/GTIN/item name. Match status: {match_type}.",
                    }
                )
                continue
            merge_adjustment(
                changes[resolved_sku],
                qty,
                normalize_cell(row.get("Reference", "")),
                path.name,
                match_type,
                parse_date(row.get("Transaction Date", "")),
            )

    return changes, issues, delivery_files, adjustment_files


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> str:
    if Workbook is None:
        return "openpyxl not available"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Receiving Update"
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    try:
        workbook.save(path)
    except PermissionError:
        return f"file locked: {path}"
    return str(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SQUARE_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    export_path = resolve_latest_square_export()
    export_rows = read_square_export_rows(export_path)
    if not export_rows:
        raise ValueError("The Square export file has no item rows.")

    master_lookup = build_master_lookup(read_master_rows(resolve_master_path()))
    changes, issues, delivery_files, adjustment_files = load_stock_changes(export_rows, master_lookup)

    fieldnames = [key for key in export_rows[0].keys() if not key.startswith("__")]
    current_col, new_col, enabled_col = choose_location_columns(fieldnames, PREFERRED_LOCATION)

    changed_rows: list[dict[str, str]] = []
    audit_rows: list[dict[str, str]] = []

    for export_row in export_rows:
        sku = normalize_cell(export_row.get("SKU", ""))
        change = changes.get(sku)
        if not change:
            continue

        current_qty = parse_decimal(export_row.get(current_col, ""), current_col, export_path, int(export_row.get("__row_number", "0") or "0"), issues)
        current_qty = current_qty if current_qty is not None else Decimal("0")
        qty_delta = change.received_qty + change.adjusted_qty
        new_qty = current_qty + qty_delta

        updated = {key: export_row.get(key, "") for key in fieldnames}
        updated[new_col] = format_quantity(new_qty)
        if enabled_col and not normalize_cell(updated.get(enabled_col, "")):
            updated[enabled_col] = "Y"
        changed_rows.append(updated)

        audit_rows.append(
            {
                "SKU": sku,
                "Item Name": normalize_cell(export_row.get("Item Name", "")),
                "Current Quantity": format_quantity(current_qty),
                "Quantity Received": format_quantity(change.received_qty),
                "Quantity Adjusted": format_quantity(change.adjusted_qty),
                "New Quantity": format_quantity(new_qty),
                "Location Current Quantity Column": current_col,
                "Location New Quantity Column": new_col,
                "Match Types": ", ".join(sorted(change.match_types)),
                "Source Files": ", ".join(sorted(change.source_files)),
                "References": ", ".join(sorted(change.references)),
                "Last Transaction Date": change.last_transaction_date,
                "Last Unit Cost": change.last_unit_cost,
            }
        )

    audit_fieldnames = [
        "SKU",
        "Item Name",
        "Current Quantity",
        "Quantity Received",
        "Quantity Adjusted",
        "New Quantity",
        "Location Current Quantity Column",
        "Location New Quantity Column",
        "Match Types",
        "Source Files",
        "References",
        "Last Transaction Date",
        "Last Unit Cost",
    ]

    write_csv(RECEIVING_UPDATE_PATH, fieldnames, changed_rows)
    xlsx_status = write_xlsx(RECEIVING_UPDATE_XLSX_PATH, fieldnames, changed_rows)
    write_csv(RECEIVING_AUDIT_PATH, audit_fieldnames, audit_rows)
    write_csv(RECEIVING_ISSUES_PATH, ["source_file", "row_number", "issue_type", "sku", "details"], issues)

    total_received = sum((change.received_qty for change in changes.values()), Decimal("0"))
    total_adjusted = sum((change.adjusted_qty for change in changes.values()), Decimal("0"))
    summary_lines = [
        f"Square export source: {export_path}",
        f"Square receiving update CSV: {RECEIVING_UPDATE_PATH}",
        f"Square receiving update Excel: {xlsx_status}",
        f"Receiving audit: {RECEIVING_AUDIT_PATH}",
        f"Issues file: {RECEIVING_ISSUES_PATH}",
        f"Delivery files processed: {len(delivery_files)}",
        f"Adjustment files processed: {len(adjustment_files)}",
        f"Changed SKUs in output: {len(changed_rows)}",
        f"Total quantity received: {format_quantity(total_received)}",
        f"Total quantity adjusted: {format_quantity(total_adjusted)}",
        f"Updated location current quantity column: {current_col}",
        f"Updated location new quantity column: {new_col}",
        f"Updated location enabled column: {enabled_col or '[unchanged]'}",
        f"Issue rows: {len(issues)}",
        "Workflow note: this file only includes rows with stock activity, so the import touches only the SKUs from the current delivery batch.",
        "Workflow note: use a fresh Square export and import the receiving update in the same after-hours session so the exported current quantities stay accurate.",
    ]
    RECEIVING_SUMMARY_PATH.write_text("\n".join(summary_lines), encoding="utf-8")
    print("\n".join(summary_lines))


if __name__ == "__main__":
    main()
