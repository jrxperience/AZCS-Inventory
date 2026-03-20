from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    Font = None
    get_column_letter = None


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
SALES_DIR = INPUT_DIR / "sales"
SALES_OVERRIDE_DIR = INPUT_DIR / "sales_match_overrides"
OUTPUT_DIR = BASE_DIR / "outputs"

MASTER_PATH = OUTPUT_DIR / "square_master_inventory.csv"
LEGACY_MASTER_PATH = BASE_DIR / "square_master_inventory.csv"
SALES_MATCH_AUDIT_PATH = OUTPUT_DIR / "sales_item_match_audit.csv"
SALES_MATCH_AUDIT_XLSX_PATH = OUTPUT_DIR / "sales_item_match_audit.xlsx"
SALES_MATCH_REVIEW_PATH = OUTPUT_DIR / "sales_item_match_review.csv"
SALES_CATALOG_SIGNALS_PATH = OUTPUT_DIR / "sales_catalog_signals.csv"
SALES_MATCH_SUMMARY_PATH = OUTPUT_DIR / "sales_match_summary.txt"
SALES_MATCH_ISSUES_PATH = OUTPUT_DIR / "sales_match_issues.csv"

STOPWORDS = {
    "A",
    "AN",
    "AND",
    "AT",
    "AVAILABLE",
    "BLACK",
    "BLUE",
    "BROWN",
    "DEG",
    "FOR",
    "GRAY",
    "GREEN",
    "IN",
    "LT",
    "MAX",
    "NOW",
    "OF",
    "OR",
    "ORANGE",
    "PER",
    "PSI",
    "RED",
    "REGULAR",
    "SAFETY",
    "THE",
    "TO",
    "W",
    "WHITE",
    "WITH",
    "X",
    "YELLOW",
}

CATEGORY_VENDOR_MAP = {
    "CHEMICALS": {"EACOCHEM", "TRIDENT", "FRONT9", "ENVIROBIOCLEANER", "MPWSR"},
    "EACO CHEM": {"EACOCHEM"},
    "EBC": {"ENVIROBIOCLEANER"},
    "F9": {"FRONT9"},
    "MANATEE": {"MPWSR"},
    "PARTS": {"MPWSR", "BARRENS", "BE"},
    "SAND": {"TRIDENT"},
    "SORBO": {"JRACENSTEIN"},
    "SWEEP AWAY": {"TRIDENT"},
    "TAGAWAY TAGINATOR": {"JRACENSTEIN"},
    "TAGINATOR": {"JRACENSTEIN"},
    "TOOLS": {"MPWSR", "BARRENS", "JRACENSTEIN", "BE"},
    "TRIDENT": {"TRIDENT"},
    "TUCKER": {"JRACENSTEIN", "TUCKER"},
    "WINDOW CLEANING": {"JRACENSTEIN"},
}

NON_INVENTORY_CATEGORIES = {"SERVICE"}
NON_INVENTORY_ITEMS = {
    "REPAIR SERVICE LABOR",
    "SHIPPING AND DELIVERY",
}


@dataclass
class MasterRecord:
    row: dict[str, str]
    display_name: str
    normalized_name: str
    base_name: str
    base_name_code: str
    vendor_key: str
    reporting_key: str
    categories_key: str
    sku_code: str
    vendor_code: str
    gtin_code: str
    tokens: set[str]
    number_tokens: set[str]
    measures: dict[str, set[str]]


@dataclass
class SalesAggregate:
    category: str
    item: str
    sku: str
    gtin: str
    lines: int = 0
    quantity: Decimal = Decimal("0")
    gross_sales: Decimal = Decimal("0")
    discounts: Decimal = Decimal("0")
    net_sales: Decimal = Decimal("0")
    tax: Decimal = Decimal("0")
    first_sale_date: str = ""
    last_sale_date: str = ""


@dataclass
class MatchOverride:
    master_sku: str
    notes: str = ""


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_decimal(value: str) -> Decimal:
    text = str(value or "").strip().replace("$", "").replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def parse_date(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def format_money(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'))}"


def format_decimal(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def normalize_text(value: str) -> str:
    text = str(value or "").upper().strip()
    if not text:
        return ""
    replacements = {
        "&": " AND ",
        "1/2": " HALF ",
        "GALLONS": " GAL ",
        "GALLON": " GAL ",
        "GAL.": " GAL ",
        "KITS": " KIT ",
        "OUNCES": " OZ ",
        "OUNCE": " OZ ",
        "OZ.": " OZ ",
        "POUNDS": " LB ",
        "POUND": " LB ",
        "LBS": " LB ",
        "GROUNDS KEEPER": " GROUNDSKEEPER ",
        "TRIGGER SPRAY": " ",
        "READY TO USE": " RTU ",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def canonical_code(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def is_non_inventory_sales_item(category: str, item: str) -> bool:
    category_key = normalize_text(category)
    item_key = normalize_text(item)
    if category_key in NON_INVENTORY_CATEGORIES:
        return True
    return item_key in NON_INVENTORY_ITEMS


def significant_tokens(normalized_text: str) -> set[str]:
    tokens: set[str] = set()
    for token in normalized_text.split():
        if len(token) <= 1:
            continue
        if token in STOPWORDS and not any(character.isdigit() for character in token):
            continue
        if len(token) < 3 and not any(character.isdigit() for character in token):
            continue
        tokens.add(token)
    return tokens


def extract_number_tokens(tokens: set[str]) -> set[str]:
    return {token for token in tokens if any(character.isdigit() for character in token)}


def extract_measures(normalized_text: str) -> dict[str, set[str]]:
    measures: dict[str, set[str]] = defaultdict(set)
    tokens = normalized_text.split()
    for index, token in enumerate(tokens[:-1]):
        next_token = tokens[index + 1]
        if token.isdigit() and next_token in {"GAL", "OZ", "LB", "KIT"}:
            measures[next_token].add(token)
    if "HALF" in tokens:
        measures["SPECIAL"].add("HALF")
    if "FULL" in tokens:
        measures["SPECIAL"].add("FULL")
    if "KIT" in tokens:
        measures["SPECIAL"].add("KIT")
    return dict(measures)


def strip_pack_descriptors(normalized_text: str) -> str:
    text = re.sub(r"\b\d+\s+(GAL|OZ|LB|KIT)\b", " ", normalized_text)
    text = re.sub(r"\b(PAIL|DRUM|KIT|HALF|FULL)\b", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def infer_sales_measures(category: str, sales_name: str, measures: dict[str, set[str]]) -> tuple[dict[str, set[str]], str]:
    inferred = {key: set(values) for key, values in measures.items()}
    base_name = strip_pack_descriptors(sales_name)
    category_key = normalize_text(category)
    tokens = sales_name.split()
    if category_key in {"CHEMICALS", "EACO CHEM", "F9", "TRIDENT", "EBC", "SWEEP AWAY"} and "GAL" not in inferred and tokens:
        last_token = tokens[-1]
        if last_token in {"1", "5", "55"}:
            inferred.setdefault("GAL", set()).add(last_token)
            base_name = re.sub(rf"\b{last_token}\b$", "", base_name).strip()
    return inferred, re.sub(r"\s+", " ", base_name).strip()


def resolve_master_path() -> Path:
    if MASTER_PATH.exists():
        return MASTER_PATH
    if LEGACY_MASTER_PATH.exists():
        return LEGACY_MASTER_PATH
    raise FileNotFoundError("square_master_inventory.csv was not found in outputs/ or the repo root.")


def resolve_sales_path() -> Path:
    local_candidates = sorted(SALES_DIR.glob("*.csv"), key=lambda path: path.stat().st_mtime, reverse=True)
    if local_candidates:
        return local_candidates[0]

    downloads_dir = Path.home() / "Downloads"
    download_candidates = sorted(downloads_dir.glob("Sales by item*.csv"), key=lambda path: path.stat().st_mtime, reverse=True)
    if download_candidates:
        return download_candidates[0]

    raise FileNotFoundError("No sales CSV was found in inputs/sales/ or Downloads matching 'Sales by item*.csv'.")


def build_master_records(
    master_rows: list[dict[str, str]],
) -> tuple[
    list[MasterRecord],
    dict[str, list[int]],
    dict[str, list[int]],
    dict[str, list[int]],
    dict[str, list[int]],
    dict[str, list[int]],
    dict[str, set[int]],
]:
    records: list[MasterRecord] = []
    sku_index: dict[str, list[int]] = defaultdict(list)
    gtin_index: dict[str, list[int]] = defaultdict(list)
    name_index: dict[str, list[int]] = defaultdict(list)
    base_name_index: dict[str, list[int]] = defaultdict(list)
    base_name_code_index: dict[str, list[int]] = defaultdict(list)
    token_index: dict[str, set[int]] = defaultdict(set)

    for index, row in enumerate(master_rows):
        display_name = str(row.get("Customer-facing Name", "")).strip() or str(row.get("Item Name", "")).strip()
        normalized_name = normalize_text(display_name)
        vendor_key = normalize_text(row.get("Default Vendor Name", ""))
        reporting_key = normalize_text(row.get("Reporting Category", ""))
        categories_key = normalize_text(row.get("Categories", ""))
        sku_code = canonical_code(row.get("SKU", ""))
        vendor_code = canonical_code(row.get("Default Vendor Code", ""))
        gtin_code = canonical_code(row.get("GTIN", ""))
        tokens = significant_tokens(normalized_name)
        measures = extract_measures(normalized_name)
        record = MasterRecord(
            row=row,
            display_name=display_name,
            normalized_name=normalized_name,
            base_name=strip_pack_descriptors(normalized_name),
            base_name_code=canonical_code(strip_pack_descriptors(normalized_name)),
            vendor_key=vendor_key,
            reporting_key=reporting_key,
            categories_key=categories_key,
            sku_code=sku_code,
            vendor_code=vendor_code,
            gtin_code=gtin_code,
            tokens=tokens,
            number_tokens=extract_number_tokens(tokens),
            measures=measures,
        )
        records.append(record)

        if sku_code:
            sku_index[sku_code].append(index)
        if vendor_code:
            sku_index[vendor_code].append(index)
        if gtin_code:
            gtin_index[gtin_code].append(index)
        if normalized_name:
            name_index[normalized_name].append(index)
        if record.base_name:
            base_name_index[record.base_name].append(index)
        if record.base_name_code:
            base_name_code_index[record.base_name_code].append(index)
        for token in tokens:
            token_index[token].add(index)

    return records, sku_index, gtin_index, name_index, base_name_index, base_name_code_index, token_index


def build_sales_aggregates(sales_rows: list[dict[str, str]]) -> dict[tuple[str, str, str, str], SalesAggregate]:
    aggregates: dict[tuple[str, str, str, str], SalesAggregate] = {}
    for row in sales_rows:
        category = str(row.get("Category", "")).strip()
        item = str(row.get("Item", "")).strip()
        sku = str(row.get("SKU", "")).strip()
        gtin = str(row.get("GTIN", "")).strip()
        key = (category, item, sku, gtin)
        current = aggregates.get(key)
        if current is None:
            current = SalesAggregate(category=category, item=item, sku=sku, gtin=gtin)
            aggregates[key] = current

        current.lines += 1
        current.quantity += parse_decimal(row.get("Qty", ""))
        current.gross_sales += parse_decimal(row.get("Gross Sales", ""))
        current.discounts += parse_decimal(row.get("Discounts", ""))
        current.net_sales += parse_decimal(row.get("Net Sales", ""))
        current.tax += parse_decimal(row.get("Tax", ""))
        sale_date = parse_date(row.get("Date", ""))
        if sale_date:
            if not current.first_sale_date or sale_date < current.first_sale_date:
                current.first_sale_date = sale_date
            if not current.last_sale_date or sale_date > current.last_sale_date:
                current.last_sale_date = sale_date

    return aggregates


def load_match_overrides(master_rows: list[dict[str, str]], issues: list[dict[str, str]]) -> tuple[dict[str, MatchOverride], dict[tuple[str, str, str], MatchOverride]]:
    known_skus = {str(row.get("SKU", "")).strip() for row in master_rows if str(row.get("SKU", "")).strip()}
    sku_overrides: dict[str, MatchOverride] = {}
    item_overrides: dict[tuple[str, str, str], MatchOverride] = {}

    for path in sorted(SALES_OVERRIDE_DIR.glob("*.csv")):
        rows = read_csv_rows(path)
        for row_number, row in enumerate(rows, start=2):
            sales_sku = canonical_code(row.get("Sales SKU", ""))
            sales_item = normalize_text(row.get("Sales Item", ""))
            sales_category = normalize_text(row.get("Sales Category", ""))
            master_sku = str(row.get("Master SKU", "")).strip()
            notes = str(row.get("Notes", "")).strip()

            if not master_sku:
                continue
            if master_sku not in known_skus:
                issues.append(
                    {
                        "source_file": path.name,
                        "row_number": str(row_number),
                        "issue_type": "unknown_master_sku",
                        "sales_sku": sales_sku,
                        "details": f"Override Master SKU '{master_sku}' does not exist in the master inventory.",
                    }
                )
                continue

            override = MatchOverride(master_sku=master_sku, notes=notes)
            if sales_sku:
                sku_overrides[sales_sku] = override
            if sales_item:
                item_overrides[(sales_category, sales_item, sales_sku)] = override

    return sku_overrides, item_overrides


def sales_category_candidates(category: str, records: list[MasterRecord]) -> set[int]:
    category_key = normalize_text(category)
    mapped_vendors = CATEGORY_VENDOR_MAP.get(category_key)
    if not mapped_vendors:
        return set(range(len(records)))
    candidate_indexes = {
        index
        for index, record in enumerate(records)
        if record.vendor_key in mapped_vendors or record.reporting_key in mapped_vendors or record.categories_key in mapped_vendors
    }
    return candidate_indexes or set(range(len(records)))


def code_tokens_from_sales_item(item: str, sku: str) -> set[str]:
    codes: set[str] = set()
    normalized_item = normalize_text(item)
    direct_code = canonical_code(sku)
    if len(direct_code) >= 4:
        codes.add(direct_code)
    compact_item = canonical_code(item)
    if len(compact_item) >= 6 and len(normalized_item.split()) <= 3:
        codes.add(compact_item)
    for token in normalized_item.split():
        compact = canonical_code(token)
        if len(compact) >= 4 and any(character.isdigit() for character in compact):
            codes.add(compact)
    return codes


def category_fit(category: str, record: MasterRecord) -> bool:
    category_key = normalize_text(category)
    mapped_vendors = CATEGORY_VENDOR_MAP.get(category_key)
    if not mapped_vendors:
        return False
    return record.vendor_key in mapped_vendors or record.reporting_key in mapped_vendors or record.categories_key in mapped_vendors


def measure_penalty(sales_measures: dict[str, set[str]], master_measures: dict[str, set[str]]) -> Decimal:
    penalty = Decimal("0")
    for unit in {"GAL", "OZ", "LB", "KIT"}:
        sales_values = sales_measures.get(unit, set())
        master_values = master_measures.get(unit, set())
        if sales_values and master_values and sales_values.isdisjoint(master_values):
            penalty += Decimal("0.12")
    sales_special = sales_measures.get("SPECIAL", set())
    master_special = master_measures.get("SPECIAL", set())
    if {"HALF", "FULL"} <= (sales_special | master_special) and sales_special != master_special:
        penalty += Decimal("0.08")
    return penalty


def score_candidate(
    sales_name: str,
    sales_tokens: set[str],
    sales_numbers: set[str],
    sales_codes: set[str],
    sales_category: str,
    sales_measures: dict[str, set[str]],
    record: MasterRecord,
) -> tuple[Decimal, str]:
    sequence_score = Decimal(str(SequenceMatcher(None, sales_name, record.normalized_name).ratio()))
    overlap_count = len(sales_tokens & record.tokens)
    overlap_score = Decimal(str(overlap_count / max(len(sales_tokens), len(record.tokens), 1)))
    if sales_numbers and record.number_tokens:
        numeric_score = Decimal(str(len(sales_numbers & record.number_tokens) / max(len(sales_numbers), len(record.number_tokens), 1)))
    else:
        numeric_score = Decimal("0.5") if not sales_numbers and not record.number_tokens else Decimal("0")
    code_score = Decimal("1") if sales_codes & {record.sku_code, record.vendor_code} else Decimal("0")
    contains_score = Decimal("1") if sales_name and (sales_name in record.normalized_name or record.normalized_name in sales_name) else Decimal("0")
    category_score = Decimal("1") if category_fit(sales_category, record) else Decimal("0")
    penalty = measure_penalty(sales_measures, record.measures)

    score = (
        (sequence_score * Decimal("0.46"))
        + (overlap_score * Decimal("0.22"))
        + (numeric_score * Decimal("0.14"))
        + (code_score * Decimal("0.10"))
        + (contains_score * Decimal("0.04"))
        + (category_score * Decimal("0.04"))
        - penalty
    )
    score = max(Decimal("0"), min(Decimal("1"), score))

    reason_parts: list[str] = []
    if code_score:
        reason_parts.append("shared code")
    if overlap_count:
        reason_parts.append(f"{overlap_count} shared tokens")
    if category_score:
        reason_parts.append("category fit")
    if contains_score:
        reason_parts.append("name contains match")
    if penalty:
        reason_parts.append("measure mismatch penalty")

    return score, ", ".join(reason_parts)


def candidate_indexes_for_sales_item(
    sales_tokens: set[str],
    sales_codes: set[str],
    category_indexes: set[int],
    sku_index: dict[str, list[int]],
    token_index: dict[str, set[int]],
) -> set[int]:
    candidates: set[int] = set()
    for code in sales_codes:
        candidates.update(index for index in sku_index.get(code, []) if index in category_indexes)

    token_candidates: list[set[int]] = []
    sorted_tokens = sorted(sales_tokens, key=lambda token: (-len(token), token))
    for token in sorted_tokens[:6]:
        token_hits = token_index.get(token)
        if token_hits:
            token_candidates.append(token_hits & category_indexes)

    if token_candidates:
        strong_hits = [hits for hits in token_candidates if hits]
        if strong_hits:
            shared = strong_hits[0].copy()
            for hits in strong_hits[1:3]:
                shared &= hits
            if shared:
                candidates.update(shared)
            else:
                for hits in strong_hits[:4]:
                    candidates.update(hits)

    if not candidates:
        candidates = set(category_indexes)

    return candidates


def choose_match_type(top_score: Decimal, second_score: Decimal, reason: str) -> str:
    gap = top_score - second_score
    if top_score >= Decimal("0.92") and gap >= Decimal("0.06"):
        return "high_confidence_fuzzy"
    if top_score >= Decimal("0.86") and gap >= Decimal("0.10") and "measure mismatch penalty" not in reason:
        return "high_confidence_fuzzy"
    if top_score >= Decimal("0.78"):
        return "review_fuzzy"
    return "unmatched"


def write_xlsx(path: Path, rows: list[dict[str, str]], fieldnames: list[str], sheet_name: str) -> bool:
    if Workbook is None or get_column_letter is None:
        return False

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(fieldnames)
    if Font is not None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}1"
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    for column_index, fieldname in enumerate(fieldnames, start=1):
        width = min(max(len(fieldname) + 2, 14), 42)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    workbook.save(path)
    return True


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SALES_DIR.mkdir(parents=True, exist_ok=True)
    SALES_OVERRIDE_DIR.mkdir(parents=True, exist_ok=True)

    issues: list[dict[str, str]] = []
    master_path = resolve_master_path()
    sales_path = resolve_sales_path()

    master_rows = read_csv_rows(master_path)
    sales_rows = read_csv_rows(sales_path)
    if not master_rows:
        raise ValueError("Master inventory is empty.")
    if not sales_rows:
        raise ValueError("Sales report is empty.")

    records, sku_index, gtin_index, name_index, base_name_index, base_name_code_index, token_index = build_master_records(master_rows)
    sales_aggregates = build_sales_aggregates(sales_rows)
    sku_overrides, item_overrides = load_match_overrides(master_rows, issues)

    audit_rows: list[dict[str, str]] = []
    review_rows: list[dict[str, str]] = []
    accepted_signals: dict[str, dict[str, object]] = {}

    total_sales_lines = 0
    total_net_sales = Decimal("0")
    matched_sales_lines = 0
    matched_net_sales = Decimal("0")

    for aggregate in sorted(sales_aggregates.values(), key=lambda entry: (entry.net_sales, entry.item), reverse=True):
        total_sales_lines += aggregate.lines
        total_net_sales += aggregate.net_sales

        sales_category_key = normalize_text(aggregate.category)
        sales_name = normalize_text(aggregate.item)
        sales_tokens = significant_tokens(sales_name)
        sales_numbers = extract_number_tokens(sales_tokens)
        raw_sales_measures = extract_measures(sales_name)
        sales_measures, sales_base_name = infer_sales_measures(aggregate.category, sales_name, raw_sales_measures)
        sales_base_name_code = canonical_code(sales_base_name)
        sales_codes = code_tokens_from_sales_item(aggregate.item, aggregate.sku)
        sales_gtin = canonical_code(aggregate.gtin)
        override = sku_overrides.get(canonical_code(aggregate.sku)) or item_overrides.get(
            (sales_category_key, sales_name, canonical_code(aggregate.sku))
        )

        matched_record: MasterRecord | None = None
        match_type = "unmatched"
        match_score = Decimal("0")
        match_reason = ""
        second_score = Decimal("0")
        candidate_details: list[tuple[Decimal, MasterRecord, str]] = []

        if is_non_inventory_sales_item(aggregate.category, aggregate.item):
            match_type = "non_inventory"
            match_reason = "Service or shipping row excluded from inventory matching."
        elif override:
            matched_record = next((record for record in records if record.row.get("SKU", "").strip() == override.master_sku), None)
            if matched_record is not None:
                match_type = "manual_override"
                match_score = Decimal("1")
                match_reason = f"Manual override. {override.notes}".strip()

        if match_type != "non_inventory" and matched_record is None and aggregate.sku:
            exact_sku_candidates = sorted(set(sku_index.get(canonical_code(aggregate.sku), [])))
            if len(exact_sku_candidates) == 1:
                matched_record = records[exact_sku_candidates[0]]
                match_type = "exact_sku"
                match_score = Decimal("1")
                match_reason = "Sales SKU exactly matches catalog SKU or vendor code."

        if match_type != "non_inventory" and matched_record is None and sales_gtin:
            exact_gtin_candidates = sorted(set(gtin_index.get(sales_gtin, [])))
            if len(exact_gtin_candidates) == 1:
                matched_record = records[exact_gtin_candidates[0]]
                match_type = "exact_gtin"
                match_score = Decimal("1")
                match_reason = "Sales GTIN exactly matches catalog GTIN."

        if match_type != "non_inventory" and matched_record is None:
            exact_name_candidates = sorted(
                set(index for index in name_index.get(sales_name, []) if index in sales_category_candidates(aggregate.category, records))
            )
            if len(exact_name_candidates) == 1:
                matched_record = records[exact_name_candidates[0]]
                match_type = "exact_name"
                match_score = Decimal("1")
                match_reason = "Normalized sales item name exactly matches the catalog name."

        if match_type != "non_inventory" and matched_record is None and sales_base_name:
            category_indexes = sales_category_candidates(aggregate.category, records)
            base_name_candidates = sorted(set(index for index in base_name_index.get(sales_base_name, []) if index in category_indexes))
            if len(base_name_candidates) == 1:
                matched_record = records[base_name_candidates[0]]
                match_type = "exact_base_name"
                match_score = Decimal("1")
                match_reason = "Base product name matches a single catalog item after stripping pack descriptors."
            elif len(base_name_candidates) > 1:
                base_name_pack_candidates = [
                    index for index in base_name_candidates if measure_penalty(sales_measures, records[index].measures) == Decimal("0")
                ]
                if len(base_name_pack_candidates) == 1:
                    matched_record = records[base_name_pack_candidates[0]]
                    match_type = "exact_base_name_pack"
                    match_score = Decimal("1")
                    match_reason = "Base product name matched and pack size isolated a single catalog item."

        if match_type != "non_inventory" and matched_record is None and sales_base_name_code:
            category_indexes = sales_category_candidates(aggregate.category, records)
            base_name_code_candidates = sorted(
                set(index for index in base_name_code_index.get(sales_base_name_code, []) if index in category_indexes)
            )
            if len(base_name_code_candidates) == 1:
                matched_record = records[base_name_code_candidates[0]]
                match_type = "exact_base_code"
                match_score = Decimal("1")
                match_reason = "Compact base product code matches a single catalog item."
            elif len(base_name_code_candidates) > 1:
                base_name_code_pack_candidates = [
                    index for index in base_name_code_candidates if measure_penalty(sales_measures, records[index].measures) == Decimal("0")
                ]
                if len(base_name_code_pack_candidates) == 1:
                    matched_record = records[base_name_code_pack_candidates[0]]
                    match_type = "exact_base_code_pack"
                    match_score = Decimal("1")
                    match_reason = "Compact base product code matched and pack size isolated a single catalog item."

        if match_type != "non_inventory" and matched_record is None and sales_codes:
            category_indexes = sales_category_candidates(aggregate.category, records)
            exact_code_candidates = sorted(
                {
                    index
                    for code in sales_codes
                    for index in sku_index.get(code, [])
                    if index in category_indexes
                }
            )
            if len(exact_code_candidates) == 1:
                matched_record = records[exact_code_candidates[0]]
                match_type = "exact_code"
                match_score = Decimal("1")
                match_reason = "Sales item exposes a unique catalog code or model number."

        if match_type != "non_inventory" and matched_record is None:
            category_indexes = sales_category_candidates(aggregate.category, records)
            candidate_indexes = candidate_indexes_for_sales_item(sales_tokens, sales_codes, category_indexes, sku_index, token_index)
            for candidate_index in candidate_indexes:
                record = records[candidate_index]
                score, reason = score_candidate(sales_name, sales_tokens, sales_numbers, sales_codes, aggregate.category, sales_measures, record)
                candidate_details.append((score, record, reason))
            candidate_details.sort(key=lambda item: (item[0], item[1].display_name), reverse=True)
            if candidate_details:
                top_score, top_record, top_reason = candidate_details[0]
                second_score = candidate_details[1][0] if len(candidate_details) > 1 else Decimal("0")
                match_type = choose_match_type(top_score, second_score, top_reason)
                match_score = top_score
                match_reason = top_reason
                if match_type in {"high_confidence_fuzzy", "review_fuzzy"}:
                    matched_record = top_record

        accepted = matched_record is not None and match_type in {
            "manual_override",
            "exact_sku",
            "exact_gtin",
            "exact_name",
            "exact_base_name",
            "exact_base_name_pack",
            "exact_base_code",
            "exact_base_code_pack",
            "exact_code",
            "high_confidence_fuzzy",
        }
        if accepted:
            matched_sales_lines += aggregate.lines
            matched_net_sales += aggregate.net_sales

            signal = accepted_signals.get(matched_record.row.get("SKU", "").strip())
            if signal is None:
                signal = {
                    "Matched Master SKU": matched_record.row.get("SKU", "").strip(),
                    "Matched Item Name": matched_record.display_name,
                    "Default Vendor Name": matched_record.row.get("Default Vendor Name", ""),
                    "Reporting Category": matched_record.row.get("Reporting Category", ""),
                    "Catalog Price": matched_record.row.get("Price", ""),
                    "Default Unit Cost": matched_record.row.get("Default Unit Cost", ""),
                    "Sales Lines": 0,
                    "Quantity Sold": Decimal("0"),
                    "Gross Sales": Decimal("0"),
                    "Discounts": Decimal("0"),
                    "Net Sales": Decimal("0"),
                    "First Sale Date": "",
                    "Last Sale Date": "",
                    "Distinct Sales Items": 0,
                    "Sales Categories Observed": set(),
                    "Match Types": set(),
                }
                accepted_signals[matched_record.row.get("SKU", "").strip()] = signal

            signal["Sales Lines"] = int(signal["Sales Lines"]) + aggregate.lines
            signal["Quantity Sold"] = Decimal(signal["Quantity Sold"]) + aggregate.quantity
            signal["Gross Sales"] = Decimal(signal["Gross Sales"]) + aggregate.gross_sales
            signal["Discounts"] = Decimal(signal["Discounts"]) + aggregate.discounts
            signal["Net Sales"] = Decimal(signal["Net Sales"]) + aggregate.net_sales
            signal["Distinct Sales Items"] = int(signal["Distinct Sales Items"]) + 1
            signal["Sales Categories Observed"].add(aggregate.category)
            signal["Match Types"].add(match_type)
            if aggregate.first_sale_date and (not signal["First Sale Date"] or aggregate.first_sale_date < signal["First Sale Date"]):
                signal["First Sale Date"] = aggregate.first_sale_date
            if aggregate.last_sale_date and (not signal["Last Sale Date"] or aggregate.last_sale_date > signal["Last Sale Date"]):
                signal["Last Sale Date"] = aggregate.last_sale_date

        top_candidates = candidate_details[:3]

        def candidate_value(position: int, value_index: int) -> str:
            if len(top_candidates) <= position:
                return ""
            score, record, reason = top_candidates[position]
            if value_index == 0:
                return f"{score.quantize(Decimal('0.001'))}"
            if value_index == 1:
                return record.display_name
            if value_index == 2:
                return record.row.get("SKU", "")
            if value_index == 3:
                return record.row.get("Default Vendor Name", "")
            return reason

        matched_sku = matched_record.row.get("SKU", "").strip() if matched_record is not None else ""
        matched_name = matched_record.display_name if matched_record is not None else ""
        matched_vendor = matched_record.row.get("Default Vendor Name", "") if matched_record is not None else ""

        audit_row = {
            "Sales Category": aggregate.category,
            "Sales Item": aggregate.item,
            "Sales SKU": aggregate.sku,
            "Sales GTIN": aggregate.gtin,
            "Sales Lines": str(aggregate.lines),
            "Quantity Sold": format_decimal(aggregate.quantity),
            "Gross Sales": format_money(aggregate.gross_sales),
            "Discounts": format_money(aggregate.discounts),
            "Net Sales": format_money(aggregate.net_sales),
            "First Sale Date": aggregate.first_sale_date,
            "Last Sale Date": aggregate.last_sale_date,
            "Match Type": match_type,
            "Match Accepted": "Y" if accepted else "N",
            "Match Score": f"{match_score.quantize(Decimal('0.001'))}",
            "Second Best Score": f"{second_score.quantize(Decimal('0.001'))}",
            "Matched Master SKU": matched_sku,
            "Matched Catalog Item": matched_name,
            "Matched Vendor": matched_vendor,
            "Matched Catalog Price": matched_record.row.get("Price", "") if matched_record is not None else "",
            "Matched Catalog Cost": matched_record.row.get("Default Unit Cost", "") if matched_record is not None else "",
            "Match Reason": match_reason,
            "Top Candidate 1 Score": candidate_value(0, 0),
            "Top Candidate 1 Item": candidate_value(0, 1),
            "Top Candidate 1 SKU": candidate_value(0, 2),
            "Top Candidate 1 Vendor": candidate_value(0, 3),
            "Top Candidate 1 Reason": candidate_value(0, 4),
            "Top Candidate 2 Score": candidate_value(1, 0),
            "Top Candidate 2 Item": candidate_value(1, 1),
            "Top Candidate 2 SKU": candidate_value(1, 2),
            "Top Candidate 2 Vendor": candidate_value(1, 3),
            "Top Candidate 2 Reason": candidate_value(1, 4),
            "Top Candidate 3 Score": candidate_value(2, 0),
            "Top Candidate 3 Item": candidate_value(2, 1),
            "Top Candidate 3 SKU": candidate_value(2, 2),
            "Top Candidate 3 Vendor": candidate_value(2, 3),
            "Top Candidate 3 Reason": candidate_value(2, 4),
        }
        audit_rows.append(audit_row)

        if not accepted and match_type != "non_inventory":
            review_rows.append(audit_row)

    signal_rows: list[dict[str, str]] = []
    for signal in sorted(accepted_signals.values(), key=lambda item: (Decimal(item["Net Sales"]), item["Matched Master SKU"]), reverse=True):
        quantity_sold = Decimal(signal["Quantity Sold"])
        net_sales = Decimal(signal["Net Sales"])
        average_realized_price = (net_sales / quantity_sold).quantize(Decimal("0.01")) if quantity_sold else Decimal("0")
        signal_rows.append(
            {
                "Matched Master SKU": str(signal["Matched Master SKU"]),
                "Matched Item Name": str(signal["Matched Item Name"]),
                "Default Vendor Name": str(signal["Default Vendor Name"]),
                "Reporting Category": str(signal["Reporting Category"]),
                "Catalog Price": str(signal["Catalog Price"]),
                "Default Unit Cost": str(signal["Default Unit Cost"]),
                "Sales Lines": str(signal["Sales Lines"]),
                "Distinct Sales Items": str(signal["Distinct Sales Items"]),
                "Quantity Sold": format_decimal(quantity_sold),
                "Gross Sales": format_money(Decimal(signal["Gross Sales"])),
                "Discounts": format_money(Decimal(signal["Discounts"])),
                "Net Sales": format_money(net_sales),
                "Average Realized Unit Price": format_money(average_realized_price),
                "First Sale Date": str(signal["First Sale Date"]),
                "Last Sale Date": str(signal["Last Sale Date"]),
                "Sales Categories Observed": ", ".join(sorted(signal["Sales Categories Observed"])),
                "Match Types": ", ".join(sorted(signal["Match Types"])),
            }
        )

    audit_fieldnames = list(audit_rows[0].keys()) if audit_rows else []
    signal_fieldnames = list(signal_rows[0].keys()) if signal_rows else [
        "Matched Master SKU",
        "Matched Item Name",
        "Default Vendor Name",
        "Reporting Category",
        "Catalog Price",
        "Default Unit Cost",
        "Sales Lines",
        "Distinct Sales Items",
        "Quantity Sold",
        "Gross Sales",
        "Discounts",
        "Net Sales",
        "Average Realized Unit Price",
        "First Sale Date",
        "Last Sale Date",
        "Sales Categories Observed",
        "Match Types",
    ]

    write_csv(SALES_MATCH_AUDIT_PATH, audit_fieldnames, audit_rows)
    write_csv(SALES_MATCH_REVIEW_PATH, audit_fieldnames, review_rows)
    write_csv(SALES_CATALOG_SIGNALS_PATH, signal_fieldnames, signal_rows)
    write_csv(SALES_MATCH_ISSUES_PATH, ["source_file", "row_number", "issue_type", "sales_sku", "details"], issues)
    xlsx_written = write_xlsx(SALES_MATCH_AUDIT_XLSX_PATH, audit_rows, audit_fieldnames, "Sales Match Audit")

    summary_lines = [
        f"Master inventory source: {master_path}",
        f"Sales source: {sales_path}",
        f"Sales match audit CSV: {SALES_MATCH_AUDIT_PATH}",
        f"Sales match audit Excel: {SALES_MATCH_AUDIT_XLSX_PATH if xlsx_written else '[openpyxl not available]'}",
        f"Sales review file: {SALES_MATCH_REVIEW_PATH}",
        f"Sales catalog signals file: {SALES_CATALOG_SIGNALS_PATH}",
        f"Issues file: {SALES_MATCH_ISSUES_PATH}",
        f"Unique sales items analyzed: {len(audit_rows)}",
        f"Total sales lines analyzed: {total_sales_lines}",
        f"Accepted matches: {sum(1 for row in audit_rows if row['Match Accepted'] == 'Y')}",
        f"Review rows: {len(review_rows)}",
        f"Matched sales lines: {matched_sales_lines}",
        f"Matched net sales: {format_money(matched_net_sales)}",
        f"Total net sales: {format_money(total_net_sales)}",
        f"Matched sales line coverage: {format_decimal((Decimal(matched_sales_lines) / Decimal(total_sales_lines) * Decimal('100')) if total_sales_lines else Decimal('0'))}%",
        f"Matched net sales coverage: {format_decimal((matched_net_sales / total_net_sales * Decimal('100')) if total_net_sales else Decimal('0'))}%",
        "Accepted match types: manual_override, exact_sku, exact_gtin, exact_name, exact_base_name, exact_base_name_pack, exact_base_code, exact_base_code_pack, exact_code, high_confidence_fuzzy.",
        "Review rows include weaker fuzzy candidates and truly unmatched items so they can be pinned with manual overrides later.",
    ]
    SALES_MATCH_SUMMARY_PATH.write_text("\n".join(summary_lines), encoding="utf-8")
    print("\n".join(summary_lines))


if __name__ == "__main__":
    main()
