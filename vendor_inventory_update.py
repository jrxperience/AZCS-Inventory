from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


MASTER_PATH = Path(r"C:\Users\JRAZC\Downloads\inventory.csv")
BARRENS_PATH = Path(r"Q:\Shared drives\AZCS\Price List\Barrens Pricelist 2025.csv")
EACO_PATH = Path(r"Q:\Shared drives\AZCS\Price List\EacoChem Full Price List.xlsx")
JR_PATH = Path(r"Q:\Shared drives\AZCS\Price List\JRacenstein Pricelist 2025.xlsx")
MPWSR_PATH = Path(r"Q:\Shared drives\AZCS\Price List\MPWSR Price List May '25.csv")

OUT_PATH = Path(r"C:\Users\JRAZC\Downloads\inventory_updated_costs_and_prices_25_markup.csv")
FLAGS_PATH = Path(r"C:\Users\JRAZC\Downloads\inventory_vendor_conflicts_25_markup.csv")
SUMMARY_PATH = Path(r"C:\Users\JRAZC\Downloads\inventory_update_summary_25_markup.txt")

MARKUP = Decimal("1.25")
CENT = Decimal("0.01")


@dataclass(frozen=True)
class Candidate:
    vendor: str
    match_type: str
    source_key: str
    source_desc: str
    price: Decimal
    source_file: str


def norm_sku(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().upper())


def norm_name(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def parse_money(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("$", "").replace(",", "").strip()
    try:
        return Decimal(text).quantize(CENT, rounding=ROUND_HALF_UP)
    except Exception:
        return None


def format_money(value: Decimal) -> str:
    return str(value.quantize(CENT, rounding=ROUND_HALF_UP))


def markup_price(cost: Decimal) -> Decimal:
    return (cost * MARKUP).quantize(CENT, rounding=ROUND_HALF_UP)


def read_csv_any(path: Path) -> list[list[str]]:
    last_error = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error  # type: ignore[misc]


def add_candidate(
    candidates: dict[int, list[Candidate]],
    row_index: int,
    candidate: Candidate,
) -> None:
    if candidate not in candidates[row_index]:
        candidates[row_index].append(candidate)


def build_master_indexes(rows: list[dict[str, str]]) -> tuple[dict[str, list[int]], dict[str, list[int]]]:
    sku_index: dict[str, list[int]] = defaultdict(list)
    name_index: dict[str, list[int]] = defaultdict(list)
    for idx, row in enumerate(rows):
        sku_key = norm_sku(row.get("SKU"))
        if sku_key:
            sku_index[sku_key].append(idx)
        name_key = norm_name(row.get("Item Name"))
        if name_key:
            name_index[name_key].append(idx)
    return sku_index, name_index


def collect_barrens_candidates(
    sku_index: dict[str, list[int]],
    candidates: dict[int, list[Candidate]],
    flags: list[dict[str, str]],
    counts: Counter[str],
    master_rows: list[dict[str, str]],
) -> None:
    rows = read_csv_any(BARRENS_PATH)
    header_row_index = next(i for i, row in enumerate(rows) if row and str(row[0]).strip() == "StockCode")
    for row in rows[header_row_index + 1:]:
        if not row:
            continue
        sku_key = norm_sku(row[0] if len(row) > 0 else "")
        price = parse_money(row[2] if len(row) > 2 else None)
        if not sku_key or price is None:
            continue

        master_matches = sku_index.get(sku_key, [])
        if len(master_matches) == 1:
            add_candidate(
                candidates,
                master_matches[0],
                Candidate(
                    vendor="Barrens",
                    match_type="sku",
                    source_key=sku_key,
                    source_desc=row[1] if len(row) > 1 else "",
                    price=price,
                    source_file=BARRENS_PATH.name,
                ),
            )
            counts["Barrens"] += 1
        elif len(master_matches) > 1:
            for idx in master_matches:
                flags.append(
                    {
                        "issue_type": "duplicate_master_sku",
                        "vendor": "Barrens",
                        "item_name": master_rows[idx].get("Item Name", ""),
                        "sku": master_rows[idx].get("SKU", ""),
                        "current_default_unit_cost": master_rows[idx].get("Default Unit Cost", ""),
                        "current_price": master_rows[idx].get("Price", ""),
                        "vendor_prices": format_money(price),
                        "details": f"Matched Barrens SKU {sku_key}, but this SKU appears multiple times in the master inventory.",
                    }
                )


def collect_mpwsr_candidates(
    sku_index: dict[str, list[int]],
    candidates: dict[int, list[Candidate]],
    flags: list[dict[str, str]],
    counts: Counter[str],
    master_rows: list[dict[str, str]],
) -> None:
    rows = read_csv_any(MPWSR_PATH)
    headers = rows[0]
    name_idx = headers.index("Name")
    desc_idx = headers.index("Description")
    dealer_idx = headers.index("Dealers")

    for row in rows[1:]:
        if len(row) <= max(name_idx, desc_idx, dealer_idx):
            continue
        sku_key = norm_sku(row[name_idx])
        price = parse_money(row[dealer_idx])
        if not sku_key or price is None:
            continue

        master_matches = sku_index.get(sku_key, [])
        if len(master_matches) == 1:
            add_candidate(
                candidates,
                master_matches[0],
                Candidate(
                    vendor="MPWSR",
                    match_type="sku",
                    source_key=sku_key,
                    source_desc=row[desc_idx],
                    price=price,
                    source_file=MPWSR_PATH.name,
                ),
            )
            counts["MPWSR"] += 1
        elif len(master_matches) > 1:
            for idx in master_matches:
                flags.append(
                    {
                        "issue_type": "duplicate_master_sku",
                        "vendor": "MPWSR",
                        "item_name": master_rows[idx].get("Item Name", ""),
                        "sku": master_rows[idx].get("SKU", ""),
                        "current_default_unit_cost": master_rows[idx].get("Default Unit Cost", ""),
                        "current_price": master_rows[idx].get("Price", ""),
                        "vendor_prices": format_money(price),
                        "details": f"Matched MPWSR SKU {sku_key}, but this SKU appears multiple times in the master inventory.",
                    }
                )


def collect_jr_candidates(
    sku_index: dict[str, list[int]],
    name_index: dict[str, list[int]],
    candidates: dict[int, list[Candidate]],
    flags: list[dict[str, str]],
    counts: Counter[str],
    master_rows: list[dict[str, str]],
) -> None:
    workbook = load_workbook(JR_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])
    code_idx = headers.index("Code")
    desc_idx = headers.index("Description")
    price_idx = headers.index("Price")

    for row in rows[1:]:
        if not row:
            continue

        code = row[code_idx] if len(row) > code_idx else None
        desc = row[desc_idx] if len(row) > desc_idx else None
        price = parse_money(row[price_idx] if len(row) > price_idx else None)
        if price is None:
            continue

        sku_key = norm_sku(code)
        if sku_key:
            master_matches = sku_index.get(sku_key, [])
            if len(master_matches) == 1:
                add_candidate(
                    candidates,
                    master_matches[0],
                    Candidate(
                        vendor="JRacenstein",
                        match_type="sku",
                        source_key=sku_key,
                        source_desc=str(desc or ""),
                        price=price,
                        source_file=JR_PATH.name,
                    ),
                )
                counts["JRacenstein"] += 1
            elif len(master_matches) > 1:
                for idx in master_matches:
                    flags.append(
                        {
                            "issue_type": "duplicate_master_sku",
                            "vendor": "JRacenstein",
                            "item_name": master_rows[idx].get("Item Name", ""),
                            "sku": master_rows[idx].get("SKU", ""),
                            "current_default_unit_cost": master_rows[idx].get("Default Unit Cost", ""),
                            "current_price": master_rows[idx].get("Price", ""),
                            "vendor_prices": format_money(price),
                            "details": f"Matched JRacenstein code {sku_key}, but this SKU appears multiple times in the master inventory.",
                        }
                    )

        name_key = norm_name(desc)
        if not name_key:
            continue

        vendor_name_matches = [
            idx
            for idx in name_index.get(name_key, [])
            if "RACENSTEIN" in (master_rows[idx].get("Default Vendor Name", "").upper())
        ]
        if len(vendor_name_matches) == 1:
            add_candidate(
                candidates,
                vendor_name_matches[0],
                Candidate(
                    vendor="JRacenstein",
                    match_type="exact_name",
                    source_key=name_key,
                    source_desc=str(desc or ""),
                    price=price,
                    source_file=JR_PATH.name,
                ),
            )
            counts["JRacenstein"] += 1

    workbook.close()


def split_eaco_item(item_name: str) -> tuple[str, str]:
    text = str(item_name or "").upper().strip()
    size = "1G"
    patterns = [
        (r"\b55\s*(?:G|GAL|GALLON|GALS|GALLONS)?\.?$", "55G"),
        (r"\b5\s*(?:G|GAL|GALLON|GALS|GALLONS)?\.?$", "5G"),
        (r"\b1\s*(?:G|GAL|GALLON|GALS|GALLONS)?\.?$", "1G"),
    ]
    base = text
    for pattern, size_name in patterns:
        if re.search(pattern, text):
            size = size_name
            base = re.sub(pattern, "", text).strip(" -.")
            break
    return norm_name(base), size


def collect_eaco_candidates(
    candidates: dict[int, list[Candidate]],
    flags: list[dict[str, str]],
    counts: Counter[str],
    master_rows: list[dict[str, str]],
) -> None:
    workbook = load_workbook(EACO_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])
    column_index = {name: headers.index(name) for name in headers}

    size_columns_b = {"55G": "55G_B", "5G": "5G_B", "1G": "1G_B"}
    size_columns_a = {"55G": "55G_A", "5G": "5G_A", "1G": "1G_A"}
    catalog_b: dict[str, dict[str, set[Decimal]]] = defaultdict(lambda: defaultdict(set))
    catalog_a: dict[str, dict[str, set[Decimal]]] = defaultdict(lambda: defaultdict(set))

    for row in rows[1:]:
        if not row or len(row) <= column_index["Product"]:
            continue
        product_key = norm_name(row[column_index["Product"]])
        if not product_key:
            continue

        for size_name, column_name in size_columns_b.items():
            price = parse_money(row[column_index[column_name]] if len(row) > column_index[column_name] else None)
            if price is not None:
                catalog_b[product_key][size_name].add(price)

        for size_name, column_name in size_columns_a.items():
            price = parse_money(row[column_index[column_name]] if len(row) > column_index[column_name] else None)
            if price is not None:
                catalog_a[product_key][size_name].add(price)

    workbook.close()

    for idx, row in enumerate(master_rows):
        vendor_name = (row.get("Default Vendor Name") or "").upper()
        if "EACO" not in vendor_name:
            continue

        base_key, size = split_eaco_item(row.get("Item Name", ""))
        if not base_key:
            continue

        prices = set(catalog_b.get(base_key, {}).get(size, set()))
        if not prices:
            prices = set(catalog_a.get(base_key, {}).get(size, set()))

        if len(prices) == 1:
            price = next(iter(prices))
            add_candidate(
                candidates,
                idx,
                Candidate(
                    vendor="EacoChem",
                    match_type=f"product_name_{size}",
                    source_key=base_key,
                    source_desc=row.get("Item Name", ""),
                    price=price,
                    source_file=EACO_PATH.name,
                ),
            )
            counts["EacoChem"] += 1
        elif len(prices) > 1:
            flags.append(
                {
                    "issue_type": "multiple_vendor_prices",
                    "vendor": "EacoChem",
                    "item_name": row.get("Item Name", ""),
                    "sku": row.get("SKU", ""),
                    "current_default_unit_cost": row.get("Default Unit Cost", ""),
                    "current_price": row.get("Price", ""),
                    "vendor_prices": "; ".join(format_money(price) for price in sorted(prices)),
                    "details": f"EacoChem returned multiple prices for base product {base_key} size {size}.",
                }
            )


def dedupe_flags(flags: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, ...]] = set()
    result: list[dict[str, str]] = []
    for flag in flags:
        key = (
            flag["issue_type"],
            flag["vendor"],
            flag["item_name"],
            flag["sku"],
            flag["vendor_prices"],
            flag["details"],
        )
        if key not in seen:
            seen.add(key)
            result.append(flag)
    return result


def main() -> None:
    with MASTER_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        master_rows = list(reader)

    sku_index, name_index = build_master_indexes(master_rows)

    candidates: dict[int, list[Candidate]] = defaultdict(list)
    flags: list[dict[str, str]] = []
    counts: Counter[str] = Counter()

    collect_barrens_candidates(sku_index, candidates, flags, counts, master_rows)
    collect_mpwsr_candidates(sku_index, candidates, flags, counts, master_rows)
    collect_jr_candidates(sku_index, name_index, candidates, flags, counts, master_rows)
    collect_eaco_candidates(candidates, flags, counts, master_rows)

    matched_rows = 0
    conflict_rows = 0
    cost_updates = 0
    cost_already_current = 0
    price_updates = 0
    price_already_current = 0
    both_updated = 0

    for idx, row in enumerate(master_rows):
        row_candidates = candidates.get(idx, [])
        if not row_candidates:
            continue

        matched_rows += 1
        unique_costs = sorted({candidate.price for candidate in row_candidates})
        if len(unique_costs) > 1:
            conflict_rows += 1
            flags.append(
                {
                    "issue_type": "multiple_vendor_prices",
                    "vendor": "; ".join(sorted({candidate.vendor for candidate in row_candidates})),
                    "item_name": row.get("Item Name", ""),
                    "sku": row.get("SKU", ""),
                    "current_default_unit_cost": row.get("Default Unit Cost", ""),
                    "current_price": row.get("Price", ""),
                    "vendor_prices": "; ".join(
                        f"{candidate.vendor}={format_money(candidate.price)}"
                        for candidate in sorted(row_candidates, key=lambda candidate: (candidate.vendor, candidate.price))
                    ),
                    "details": "Multiple distinct vendor prices matched this inventory item, so no cost or selling price was auto-updated.",
                }
            )
            continue

        vendor_cost = unique_costs[0]
        target_price = markup_price(vendor_cost)

        current_cost = parse_money(row.get("Default Unit Cost"))
        current_price = parse_money(row.get("Price"))

        cost_changed = current_cost != vendor_cost
        price_changed = current_price != target_price

        if cost_changed:
            row["Default Unit Cost"] = format_money(vendor_cost)
            cost_updates += 1
        else:
            cost_already_current += 1

        if price_changed:
            row["Price"] = format_money(target_price)
            price_updates += 1
        else:
            price_already_current += 1

        if cost_changed and price_changed:
            both_updated += 1

    flags = dedupe_flags(flags)

    with OUT_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(master_rows)

    flag_fieldnames = [
        "issue_type",
        "vendor",
        "item_name",
        "sku",
        "current_default_unit_cost",
        "current_price",
        "vendor_prices",
        "details",
    ]
    with FLAGS_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=flag_fieldnames)
        writer.writeheader()
        writer.writerows(flags)

    summary_lines = [
        f"Updated inventory written to: {OUT_PATH}",
        f"Conflict report written to: {FLAGS_PATH}",
        f"Master rows: {len(master_rows)}",
        f"Rows with at least one high-confidence vendor match: {matched_rows}",
        f"Rows flagged for conflicting vendor prices: {conflict_rows}",
        f"Default Unit Cost updated: {cost_updates}",
        f"Default Unit Cost already current: {cost_already_current}",
        f"Price updated to 25% markup: {price_updates}",
        f"Price already at 25% markup: {price_already_current}",
        f"Rows where both cost and price changed: {both_updated}",
        "Candidate matches by vendor:",
    ]
    for vendor in sorted(counts):
        summary_lines.append(f"  {vendor}: {counts[vendor]}")

    SUMMARY_PATH.write_text("\n".join(summary_lines), encoding="utf-8")
    print("\n".join(summary_lines))


if __name__ == "__main__":
    main()
