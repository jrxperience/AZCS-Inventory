from __future__ import annotations

import csv
import html
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field, replace
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlparse

import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
PRICE_LIST_DIR = INPUT_DIR / "price_lists"
TEMPLATE_DIR = BASE_DIR / "templates"
OUTPUT_DIR = BASE_DIR / "outputs"

TEMPLATE_PATH = TEMPLATE_DIR / "Square Import Template.csv"
LEGACY_TEMPLATE_PATH = BASE_DIR / "Square Import Template.csv"
VERIFIED_ENRICHMENT_PATH = INPUT_DIR / "verified_product_enrichment.csv"

MASTER_OUT_PATH = OUTPUT_DIR / "square_master_inventory.csv"
REVIEW_OUT_PATH = OUTPUT_DIR / "square_master_inventory_overlap_review.csv"
SUMMARY_OUT_PATH = OUTPUT_DIR / "square_master_inventory_summary.txt"
ENRICHMENT_AUDIT_OUT_PATH = OUTPUT_DIR / "product_enrichment_audit.csv"
IMAGE_DATABASE_OUT_PATH = OUTPUT_DIR / "inventory_database_with_images.csv"
IMAGE_DATABASE_XLSX_OUT_PATH = OUTPUT_DIR / "inventory_database_with_images.xlsx"
IMAGE_AUDIT_OUT_PATH = OUTPUT_DIR / "product_image_match_audit.csv"
IMAGE_DIR = BASE_DIR / "Images"
JRACENSTEIN_IMAGE_DIR = IMAGE_DIR / "J.racenstein images"
MANATEE_IMAGE_DIR = IMAGE_DIR / "Manatee images"
TRIDENT_IMAGE_DIR = IMAGE_DIR / "Trident images" / "media"

CENT = Decimal("0.01")
MONEY_RE = re.compile(r"\$?\s*\d[\d,\s]*\.\d{2}")
GTIN_LENGTHS = {8, 12, 13, 14}
GENERIC_CATEGORY_SEGMENTS = {"EQUIPMENT"}
GENERIC_SEO_WORDS = {"AND", "THE", "FOR", "WITH", "KIT", "PACK", "CASE", "REGULAR"}
WEIGHT_PRECISION = Decimal("0.001")
GRAMS_PER_POUND = Decimal("453.59237")
SHOPIFY_VENDOR_SOURCES = {
    "MPWSR": {
        "feed_url": "https://mpwsr.com/products.json",
        "product_base": "https://mpwsr.com/products/",
    },
    "Barrens": {
        "feed_url": "https://www.barens.com/products.json",
        "product_base": "https://www.barens.com/products/",
    },
}
HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; AZCSInventoryBot/1.0)"}
JRACENSTEIN_BASE_URL = "https://jracenstein.com"
JRACENSTEIN_LISTING_URL = f"{JRACENSTEIN_BASE_URL}/jracenstein/"
JRACENSTEIN_GRAPHQL_URL = f"{JRACENSTEIN_BASE_URL}/graphql"
TRIDENT_SITEMAP_URL = "https://www.tridentprotects.com/sitemap.xml"
EACOCHEM_ALL_PRODUCTS_URL = "https://eacochem.com/all-products/"
GOLD_ASSASSIN_BASE_URL = "https://www.goldassassin.com"
GOLD_ASSASSIN_SITEMAP_URL = f"{GOLD_ASSASSIN_BASE_URL}/store-products-sitemap.xml"
JRACENSTEIN_DISTINCTIVE_KEYWORDS = (
    "SCRAPER",
    "BRUSH",
    "DUSTER",
    "SQUEEGEE",
    "CHANNEL",
    "SLEEVE",
    "HANDLE",
    "POLE",
    "CLIP",
    "RUBBER",
    "HOLSTER",
    "KIT",
    "TIP",
)
JRACENSTEIN_SEARCH_QUERY = """
query SearchProducts($term: String!) {
  site {
    search {
      searchProducts(filters: {searchTerm: $term}) {
        products(first: 8) {
          edges {
            node {
              path
              name
              sku
              upc
            }
          }
        }
      }
    }
  }
}
"""
JRACENSTEIN_PRODUCT_QUERY = """
query ProductByPath($path: String!) {
  site {
    route(path: $path) {
      node {
        __typename
        ... on Product {
          path
          name
          sku
          upc
          gtin
          mpn
          weight {
            value
            unit
          }
          description
          variants(first: 250) {
            edges {
              node {
                sku
                upc
                gtin
                mpn
                weight {
                  value
                  unit
                }
                productOptions {
                  edges {
                    node {
                      displayName
                      ... on MultipleChoiceOption {
                        values(first: 25) {
                          edges {
                            node {
                              label
                            }
                          }
                        }
                      }
                    }
                  }
                }
              }
            }
          }
        }
      }
    }
  }
}
"""


@dataclass
class SourceItem:
    vendor: str
    source_file: str
    item_name: str
    sku: str = ""
    gtin: str = ""
    description: str = ""
    category: str = ""
    reporting_category: str = ""
    default_unit_cost: Decimal | None = None
    price: Decimal | None = None
    vendor_code: str = ""
    notes: list[str] = field(default_factory=list)
    generated_sku: bool = False
    customer_facing_name_override: str = ""
    description_override: str = ""
    permalink_override: str = ""
    seo_title_override: str = ""
    seo_description_override: str = ""
    weight_lb_override: Decimal | None = None
    image_relative_path: str = ""
    image_absolute_path: str = ""
    image_match_type: str = ""
    image_source_folder: str = ""
    image_lookup_tokens: list[str] = field(default_factory=list)
    image_lookup_slug: str = ""


@dataclass
class ReviewIssue:
    issue_type: str
    vendor: str
    source_file: str
    item_name: str = ""
    sku: str = ""
    gtin: str = ""
    category: str = ""
    default_unit_cost: str = ""
    price: str = ""
    details: str = ""


@dataclass
class EnrichmentAuditEntry:
    enrichment_type: str
    vendor: str
    sku: str
    vendor_code: str
    item_name: str
    field: str
    value: str
    source: str
    details: str = ""


@dataclass
class ImageMatchAuditEntry:
    vendor: str
    sku: str
    vendor_code: str
    item_name: str
    image_relative_path: str
    image_absolute_path: str
    match_type: str
    source_folder: str
    details: str = ""


def clean_text(value: object) -> str:
    text = str(value or "")
    replacements = {
        "\u2019": "'",
        "\u2018": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2013": "-",
        "\u2014": "-",
        "\u2022": "",
        "\uf0b7": "",
        "\u00bd": "1/2",
        "\xa0": " ",
        "\n": " ",
        "\r": " ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" -")


def clean_code(value: object) -> str:
    text = clean_text(value).upper()
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", "", text)
    return text


def normalize_name(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean_text(value).upper())


def normalize_digits(value: object) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def normalize_sku(value: object) -> str:
    return re.sub(r"\s+", "", clean_text(value).upper())


def gtin_checksum_valid(value: str) -> bool:
    if not value.isdigit() or len(value) not in GTIN_LENGTHS:
        return False
    digits = [int(char) for char in value]
    total = 0
    for index, digit in enumerate(reversed(digits[:-1]), start=1):
        total += digit * (3 if index % 2 == 1 else 1)
    expected_check_digit = (10 - (total % 10)) % 10
    return expected_check_digit == digits[-1]


def parse_money(value: object) -> Decimal | None:
    if value is None:
        return None
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("$", "").replace(",", "").replace(" ", "")
    if not text:
        return None
    try:
        return Decimal(text).quantize(CENT, rounding=ROUND_HALF_UP)
    except Exception:
        return None


def extract_money_values(value: object) -> list[Decimal]:
    text = clean_text(value)
    results: list[Decimal] = []
    for match in MONEY_RE.findall(text):
        money = parse_money(match)
        if money is not None:
            results.append(money)
    return results


def format_money(value: Decimal | None) -> str:
    if value is None:
        return ""
    return str(value.quantize(CENT, rounding=ROUND_HALF_UP))


def strip_product_markers(name: str) -> tuple[str, list[str]]:
    notes: list[str] = []
    cleaned = clean_text(name)
    while cleaned.endswith(("+", "*")):
        marker = cleaned[-1]
        cleaned = cleaned[:-1].rstrip()
        if marker == "+":
            notes.append("Source sheet marks this as a hazmat item.")
        elif marker == "*":
            notes.append("Source sheet marks this item with a special shipping note.")
    return cleaned, notes


def valid_gtin(value: object) -> str:
    digits = normalize_digits(value)
    return digits if gtin_checksum_valid(digits) else ""


def build_description(*parts: str) -> str:
    cleaned_parts = [clean_text(part) for part in parts if clean_text(part)]
    return " | ".join(cleaned_parts)


def strip_html_to_text(value: object) -> str:
    text = str(value or "")
    text = re.sub(r"(?i)<br\s*/?>", " ", text)
    text = re.sub(r"(?i)</p>", " ", text)
    text = re.sub(r"(?i)</li>", " ", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    return clean_text(text)


def trim_words(text: str, max_length: int) -> str:
    cleaned = clean_text(text)
    if len(cleaned) <= max_length:
        return cleaned
    truncated = cleaned[: max_length - 1].rstrip(" ,;:-")
    if " " in truncated:
        truncated = truncated.rsplit(" ", 1)[0]
    return truncated.rstrip(" ,;:-") + "..."


def slugify(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def local_image_stem_key(value: str) -> str:
    stem = clean_text(Path(str(value)).stem).lower()
    stem = re.sub(r"_[0-9]+$", "", stem)
    stem = re.sub(r"^\d+[-_]+", "", stem)
    return slugify(stem)


def relative_repo_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(BASE_DIR.resolve()))
    except Exception:
        return str(path.resolve())


def format_weight(value: Decimal | None) -> str:
    if value is None:
        return ""
    quantized = value.quantize(WEIGHT_PRECISION, rounding=ROUND_HALF_UP)
    text = format(quantized, "f").rstrip("0").rstrip(".")
    return text or "0"


def product_display_name(item: SourceItem) -> str:
    return clean_text(item.customer_facing_name_override or item.item_name)


def product_description_text(item: SourceItem) -> str:
    return clean_text(item.description_override or item.description)


def split_variant_suffix(item_name: str) -> tuple[str, str]:
    base, separator, suffix = clean_text(item_name).partition(" - ")
    return clean_text(base), clean_text(suffix) if separator else ""


def combine_title_with_suffix(title: str, suffix: str) -> str:
    return clean_text(f"{clean_text(title)} - {clean_text(suffix)}" if clean_text(suffix) else title)


def path_slug_from_url(url: str) -> str:
    path = urlparse(url).path.rstrip("/")
    return clean_text(path.split("/")[-1]) if path else ""


def category_segments(value: str) -> list[str]:
    return [clean_text(segment) for segment in str(value or "").split(">") if clean_text(segment)]


def category_tail(value: str) -> str:
    segments = category_segments(value)
    if len(segments) <= 1:
        return ""
    for segment in reversed(segments):
        if segment.upper() not in GENERIC_CATEGORY_SEGMENTS:
            return segment
    return segments[-1] if segments else ""


def seo_keyword_base(item_name: str) -> str:
    return clean_text(re.sub(r"\[[^\]]+\]", "", item_name))


def build_seo_title(item: SourceItem) -> str:
    if clean_text(item.seo_title_override):
        return trim_words(item.seo_title_override, 78)

    base = seo_keyword_base(product_display_name(item))
    context = category_tail(item.category)
    suffix = "AZ Cleaning Supplies"
    candidates = []
    if context and context.upper() not in base.upper():
        candidates.append(f"{base} | {context} | {suffix}")
    candidates.append(f"{base} | {suffix}")
    for candidate in candidates:
        if len(candidate) <= 78:
            return candidate
    return trim_words(candidates[-1], 78)


def build_seo_description(item: SourceItem, seo_title: str) -> str:
    if clean_text(item.seo_description_override):
        return trim_words(item.seo_description_override, 160)

    description_override = clean_text(item.description_override)
    if description_override and normalize_name(description_override) != normalize_name(product_display_name(item)):
        return trim_words(f"{description_override} Available at AZ Cleaning Supplies.", 160)

    base = seo_keyword_base(product_display_name(item))
    description_parts = [base]
    context = category_tail(item.category)
    if context and context.upper() not in base.upper():
        description_parts.append(f"Category: {context}")
    vendor_code = clean_text(item.vendor_code or item.sku)
    if vendor_code:
        description_parts.append(f"Vendor code: {vendor_code}")
    description_parts.append("Available at AZ Cleaning Supplies")
    return trim_words(". ".join(description_parts) + ".", 160)


def make_category(vendor: str, category: str = "") -> str:
    if category:
        return f"{vendor} > {clean_text(category)}"
    return vendor


def read_csv_any(path: Path) -> list[list[str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    raise RuntimeError(f"Could not read {path}")


def resolve_template_path() -> Path:
    if TEMPLATE_PATH.exists():
        return TEMPLATE_PATH
    if LEGACY_TEMPLATE_PATH.exists():
        return LEGACY_TEMPLATE_PATH
    raise FileNotFoundError("Square Import Template.csv was not found in templates/ or the repo root.")


def resolve_latest_source(patterns: list[str]) -> Path:
    candidates: list[Path] = []
    search_roots = [PRICE_LIST_DIR, BASE_DIR]
    for root in search_roots:
        if not root.exists():
            continue
        for pattern in patterns:
            candidates.extend(root.glob(pattern))
    unique_candidates = sorted({path.resolve() for path in candidates if path.is_file()})
    if not unique_candidates:
        raise FileNotFoundError(f"No source file matched patterns: {patterns}")
    return max(
        unique_candidates,
        key=lambda path: (path.stat().st_mtime, 1 if PRICE_LIST_DIR in path.parents else 0),
    )


def load_square_headers(path: Path) -> list[str]:
    rows = read_csv_any(path)
    for row in rows:
        if row and clean_text(row[0]) == "Token":
            return [clean_text(cell) for cell in row]
    raise ValueError(f"Could not find Square header row in {path}")


def parse_barrens(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    rows = read_csv_any(path)
    header_row = next(i for i, row in enumerate(rows) if row and clean_text(row[0]) == "StockCode")
    items: list[SourceItem] = []
    current_category = ""

    for row in rows[header_row + 1 :]:
        first = clean_text(row[0] if len(row) > 0 else "")
        second = clean_text(row[1] if len(row) > 1 else "")
        if not first and not second:
            continue

        price = parse_money(row[2] if len(row) > 2 else None)
        if first and not second and price is None:
            current_category = first
            continue

        if not first or not second or price is None:
            continue

        items.append(
            SourceItem(
                vendor="Barrens",
                source_file=path.name,
                item_name=second,
                sku=first,
                description=build_description(second, f"Category: {current_category}" if current_category else ""),
                category=make_category("Barrens", current_category),
                reporting_category="Barrens",
                default_unit_cost=price,
                vendor_code=first,
            )
        )

    return items, []


def parse_mpwsr(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    rows = read_csv_any(path)
    headers = [clean_text(value) for value in rows[0]]
    index = {name: headers.index(name) for name in headers}
    items: list[SourceItem] = []

    for row in rows[1:]:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))

        sku = clean_text(row[index["Name"]])
        item_name = clean_text(row[index["Description"]])
        cost = parse_money(row[index["Dealers"]])
        price = parse_money(row[index["Base Price"]])
        if not item_name:
            continue

        items.append(
            SourceItem(
                vendor="MPWSR",
                source_file=path.name,
                item_name=item_name,
                sku=sku,
                description=item_name,
                category=make_category("MPWSR"),
                reporting_category="MPWSR",
                default_unit_cost=cost,
                price=price,
                vendor_code=sku,
            )
        )

    return items, []


def parse_inseco(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    items: list[SourceItem] = []
    current_category = ""

    for row in sheet.iter_rows(values_only=True):
        cells = list(row[:4])
        values = [clean_text(value) for value in cells]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue

        if len(non_empty) == 1:
            text = non_empty[0]
            if "DISTRIBUTOR PRICE LIST" in text.upper():
                continue
            if text.upper() in {"SKU", "PRODUCT#", "PRICE"}:
                continue
            current_category = text
            continue

        item_name = clean_text(cells[0])
        gtin = valid_gtin(cells[1])
        sku = clean_text(cells[2])
        cost = parse_money(cells[3])
        if not item_name or cost is None:
            continue

        items.append(
            SourceItem(
                vendor="INSECO",
                source_file=path.name,
                item_name=item_name,
                sku=sku,
                gtin=gtin,
                description=build_description(item_name, f"Category: {current_category}" if current_category else ""),
                category=make_category("INSECO", current_category),
                reporting_category="INSECO",
                default_unit_cost=cost,
                vendor_code=sku,
            )
        )

    workbook.close()
    return items, []


def parse_jracenstein(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [clean_text(value) for value in rows[0]]
    index = {name: headers.index(name) for name in headers}
    items: list[SourceItem] = []

    for row in rows[1:]:
        if not row:
            continue
        sku = clean_text(row[index["Code"]] if len(row) > index["Code"] else "")
        item_name = clean_text(row[index["Model"]] if len(row) > index["Model"] else "")
        category = clean_text(row[index["Category"]] if len(row) > index["Category"] else "")
        case_qty = clean_text(row[index["Case"]] if len(row) > index["Case"] else "")
        price = parse_money(row[index["2026 List Price"]] if len(row) > index["2026 List Price"] else None)
        cost = parse_money(row[index["2026 Distributor Price"]] if len(row) > index["2026 Distributor Price"] else None)
        if not item_name:
            continue

        description = build_description(
            item_name,
            f"Category: {category}" if category else "",
            f"Case pack: {case_qty}" if case_qty else "",
        )
        items.append(
            SourceItem(
                vendor="JRacenstein",
                source_file=path.name,
                item_name=item_name,
                sku=sku,
                description=description,
                category=make_category("JRacenstein", category),
                reporting_category="JRacenstein",
                default_unit_cost=cost,
                price=price,
                vendor_code=sku,
            )
        )

    workbook.close()
    return items, []


def infer_be_category(description: str) -> str:
    text = clean_text(description).upper()
    if text.startswith("HW"):
        return "Hot Water Equipment"
    if text.startswith("CW"):
        return "Cold Water Equipment"
    if text.startswith("PW"):
        return "Pressure Washers"
    return "Equipment"


def parse_be(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    items: list[SourceItem] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue

            for row in table:
                row = list(row or [])
                if len(row) < 6:
                    continue
                stockcode = clean_text(row[0])
                if not stockcode or stockcode == "Stockcode":
                    continue
                description = clean_text(f"{clean_text(row[1])} {clean_text(row[2])}")
                upc = valid_gtin(row[4])
                cost = parse_money(row[5])
                if not description:
                    continue

                category = infer_be_category(description)
                items.append(
                    SourceItem(
                        vendor="BE",
                        source_file=path.name,
                        item_name=description,
                        sku=stockcode,
                        gtin=upc,
                        description=description,
                        category=make_category("BE", category),
                        reporting_category="BE",
                        default_unit_cost=cost,
                        vendor_code=stockcode,
                    )
                )

    return items, []


def parse_trident(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    items: list[SourceItem] = []
    current_section = ""
    current_pack = ""

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue

            for row in table:
                row = list(row or [])
                row += [""] * (15 - len(row))
                name_cell = clean_text(row[1])
                joined = " ".join(clean_text(cell) for cell in row if clean_text(cell))

                if not name_cell and not joined:
                    continue

                if name_cell in {"Sealers", "Cleaners", "Sand"} or name_cell.startswith("2 Part Sealers"):
                    current_section = name_cell
                    if "5 Gallon Pail" in joined:
                        current_pack = "5 Gallon Pail"
                    elif "4 Gallon Case" in joined:
                        current_pack = "4 Gallon Case (per gallon)"
                    elif "50 lb bag" in joined:
                        current_pack = "50 lb bag"
                    elif "Kits - Dealer" in joined:
                        current_pack = "Kit"
                    continue

                if "Pricing" in joined or "Free Shipping" in joined or name_cell.startswith("$2,500"):
                    continue
                if name_cell.startswith("Sand:"):
                    continue
                if name_cell.startswith("*Full Pallet") or name_cell.startswith("**Mixed Pallet"):
                    continue

                dealer_prices = extract_money_values(row[4])
                direct_price = parse_money(row[7])
                vendor_number = clean_text(row[10])
                product_number = clean_code(row[13])
                if not name_cell or not dealer_prices:
                    continue

                base_name, notes = strip_product_markers(name_cell)
                item_name = f"{base_name} - {current_pack}" if current_pack else base_name
                cost = max(dealer_prices)
                if len(dealer_prices) > 1:
                    notes.append("Default Unit Cost uses the higher dealer tier shown in the source sheet.")

                items.append(
                    SourceItem(
                        vendor="Trident",
                        source_file=path.name,
                        item_name=item_name,
                        sku=product_number,
                        description=build_description(base_name, f"Section: {current_section}", f"Pack: {current_pack}" if current_pack else ""),
                        category=make_category("Trident", current_section),
                        reporting_category="Trident",
                        default_unit_cost=cost,
                        price=direct_price,
                        vendor_code=vendor_number or product_number,
                        notes=notes,
                    )
                )

    return items, []


def append_eaco_item(
    items: list[SourceItem],
    vendor: str,
    source_file: str,
    base_name: str,
    section: str,
    size_label: str,
    prices: list[Decimal],
    notes: list[str] | None = None,
) -> None:
    if not prices:
        return
    if len(prices) >= 3:
        cost = max(prices[:2])
    else:
        cost = max(prices)
    price = prices[2] if len(prices) >= 3 else None
    item_name = f"{base_name} - {size_label}"
    items.append(
        SourceItem(
            vendor=vendor,
            source_file=source_file,
            item_name=item_name,
            description=build_description(base_name, f"Section: {section}", f"Pack: {size_label}"),
            category=make_category(vendor, section),
            reporting_category=vendor,
            default_unit_cost=cost,
            price=price,
            notes=list(notes or []),
        )
    )


def parse_eaco_new_construction(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    items: list[SourceItem] = []
    issues: list[ReviewIssue] = []
    current_section = "New Construction & Restoration"

    with pdfplumber.open(path) as pdf:
        table = pdf.pages[0].extract_table()
        if not table:
            issues.append(
                ReviewIssue(
                    issue_type="parse_failure",
                    vendor="EacoChem",
                    source_file=path.name,
                    details="Could not extract table from the EacoChem new construction PDF.",
                )
            )
            return items, issues

        for row in table[2:]:
            row = list(row or [])
            row += [""] * (10 - len(row))
            name_cell = clean_text(row[0])
            if not name_cell:
                continue
            if name_cell == "Sealers":
                current_section = "Sealers"
                continue

            base_name, notes = strip_product_markers(name_cell)
            drum_prices = extract_money_values(" ".join(clean_text(cell) for cell in row[1:4]))
            pail_prices = extract_money_values(" ".join(clean_text(cell) for cell in row[4:7]))
            gallon_prices = extract_money_values(" ".join(clean_text(cell) for cell in row[7:10]))
            gallon_label = "2 Gallon" if "2Gal" in " ".join(clean_text(cell) for cell in row[7:10]) else "1 Gallon"

            if drum_prices:
                append_eaco_item(items, "EacoChem", path.name, base_name, current_section, "55 Gallon Drum", drum_prices, notes)
            if pail_prices:
                append_eaco_item(items, "EacoChem", path.name, base_name, current_section, "5 Gallon Pail", pail_prices, notes)
            if gallon_prices:
                append_eaco_item(items, "EacoChem", path.name, base_name, current_section, gallon_label, gallon_prices, notes)

            if not drum_prices and not pail_prices and not gallon_prices:
                issues.append(
                    ReviewIssue(
                        issue_type="ambiguous_source_row",
                        vendor="EacoChem",
                        source_file=path.name,
                        item_name=base_name,
                        category=current_section,
                        details="No usable prices were extracted from this EacoChem row.",
                    )
                )

    return items, issues


def parse_eaco_fleet(path: Path) -> tuple[list[SourceItem], list[ReviewIssue]]:
    items: list[SourceItem] = []
    issues: list[ReviewIssue] = []
    reader = PdfReader(str(path))
    text = reader.pages[0].extract_text() or ""
    capture = False

    for raw_line in text.splitlines():
        line = clean_text(raw_line)
        if not line:
            continue
        if "per Case" in line:
            capture = True
            continue
        if not capture:
            continue
        if line.startswith("***GLORY") or line.startswith("Rev:") or "EaCo Chem" in line:
            break

        prices = extract_money_values(line)
        if not prices:
            continue

        first_price_match = MONEY_RE.search(line)
        if not first_price_match:
            continue

        name_text = clean_text(line[: first_price_match.start()])
        base_name, notes = strip_product_markers(name_text)

        if len(prices) >= 8:
            append_eaco_item(items, "EacoChem", path.name, base_name, "Fleet Wash Products", "55 Gallon Drum", prices[0:3], notes)
            append_eaco_item(items, "EacoChem", path.name, base_name, "Fleet Wash Products", "5 Gallon Pail", prices[3:6], notes)
            append_eaco_item(
                items,
                "EacoChem",
                path.name,
                base_name,
                "Fleet Wash Products",
                "1 Gallon",
                [prices[6], prices[6], prices[7]],
                notes,
            )
        elif len(prices) == 6:
            append_eaco_item(items, "EacoChem", path.name, base_name, "Fleet Wash Products", "55 Gallon Drum", prices[0:3], notes)
            append_eaco_item(items, "EacoChem", path.name, base_name, "Fleet Wash Products", "5 Gallon Pail", prices[3:6], notes)
        elif len(prices) == 1:
            items.append(
                SourceItem(
                    vendor="EacoChem",
                    source_file=path.name,
                    item_name=base_name,
                    description=build_description(base_name, "Section: Fleet Wash Products", "Special pack pricing from source sheet"),
                    category=make_category("EacoChem", "Fleet Wash Products"),
                    reporting_category="EacoChem",
                    default_unit_cost=prices[0],
                    notes=notes + ["This source row only provided one price; selling price was left blank."],
                )
            )
        else:
            issues.append(
                ReviewIssue(
                    issue_type="ambiguous_source_row",
                    vendor="EacoChem",
                    source_file=path.name,
                    item_name=base_name,
                    details=f"Unexpected number of price columns ({len(prices)}) in fleet sheet line: {line}",
                )
            )

    return items, issues


def load_verified_enrichments(path: Path) -> dict[tuple[str, str, str], dict[str, str]]:
    if not path.exists():
        return {}

    rows = read_csv_any(path)
    if not rows:
        return {}

    headers = [clean_text(value) for value in rows[0]]
    index = {name: position for position, name in enumerate(headers)}
    enrichments: dict[tuple[str, str, str], dict[str, str]] = {}

    def optional_value(values: list[str], column_name: str) -> str:
        if column_name not in index:
            return ""
        position = index[column_name]
        return values[position] if position < len(values) else ""

    for row in rows[1:]:
        padded = row + [""] * (len(headers) - len(row))
        vendor = clean_text(padded[index["vendor"]]).upper()
        match_field = clean_text(padded[index["match_field"]]).lower()
        match_value = clean_text(padded[index["match_value"]])
        if not vendor or not match_field or not match_value:
            continue
        enrichments[(vendor, match_field, clean_text(match_value).upper())] = {
            "gtin": valid_gtin(optional_value(padded, "gtin")),
            "seo_title": clean_text(optional_value(padded, "seo_title")),
            "seo_description": clean_text(optional_value(padded, "seo_description")),
            "source_url": clean_text(optional_value(padded, "source_url")),
            "notes": clean_text(optional_value(padded, "notes")),
        }

    return enrichments


def fetch_shopify_products(feed_url: str) -> list[dict[str, object]]:
    page = 1
    products: list[dict[str, object]] = []

    while True:
        response = requests.get(
            feed_url,
            params={"limit": 250, "page": page},
            headers=HTTP_HEADERS,
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        batch = payload.get("products", [])
        if not isinstance(batch, list) or not batch:
            break
        products.extend(batch)
        if len(batch) < 250:
            break
        page += 1

    return products


def shopify_exact_title_key(value: str) -> str:
    return normalize_name(value)


def shopify_description_text(body_html: object) -> str:
    text = strip_html_to_text(body_html)
    if len(text) < 32:
        return ""
    return trim_words(text, 1000)


def is_descriptive_shopify_title(title: str, sku: str) -> bool:
    cleaned = clean_text(title)
    if not cleaned:
        return False
    if normalize_name(cleaned) == normalize_name(sku):
        return False
    descriptive_tokens = [
        token
        for token in re.findall(r"[A-Za-z][A-Za-z0-9./-]*", cleaned)
        if len(re.sub(r"[^A-Za-z]", "", token)) >= 4
    ]
    if any(char.isdigit() for char in cleaned) and len(descriptive_tokens) < 2:
        return False
    return bool(descriptive_tokens)


def is_meaningful_shopify_handle(handle: str) -> bool:
    cleaned = clean_text(handle)
    if not cleaned or not re.search(r"[A-Za-z]", cleaned):
        return False
    if re.fullmatch(r"product[_-][0-9a-f-]{16,}", cleaned.lower()):
        return False
    return True


def shopify_weight_lb(grams: object) -> Decimal | None:
    try:
        value = Decimal(str(grams or 0))
    except Exception:
        return None
    if value <= 0:
        return None
    return (value / GRAMS_PER_POUND).quantize(WEIGHT_PRECISION, rounding=ROUND_HALF_UP)


def apply_shopify_vendor_enrichments(
    items: list[SourceItem],
) -> tuple[list[EnrichmentAuditEntry], Counter[str], Counter[str], list[str]]:
    audit_entries: list[EnrichmentAuditEntry] = []
    match_counts: Counter[str] = Counter()
    detail_counts: Counter[str] = Counter()
    notes: list[str] = []

    for vendor, config in SHOPIFY_VENDOR_SOURCES.items():
        try:
            products = fetch_shopify_products(config["feed_url"])
        except Exception as exc:
            notes.append(f"{vendor} website enrichment skipped: {exc}")
            continue

        sku_matches: dict[str, list[tuple[dict[str, object], dict[str, object]]]] = defaultdict(list)
        title_matches: dict[str, list[tuple[dict[str, object], dict[str, object]]]] = defaultdict(list)
        for product in products:
            product_title = clean_text(product.get("title", ""))
            variants = product.get("variants", [])
            if not isinstance(variants, list):
                continue
            if len(variants) == 1 and product_title:
                variant = variants[0]
                if isinstance(variant, dict):
                    title_matches[shopify_exact_title_key(product_title)].append((product, variant))
            for variant in variants:
                if not isinstance(variant, dict):
                    continue
                sku_key = normalize_sku(variant.get("sku", ""))
                if sku_key:
                    sku_matches[sku_key].append((product, variant))

        unique_sku_matches = {
            key: matches[0]
            for key, matches in sku_matches.items()
            if len(matches) == 1
        }
        unique_title_matches = {
            key: matches[0]
            for key, matches in title_matches.items()
            if len(matches) == 1
        }

        for item in [candidate for candidate in items if candidate.vendor == vendor]:
            match: tuple[dict[str, object], dict[str, object]] | None = None
            match_type = ""
            for token in (normalize_sku(item.vendor_code), normalize_sku(item.sku)):
                if token and token in unique_sku_matches:
                    match = unique_sku_matches[token]
                    match_type = "shopify_sku_match"
                    break
            if match is None:
                title_key = shopify_exact_title_key(item.item_name)
                if title_key and title_key in unique_title_matches:
                    match = unique_title_matches[title_key]
                    match_type = "shopify_title_match"
            if match is None:
                continue

            product, variant = match
            product_title = clean_text(product.get("title", ""))
            product_url = f"{config['product_base']}{clean_text(product.get('handle', ''))}"
            if (
                product_title
                and is_descriptive_shopify_title(product_title, item.vendor_code or item.sku)
                and product_title != product_display_name(item)
            ):
                item.customer_facing_name_override = product_title
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type=match_type,
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Customer-facing Name",
                        value=product_title,
                        source=product_url,
                        details="Pulled from the vendor website product title.",
                    )
                )
                detail_counts["customer_names"] += 1

            description_text = shopify_description_text(product.get("body_html", ""))
            if description_text and normalize_name(description_text) != normalize_name(product_description_text(item)):
                item.description_override = description_text
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type=match_type,
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Description",
                        value=trim_words(description_text, 120),
                        source=product_url,
                        details="Pulled from the vendor website product description.",
                    )
                )
                detail_counts["descriptions"] += 1

            handle = clean_text(product.get("handle", ""))
            if handle and is_meaningful_shopify_handle(handle):
                item.permalink_override = handle
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type=match_type,
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Permalink",
                        value=handle,
                        source=product_url,
                        details="Using the vendor website product handle as the preferred permalink.",
                    )
                )
                detail_counts["permalinks"] += 1

            weight_lb = shopify_weight_lb(variant.get("grams"))
            if weight_lb is not None:
                item.weight_lb_override = weight_lb
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type=match_type,
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Weight (lb)",
                        value=format_weight(weight_lb),
                        source=product_url,
                        details="Converted from the vendor website shipping weight in grams.",
                    )
                )
                detail_counts["weights"] += 1

            match_counts[vendor] += 1

        notes.append(
            f"{vendor} website enrichment used {len(products)} live products and matched {match_counts[vendor]} catalog rows."
        )

    return audit_entries, match_counts, detail_counts, notes


def jracenstein_clean_identifier(value: object) -> str:
    return clean_text(str(value or "").replace("\x00", ""))


def jracenstein_stripped_numeric_code(value: object) -> str:
    digits = normalize_digits(jracenstein_clean_identifier(value))
    return digits.lstrip("0") or digits


def measurement_to_lb(value: object) -> Decimal | None:
    if not isinstance(value, dict):
        return None
    try:
        amount = Decimal(str(value.get("value")))
    except Exception:
        return None
    if amount <= 0:
        return None
    unit = clean_text(value.get("unit", "")).lower()
    if unit in {"lb", "lbs", "pound", "pounds"}:
        pounds = amount
    elif unit in {"oz", "ounce", "ounces"}:
        pounds = amount / Decimal("16")
    elif unit in {"g", "gram", "grams"}:
        pounds = amount / GRAMS_PER_POUND
    else:
        return None
    return pounds.quantize(WEIGHT_PRECISION, rounding=ROUND_HALF_UP)


def normalize_numeric_hint(value: str) -> str:
    cleaned = clean_text(value)
    if "." in cleaned:
        cleaned = cleaned.rstrip("0").rstrip(".")
    return cleaned


def extract_size_hints(value: str) -> set[str]:
    text = clean_text(re.sub(r"\[\s*Case[^\]]*\]", " ", value, flags=re.I)).upper()
    hints: set[str] = set()

    for match in re.finditer(r'(\d+(?:\.\d+)?)\s*(?:INCH(?:ES)?|IN\b|")', text):
        hints.add(normalize_numeric_hint(match.group(1)))

    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        hints.add(normalize_numeric_hint(text))
    elif any(
        token in text
        for token in ("INCH", "BRASS", "STAINLESS", "COMPLETE", "CHANNEL", "SQUEEGEE", "SLEEVE", "RUBBER")
    ):
        for match in re.finditer(r"(?<![A-Z0-9])(\d+(?:\.\d+)?)(?![A-Z0-9])", text):
            hints.add(normalize_numeric_hint(match.group(1)))

    return {hint for hint in hints if hint}


def extract_case_suffix_from_name(value: str) -> str:
    match = re.search(r"(\[Case[^\]]+\])", value, re.I)
    return clean_text(match.group(1)) if match else ""


def jracenstein_similarity_key(value: str) -> str:
    text = clean_text(value).upper().replace("PRO+", "PROPLUS").replace("W/", "WITH ")
    text = re.sub(r"\[\s*CASE[^\]]*\]", " ", text, flags=re.I)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def jracenstein_variant_labels(variant: dict[str, object]) -> list[str]:
    labels: list[str] = []
    option_edges = variant.get("productOptions", {}).get("edges", []) if isinstance(variant, dict) else []
    for option_edge in option_edges:
        option = option_edge.get("node", {}) if isinstance(option_edge, dict) else {}
        value_edges = option.get("values", {}).get("edges", []) if isinstance(option, dict) else []
        option_values = [
            clean_text(value_edge.get("node", {}).get("label", ""))
            for value_edge in value_edges
            if isinstance(value_edge, dict) and clean_text(value_edge.get("node", {}).get("label", ""))
        ]
        if option_values:
            labels.append(", ".join(option_values))
    return labels


def fetch_jracenstein_storefront_token(session: requests.Session) -> str:
    response = session.get(JRACENSTEIN_LISTING_URL, headers=HTTP_HEADERS, timeout=30)
    response.raise_for_status()
    match = re.search(r'const STOREFRONT_TOKEN = "([^"]+)"', response.text)
    if not match:
        raise RuntimeError("Could not extract the JRacenstein Storefront token.")
    return clean_text(match.group(1))


def jracenstein_graphql(
    session: requests.Session,
    token: str,
    query: str,
    variables: dict[str, object],
) -> dict[str, object]:
    response = session.post(
        JRACENSTEIN_GRAPHQL_URL,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            **HTTP_HEADERS,
        },
        json={"query": query, "variables": variables},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("errors"):
        raise RuntimeError(str(payload["errors"]))
    return payload.get("data", {})


def search_jracenstein_products(
    session: requests.Session,
    token: str,
    term: str,
    search_cache: dict[str, list[dict[str, object]]],
) -> list[dict[str, object]]:
    cleaned_term = jracenstein_clean_identifier(term)
    if not cleaned_term:
        return []
    if cleaned_term not in search_cache:
        data = jracenstein_graphql(session, token, JRACENSTEIN_SEARCH_QUERY, {"term": cleaned_term})
        edges = data.get("site", {}).get("search", {}).get("searchProducts", {}).get("products", {}).get("edges", [])
        search_cache[cleaned_term] = [
            edge.get("node", {})
            for edge in edges
            if isinstance(edge, dict) and isinstance(edge.get("node"), dict)
        ]
    return search_cache[cleaned_term]


def fetch_jracenstein_product_data(
    session: requests.Session,
    token: str,
    path: str,
    product_cache: dict[str, dict[str, object]],
) -> dict[str, object]:
    cleaned_path = clean_text(path)
    if not cleaned_path:
        return {}
    if cleaned_path not in product_cache:
        data = jracenstein_graphql(session, token, JRACENSTEIN_PRODUCT_QUERY, {"path": cleaned_path})
        product_cache[cleaned_path] = data.get("site", {}).get("route", {}).get("node", {}) or {}
    return product_cache[cleaned_path]


def jracenstein_code_matches(value: object, target_code: str) -> bool:
    candidate = jracenstein_clean_identifier(value)
    target = jracenstein_clean_identifier(target_code)
    if not candidate or not target:
        return False
    if candidate == target:
        return True
    if target.startswith("0"):
        stripped_target = jracenstein_stripped_numeric_code(target)
        return bool(stripped_target and jracenstein_stripped_numeric_code(candidate) == stripped_target)
    return False


def jracenstein_code_match_rank(kind: str, field: str) -> int:
    field_rank = {"mpn": 4, "sku": 3, "gtin": 2, "upc": 1}.get(field, 0)
    return (10 if kind == "variant" else 0) + field_rank


def build_jracenstein_exact_candidates(
    session: requests.Session,
    token: str,
    code: str,
    search_cache: dict[str, list[dict[str, object]]],
    product_cache: dict[str, dict[str, object]],
) -> list[dict[str, object]]:
    cleaned_code = jracenstein_clean_identifier(code)
    if not cleaned_code:
        return []

    search_terms = [cleaned_code]
    stripped_code = jracenstein_stripped_numeric_code(cleaned_code)
    if cleaned_code.startswith("0") and stripped_code and stripped_code not in search_terms:
        search_terms.append(stripped_code)

    candidates: list[dict[str, object]] = []
    seen_paths: set[str] = set()

    for term in search_terms:
        for result in search_jracenstein_products(session, token, term, search_cache):
            product_path = clean_text(result.get("path", ""))
            if not product_path or product_path in seen_paths:
                continue
            seen_paths.add(product_path)
            product = fetch_jracenstein_product_data(session, token, product_path, product_cache)
            if product.get("__typename") != "Product":
                continue

            product_name = clean_text(product.get("name", ""))
            product_url = f"{JRACENSTEIN_BASE_URL}{product_path}"
            description_text = trim_words(strip_html_to_text(product.get("description", "")), 1000)
            if len(description_text) < 40:
                description_text = ""

            exact_matches: list[dict[str, object]] = []
            for field in ("sku", "upc", "gtin", "mpn"):
                if jracenstein_code_matches(product.get(field), cleaned_code):
                    exact_matches.append(
                        {
                            "product_name": product_name,
                            "product_path": product_path,
                            "product_url": product_url,
                            "product_description": description_text,
                            "product_sku": jracenstein_clean_identifier(product.get("sku")),
                            "product_upc": jracenstein_clean_identifier(product.get("upc")),
                            "product_gtin": jracenstein_clean_identifier(product.get("gtin")),
                            "product_mpn": jracenstein_clean_identifier(product.get("mpn")),
                            "product_weight_lb": measurement_to_lb(product.get("weight")),
                            "variant_labels": [],
                            "variant_sku": "",
                            "variant_upc": "",
                            "variant_gtin": "",
                            "variant_mpn": "",
                            "variant_weight_lb": None,
                            "matched_kind": "product",
                            "matched_field": field,
                            "matched_value": jracenstein_clean_identifier(product.get(field)),
                        }
                    )

            variant_edges = product.get("variants", {}).get("edges", [])
            for variant_edge in variant_edges:
                variant = variant_edge.get("node", {}) if isinstance(variant_edge, dict) else {}
                for field in ("sku", "upc", "gtin", "mpn"):
                    if not jracenstein_code_matches(variant.get(field), cleaned_code):
                        continue
                    exact_matches.append(
                        {
                            "product_name": product_name,
                            "product_path": product_path,
                            "product_url": product_url,
                            "product_description": description_text,
                            "product_sku": jracenstein_clean_identifier(product.get("sku")),
                            "product_upc": jracenstein_clean_identifier(product.get("upc")),
                            "product_gtin": jracenstein_clean_identifier(product.get("gtin")),
                            "product_mpn": jracenstein_clean_identifier(product.get("mpn")),
                            "product_weight_lb": measurement_to_lb(product.get("weight")),
                            "variant_labels": jracenstein_variant_labels(variant),
                            "variant_sku": jracenstein_clean_identifier(variant.get("sku")),
                            "variant_upc": jracenstein_clean_identifier(variant.get("upc")),
                            "variant_gtin": jracenstein_clean_identifier(variant.get("gtin")),
                            "variant_mpn": jracenstein_clean_identifier(variant.get("mpn")),
                            "variant_weight_lb": measurement_to_lb(variant.get("weight")),
                            "matched_kind": "variant",
                            "matched_field": field,
                            "matched_value": jracenstein_clean_identifier(variant.get(field)),
                        }
                    )

            if not exact_matches:
                continue

            candidates.append(
                max(
                    exact_matches,
                    key=lambda match: jracenstein_code_match_rank(match["matched_kind"], match["matched_field"]),
                )
            )

        if candidates:
            break

    return candidates


def jracenstein_candidate_name(item: SourceItem, candidate: dict[str, object]) -> str:
    base = clean_text(candidate.get("product_name", ""))
    for label in candidate.get("variant_labels", []):
        cleaned_label = clean_text(label)
        if cleaned_label and normalize_name(cleaned_label) not in normalize_name(base):
            base = clean_text(f"{base} {cleaned_label}")
    case_suffix = extract_case_suffix_from_name(item.item_name)
    if case_suffix and normalize_name(case_suffix) not in normalize_name(base):
        base = clean_text(f"{base} {case_suffix}")
    return base


def jracenstein_candidate_gtin(candidate: dict[str, object]) -> str:
    for value in (
        candidate.get("variant_gtin", ""),
        candidate.get("variant_upc", ""),
        candidate.get("product_gtin", ""),
        candidate.get("product_upc", ""),
    ):
        gtin = valid_gtin(value)
        if gtin:
            return gtin
    return ""


def jracenstein_candidate_weight(candidate: dict[str, object]) -> Decimal | None:
    return candidate.get("variant_weight_lb") or candidate.get("product_weight_lb")


def build_jracenstein_permalink(item: SourceItem, candidate: dict[str, object]) -> str:
    base = path_slug_from_url(candidate.get("product_url", ""))
    if not base:
        base = slugify(jracenstein_candidate_name(item, candidate))
    combined = base
    for label in candidate.get("variant_labels", []):
        label_slug = slugify(label)
        if label_slug and normalize_name(label_slug) not in normalize_name(combined):
            combined = f"{combined}-{label_slug}"
    case_pack = extract_case_pack(item.description) or extract_case_suffix_from_name(item.item_name)
    case_slug = slugify(case_pack)
    if case_slug:
        case_segment = case_slug if case_slug.startswith("case-") else f"case-{case_slug}"
        if normalize_name(case_segment) not in normalize_name(combined):
            combined = f"{combined}-{case_segment}"
    return clean_text(combined)


def unique_nonempty(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = clean_text(value)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def filter_jracenstein_candidates_by_keywords(
    item: SourceItem,
    candidates: list[dict[str, object]],
) -> list[dict[str, object]]:
    filtered = candidates
    item_upper = clean_text(item.item_name).upper()
    for keyword in JRACENSTEIN_DISTINCTIVE_KEYWORDS:
        if keyword not in item_upper:
            continue
        keyword_matches = [
            candidate
            for candidate in filtered
            if keyword in jracenstein_candidate_name(item, candidate).upper()
        ]
        if len(keyword_matches) == 1:
            return keyword_matches
        if keyword_matches:
            filtered = keyword_matches
    return filtered


def jracenstein_match_score(item: SourceItem, candidate: dict[str, object]) -> float:
    item_key = jracenstein_similarity_key(item.item_name)
    candidate_name = jracenstein_candidate_name(item, candidate)
    candidate_key = jracenstein_similarity_key(candidate_name)
    score = SequenceMatcher(None, item_key, candidate_key).ratio()

    item_sizes = extract_size_hints(item.item_name)
    candidate_sizes = extract_size_hints(candidate_name)
    if item_sizes and candidate_sizes:
        if item_sizes & candidate_sizes:
            score += 0.20
        else:
            score -= 0.25

    for keyword, bonus in (
        ("KIT", 0.06),
        ("PROPLUS", 0.08),
        ("BRUSH", 0.05),
        ("DUSTER", 0.05),
        ("SCRAPER", 0.05),
        ("CHANNEL", 0.05),
        ("SQUEEGEE", 0.05),
        ("HANDLE", 0.04),
        ("POLE", 0.04),
    ):
        if keyword in item_key and keyword in candidate_key:
            score += bonus

    return score


def resolve_jracenstein_candidate(
    item: SourceItem,
    candidates: list[dict[str, object]],
) -> dict[str, object] | None:
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    gtins = {jracenstein_candidate_gtin(candidate) for candidate in candidates if jracenstein_candidate_gtin(candidate)}
    filtered = candidates
    item_sizes = extract_size_hints(item.item_name)
    if item_sizes:
        size_matches = [
            candidate
            for candidate in candidates
            if extract_size_hints(jracenstein_candidate_name(item, candidate)) & item_sizes
        ]
        if len(size_matches) == 1:
            return size_matches[0]
        if size_matches:
            filtered = size_matches

    filtered = filter_jracenstein_candidates_by_keywords(item, filtered)
    scored = sorted(
        ((jracenstein_match_score(item, candidate), candidate) for candidate in filtered),
        key=lambda pair: pair[0],
        reverse=True,
    )

    if not scored:
        return None
    if len(gtins) == 1 and gtins:
        return scored[0][1]
    if len(scored) == 1:
        return scored[0][1]
    if scored[0][0] >= 0.56 and scored[0][0] - scored[1][0] >= 0.08:
        return scored[0][1]
    return None


TRIDENT_URL_ALIASES = {
    "Hurricane Cat 4": "hurricane-cat-4",
    "Hurricane Cat 5": "hurricane-cat-5",
    "Hurricane Cat 5 1/2": "hurricane-cat-5-half-kit",
    "Tidal Wave Gel": "tidal-wave",
    "Tidal Wave Spray": "tidal-wave",
}


def fetch_trident_product_pages(target_names: set[str]) -> dict[str, dict[str, str]]:
    response = requests.get(TRIDENT_SITEMAP_URL, headers=HTTP_HEADERS, timeout=30)
    response.raise_for_status()
    urls = re.findall(r"<loc>(.*?)</loc>", response.text)
    slug_to_url = {url.rstrip("/").split("/")[-1]: url for url in urls}
    candidate_urls: set[str] = set()
    for target_name in target_names:
        cleaned_name = clean_text(re.sub(r"\s*\([^)]*\)", "", target_name))
        alias_slug = TRIDENT_URL_ALIASES.get(cleaned_name)
        slug = alias_slug or slugify(cleaned_name.replace("1/2", "half"))
        if slug in slug_to_url:
            candidate_urls.add(slug_to_url[slug])

    products: dict[str, dict[str, str]] = {}

    for url in sorted(candidate_urls):
        html = requests.get(url, headers=HTTP_HEADERS, timeout=30).text
        if "html-product-details-page" not in html:
            continue
        soup = BeautifulSoup(html, "html.parser")
        title_node = soup.select_one(".product-name")
        desc_node = soup.select_one(".full-description")
        short_node = soup.select_one(".short-description")
        title = clean_text(title_node.get_text(" ", strip=True) if title_node else "")
        if not title:
            continue
        description = clean_text(desc_node.get_text(" ", strip=True) if desc_node else "")
        short_description = clean_text(short_node.get_text(" ", strip=True) if short_node else "")
        if short_description and short_description.upper() not in description.upper():
            description = build_description(short_description, description)
        products[normalize_name(title)] = {
            "title": title,
            "description": trim_words(description, 1000),
            "url": url,
        }

    return products


def fetch_eacochem_product_pages(target_names: set[str]) -> dict[str, dict[str, str]]:
    response = requests.get(EACOCHEM_ALL_PRODUCTS_URL, headers=HTTP_HEADERS, timeout=30)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    urls: dict[str, tuple[str, str]] = {}
    for link in soup.find_all("a", href=True):
        href = clean_text(link.get("href", ""))
        if "/eaco_products/" not in href:
            continue
        text = clean_text(link.get_text(" ", strip=True))
        key = normalize_name(text)
        if href and key and href not in urls:
            urls[href] = (text, key)

    products: dict[str, dict[str, str]] = {}
    target_keys = {normalize_name(name) for name in target_names if normalize_name(name)}
    candidate_urls = [url for url, (_, key) in urls.items() if key in target_keys]
    for url in candidate_urls:
        html = requests.get(url, headers=HTTP_HEADERS, timeout=30).text
        soup = BeautifulSoup(html, "html.parser")
        title_node = soup.select_one("h1") or soup.select_one(".entry-title")
        if title_node is None:
            continue
        title = clean_text(title_node.get_text(" ", strip=True))
        subtitle_node = soup.select_one("h2")
        subtitle = clean_text(subtitle_node.get_text(" ", strip=True) if subtitle_node else "")
        meta_match = re.search(r'<meta name="description" content="(.*?)"', html, re.I | re.S)
        meta_description = clean_text(meta_match.group(1)) if meta_match else ""
        if not title:
            continue
        description = build_description(subtitle, meta_description)
        products[normalize_name(title)] = {
            "title": title,
            "description": trim_words(description, 1000),
            "url": url,
        }

    return products


def fetch_gold_assassin_products() -> dict[str, dict[str, object]]:
    response = requests.get(GOLD_ASSASSIN_SITEMAP_URL, headers=HTTP_HEADERS, timeout=30)
    response.raise_for_status()
    urls = [html.unescape(url) for url in re.findall(r"<loc>(.*?)</loc>", response.text)]
    products: dict[str, dict[str, object]] = {}

    for url in urls:
        page_html = ""
        for _ in range(3):
            try:
                page_response = requests.get(url, headers=HTTP_HEADERS, timeout=30)
                page_response.raise_for_status()
                page_html = page_response.text
                break
            except Exception:
                continue
        if not page_html:
            continue

        soup = BeautifulSoup(page_html, "html.parser")
        product_data: dict[str, object] | None = None
        for script in soup.select('script[type="application/ld+json"]'):
            try:
                payload = json.loads(script.get_text(strip=True))
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue
            if clean_text(payload.get("@type", "")).lower() != "product":
                continue
            sku = normalize_sku(payload.get("sku", ""))
            if not sku:
                continue
            product_data = payload
            break
        if product_data is None:
            continue

        sku = normalize_sku(product_data.get("sku", ""))
        title = clean_text(product_data.get("name", ""))
        description = trim_words(strip_html_to_text(product_data.get("description", "")), 1000)
        offer_data = product_data.get("Offers") or product_data.get("offers") or {}
        weight_match = re.search(
            rf'"sku":"{re.escape(sku)}","weight":(\d+(?:\.\d+)?)',
            page_html,
            re.I,
        )
        weight_lb = None
        if weight_match:
            try:
                weight_lb = Decimal(weight_match.group(1)).quantize(WEIGHT_PRECISION, rounding=ROUND_HALF_UP)
            except Exception:
                weight_lb = None

        products[sku] = {
            "title": title,
            "description": description,
            "url": clean_text(offer_data.get("url", "")) or url,
            "weight_lb": weight_lb,
            "sku": sku,
        }

    return products


def apply_direct_vendor_enrichments(
    items: list[SourceItem],
) -> tuple[list[EnrichmentAuditEntry], Counter[str], Counter[str], list[str]]:
    audit_entries: list[EnrichmentAuditEntry] = []
    match_counts: Counter[str] = Counter()
    detail_counts: Counter[str] = Counter()
    notes: list[str] = []

    # JRacenstein exact product/variant code matches via the live Storefront GraphQL catalog.
    try:
        session = requests.Session()
        token = fetch_jracenstein_storefront_token(session)
        search_cache: dict[str, list[dict[str, object]]] = {}
        product_cache: dict[str, dict[str, object]] = {}
        unresolved_candidates = 0

        for item in [candidate for candidate in items if candidate.vendor == "JRacenstein"]:
            vendor_code = clean_text(item.vendor_code or item.sku)
            candidates = build_jracenstein_exact_candidates(session, token, vendor_code, search_cache, product_cache)
            resolved = resolve_jracenstein_candidate(item, candidates)
            if resolved is None:
                if len(candidates) > 1:
                    unresolved_candidates += 1
                continue

            item.image_lookup_tokens = unique_nonempty(
                [
                    clean_text(resolved.get("variant_sku", "")),
                    clean_text(resolved.get("product_sku", "")),
                    clean_text(item.vendor_code),
                    clean_text(item.sku),
                ]
            )
            item.image_lookup_slug = path_slug_from_url(clean_text(resolved.get("product_url", "")))

            product_url = clean_text(resolved.get("product_url", ""))
            candidate_name = jracenstein_candidate_name(item, resolved)
            match_detail = (
                f"Matched on exact {resolved['matched_kind']} {resolved['matched_field'].upper()} "
                f"for vendor code {vendor_code}."
            )

            if candidate_name and candidate_name != product_display_name(item):
                item.customer_facing_name_override = candidate_name
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="jracenstein_exact_code_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=vendor_code,
                        item_name=item.item_name,
                        field="Customer-facing Name",
                        value=candidate_name,
                        source=product_url,
                        details=match_detail,
                    )
                )
                detail_counts["customer_names"] += 1

            description_text = clean_text(resolved.get("product_description", ""))
            if description_text and normalize_name(description_text) != normalize_name(product_description_text(item)):
                item.description_override = description_text
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="jracenstein_exact_code_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=vendor_code,
                        item_name=item.item_name,
                        field="Description",
                        value=trim_words(description_text, 120),
                        source=product_url,
                        details="Pulled from the matching JRacenstein product page description.",
                    )
                )
                detail_counts["descriptions"] += 1

            gtin = jracenstein_candidate_gtin(resolved)
            if gtin and gtin != valid_gtin(item.gtin):
                item.gtin = gtin
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="jracenstein_exact_code_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=vendor_code,
                        item_name=item.item_name,
                        field="GTIN",
                        value=gtin,
                        source=product_url,
                        details=match_detail,
                    )
                )
                detail_counts["gtins"] += 1

            weight_lb = jracenstein_candidate_weight(resolved)
            if weight_lb is not None:
                item.weight_lb_override = weight_lb
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="jracenstein_exact_code_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=vendor_code,
                        item_name=item.item_name,
                        field="Weight (lb)",
                        value=format_weight(weight_lb),
                        source=product_url,
                        details="Pulled from the matching JRacenstein product or variant shipping weight.",
                    )
                )
                detail_counts["weights"] += 1

            permalink = build_jracenstein_permalink(item, resolved)
            if permalink:
                item.permalink_override = permalink
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="jracenstein_exact_code_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=vendor_code,
                        item_name=item.item_name,
                        field="Permalink",
                        value=permalink,
                        source=product_url,
                        details="Built from the JRacenstein product path plus the matched variant labels when needed.",
                    )
                )
                detail_counts["permalinks"] += 1

            match_counts["JRacenstein"] += 1

        note = f"JRacenstein website enrichment matched {match_counts['JRacenstein']} exact code rows."
        if unresolved_candidates:
            note += f" Left {unresolved_candidates} ambiguous exact-code candidates unchanged."
        notes.append(note)
    except Exception as exc:
        notes.append(f"JRacenstein website enrichment skipped: {exc}")

    # Trident exact base-title matches against product pages.
    try:
        trident_items = [candidate for candidate in items if candidate.vendor == "Trident"]
        trident_products = fetch_trident_product_pages({split_variant_suffix(item.item_name)[0] for item in trident_items})
        for item in trident_items:
            base_name, pack_suffix = split_variant_suffix(item.item_name)
            key = normalize_name(base_name)
            if key not in trident_products:
                continue
            product = trident_products[key]
            item.image_lookup_slug = path_slug_from_url(product["url"])
            customer_name = combine_title_with_suffix(product["title"], pack_suffix)
            if customer_name and customer_name != product_display_name(item):
                item.customer_facing_name_override = customer_name
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="trident_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Customer-facing Name",
                        value=customer_name,
                        source=product["url"],
                        details="Matched on exact base product title from the Trident site.",
                    )
                )
                detail_counts["customer_names"] += 1
            description = product["description"]
            if description and normalize_name(description) != normalize_name(product_description_text(item)):
                item.description_override = description
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="trident_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Description",
                        value=trim_words(description, 120),
                        source=product["url"],
                        details="Pulled from the Trident product page description.",
                    )
                )
                detail_counts["descriptions"] += 1
            slug = slugify(customer_name)
            if slug:
                item.permalink_override = slug
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="trident_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Permalink",
                        value=slug,
                        source=product["url"],
                        details="Using a slug derived from the Trident product title and pack size.",
                    )
                )
                detail_counts["permalinks"] += 1
            match_counts["Trident"] += 1
        notes.append(f"Trident website enrichment matched {match_counts['Trident']} catalog rows.")
    except Exception as exc:
        notes.append(f"Trident website enrichment skipped: {exc}")

    # EacoChem exact base-title matches against product pages.
    try:
        eacochem_items = [candidate for candidate in items if candidate.vendor == "EacoChem"]
        eacochem_products = fetch_eacochem_product_pages({split_variant_suffix(item.item_name)[0] for item in eacochem_items})
        for item in eacochem_items:
            base_name, pack_suffix = split_variant_suffix(item.item_name)
            key = normalize_name(base_name)
            if key not in eacochem_products:
                continue
            product = eacochem_products[key]
            customer_name = combine_title_with_suffix(product["title"], pack_suffix)
            if customer_name and customer_name != product_display_name(item):
                item.customer_facing_name_override = customer_name
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="eacochem_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Customer-facing Name",
                        value=customer_name,
                        source=product["url"],
                        details="Matched on exact base product title from the EaCo Chem site.",
                    )
                )
                detail_counts["customer_names"] += 1
            description = product["description"]
            if description and normalize_name(description) != normalize_name(product_description_text(item)):
                item.description_override = description
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="eacochem_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Description",
                        value=trim_words(description, 120),
                        source=product["url"],
                        details="Pulled from the EaCo Chem product page description.",
                    )
                )
                detail_counts["descriptions"] += 1
            slug = slugify(customer_name)
            if slug:
                item.permalink_override = slug
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="eacochem_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Permalink",
                        value=slug,
                        source=product["url"],
                        details="Using a slug derived from the EaCo Chem product title and pack size.",
                    )
                )
                detail_counts["permalinks"] += 1
            match_counts["EacoChem"] += 1
        notes.append(f"EacoChem website enrichment matched {match_counts['EacoChem']} catalog rows.")
    except Exception as exc:
        notes.append(f"EacoChem website enrichment skipped: {exc}")

    # Gold Assassin exact SKU matches against the manufacturer site.
    try:
        gold_products = fetch_gold_assassin_products()
        gold_items = [
            candidate
            for candidate in items
            if candidate.vendor == "MPWSR"
            and normalize_sku(candidate.vendor_code or candidate.sku) in gold_products
        ]
        skipped_weight_conflicts = 0

        for item in gold_items:
            gold_product = gold_products[normalize_sku(item.vendor_code or item.sku)]
            item.image_lookup_slug = path_slug_from_url(clean_text(gold_product.get("url", "")))
            customer_name = clean_text(gold_product.get("title", ""))
            product_url = clean_text(gold_product.get("url", ""))
            if customer_name and customer_name != product_display_name(item):
                item.customer_facing_name_override = customer_name
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="gold_assassin_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Customer-facing Name",
                        value=customer_name,
                        source=product_url,
                        details="Matched on exact SKU from the Gold Assassin manufacturer site.",
                    )
                )
                detail_counts["customer_names"] += 1

            description = clean_text(gold_product.get("description", ""))
            if description and normalize_name(description) != normalize_name(product_description_text(item)):
                item.description_override = description
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="gold_assassin_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Description",
                        value=trim_words(description, 120),
                        source=product_url,
                        details="Pulled from the Gold Assassin product page description.",
                    )
                )
                detail_counts["descriptions"] += 1

            permalink = slugify(customer_name)
            if permalink:
                item.permalink_override = permalink
                audit_entries.append(
                    EnrichmentAuditEntry(
                        enrichment_type="gold_assassin_direct_match",
                        vendor=item.vendor,
                        sku=item.sku,
                        vendor_code=clean_text(item.vendor_code or item.sku),
                        item_name=item.item_name,
                        field="Permalink",
                        value=permalink,
                        source=product_url,
                        details="Using a permalink derived from the Gold Assassin product title.",
                    )
                )
                detail_counts["permalinks"] += 1

            weight_lb = gold_product.get("weight_lb")
            if isinstance(weight_lb, Decimal):
                current_weight = item.weight_lb_override
                if current_weight is None or abs(current_weight - weight_lb) <= Decimal("1.0"):
                    item.weight_lb_override = weight_lb
                    audit_entries.append(
                        EnrichmentAuditEntry(
                            enrichment_type="gold_assassin_direct_match",
                            vendor=item.vendor,
                            sku=item.sku,
                            vendor_code=clean_text(item.vendor_code or item.sku),
                            item_name=item.item_name,
                            field="Weight (lb)",
                            value=format_weight(weight_lb),
                            source=product_url,
                            details="Pulled from the Gold Assassin product page item weight.",
                        )
                    )
                    detail_counts["weights"] += 1
                else:
                    skipped_weight_conflicts += 1

            match_counts["GoldAssassin"] += 1

        note = f"Gold Assassin website enrichment matched {match_counts['GoldAssassin']} exact SKU rows."
        if skipped_weight_conflicts:
            note += f" Left {skipped_weight_conflicts} conflicting existing weights unchanged."
        notes.append(note)
    except Exception as exc:
        notes.append(f"Gold Assassin website enrichment skipped: {exc}")

    return audit_entries, match_counts, detail_counts, notes


def set_item_image_match(
    item: SourceItem,
    image_path: Path,
    match_type: str,
    source_folder: str,
) -> None:
    item.image_relative_path = relative_repo_path(image_path)
    item.image_absolute_path = str(image_path.resolve())
    item.image_match_type = clean_text(match_type)
    item.image_source_folder = clean_text(source_folder)


def build_unique_path_index(paths: list[Path], key_builder) -> dict[str, Path]:
    grouped: dict[str, list[Path]] = defaultdict(list)
    for path in paths:
        if not path.is_file():
            continue
        key = clean_text(key_builder(path))
        if not key:
            continue
        grouped[key].append(path)
    return {key: values[0] for key, values in grouped.items() if len(values) == 1}


def apply_local_image_matches(
    items: list[SourceItem],
) -> tuple[list[ImageMatchAuditEntry], Counter[str], list[str]]:
    audit_entries: list[ImageMatchAuditEntry] = []
    match_counts: Counter[str] = Counter()
    notes: list[str] = []

    jr_files = list(JRACENSTEIN_IMAGE_DIR.glob("*")) if JRACENSTEIN_IMAGE_DIR.exists() else []
    trident_files = list(TRIDENT_IMAGE_DIR.glob("*")) if TRIDENT_IMAGE_DIR.exists() else []
    manatee_files = list(MANATEE_IMAGE_DIR.glob("*")) if MANATEE_IMAGE_DIR.exists() else []

    jr_index = build_unique_path_index(jr_files, lambda path: normalize_sku(path.stem))
    trident_index = build_unique_path_index(trident_files, lambda path: local_image_stem_key(path.stem))
    manatee_code_index = build_unique_path_index(manatee_files, lambda path: normalize_sku(path.stem))
    manatee_name_index = build_unique_path_index(manatee_files, lambda path: normalize_name(path.stem))

    for item in items:
        image_path: Path | None = None
        match_type = ""
        source_folder = ""
        detail = ""

        if item.reporting_category == "JRacenstein":
            for token in unique_nonempty(item.image_lookup_tokens + [item.vendor_code, item.sku]):
                key = normalize_sku(token)
                if key in jr_index:
                    image_path = jr_index[key]
                    match_type = "jracenstein_image_code"
                    source_folder = "J.racenstein images"
                    detail = f"Matched local image filename on exact JR code token {token}."
                    break

        elif item.reporting_category == "Trident":
            trident_keys = unique_nonempty(
                [
                    item.image_lookup_slug,
                    item.permalink_override,
                    slugify(product_display_name(item)),
                    slugify(item.item_name),
                ]
            )
            for token in trident_keys:
                key = local_image_stem_key(token)
                if key in trident_index:
                    image_path = trident_index[key]
                    match_type = "trident_image_slug"
                    source_folder = "Trident images/media"
                    detail = f"Matched local Trident image on slug {token}."
                    break

        else:
            for token in unique_nonempty([item.vendor_code, item.sku]):
                key = normalize_sku(token)
                if key in manatee_code_index:
                    image_path = manatee_code_index[key]
                    match_type = "manatee_exact_code"
                    source_folder = "Manatee images"
                    detail = f"Matched local image filename on exact code token {token}."
                    break

            if image_path is None:
                candidate_paths: list[tuple[Path, str, str]] = []
                for label in unique_nonempty([product_display_name(item), item.item_name]):
                    name_key = normalize_name(label)
                    if name_key in manatee_name_index:
                        candidate_paths.append((manatee_name_index[name_key], "manatee_exact_name", label))
                unique_paths = {path.resolve(): (path, match_type_value, label) for path, match_type_value, label in candidate_paths}
                if len(unique_paths) == 1:
                    image_path, match_type, label = next(iter(unique_paths.values()))
                    source_folder = "Manatee images"
                    detail = f"Matched local image on exact normalized name from {label}."

        if image_path is None:
            continue

        set_item_image_match(item, image_path, match_type, source_folder)
        audit_entries.append(
            ImageMatchAuditEntry(
                vendor=item.vendor,
                sku=item.sku,
                vendor_code=clean_text(item.vendor_code or item.sku),
                item_name=item.item_name,
                image_relative_path=item.image_relative_path,
                image_absolute_path=item.image_absolute_path,
                match_type=match_type,
                source_folder=source_folder,
                details=detail,
            )
        )
        match_counts[source_folder] += 1

    if jr_files:
        notes.append(f"JR local image matches applied: {match_counts['J.racenstein images']}.")
    if trident_files:
        notes.append(f"Trident local image matches applied: {match_counts['Trident images/media']}.")
    if manatee_files:
        notes.append(f"Manatee local image matches applied: {match_counts['Manatee images']}.")
    return audit_entries, match_counts, notes


def item_lookup_tokens(item: SourceItem) -> set[str]:
    tokens: set[str] = set()

    def maybe_add(token: str) -> None:
        candidate = clean_text(token).upper().strip(".,;:()[]{}")
        if len(candidate) < 5:
            return
        if not any(char.isdigit() for char in candidate):
            return
        if candidate.isdigit() and len(candidate) < 6:
            return
        if re.fullmatch(r"\d+(?:\.\d+)?[A-Z]+", candidate):
            return
        tokens.add(candidate)

    maybe_add(item.sku)
    maybe_add(item.vendor_code)
    first_word = clean_text(item.item_name).split(" ", 1)[0]
    maybe_add(first_word)
    return tokens


def apply_verified_enrichments(
    items: list[SourceItem],
    enrichments: dict[tuple[str, str, str], dict[str, str]],
) -> list[EnrichmentAuditEntry]:
    audit_entries: list[EnrichmentAuditEntry] = []

    for item in items:
        keys = [
            (item.vendor.upper(), "vendor_code", clean_text(item.vendor_code).upper()),
            (item.vendor.upper(), "sku", clean_text(item.sku).upper()),
            (item.vendor.upper(), "item_name", clean_text(item.item_name).upper()),
        ]
        enrichment = next((enrichments[key] for key in keys if key in enrichments), None)
        if enrichment is None:
            continue

        gtin = valid_gtin(enrichment.get("gtin", ""))
        if gtin and gtin != valid_gtin(item.gtin):
            item.gtin = gtin
            audit_entries.append(
                EnrichmentAuditEntry(
                    enrichment_type="verified_override",
                    vendor=item.vendor,
                    sku=item.sku,
                    vendor_code=clean_text(item.vendor_code or item.sku),
                    item_name=item.item_name,
                    field="GTIN",
                    value=gtin,
                    source=enrichment.get("source_url", ""),
                    details=enrichment.get("notes", ""),
                )
            )

        if clean_text(enrichment.get("seo_title", "")):
            item.seo_title_override = enrichment["seo_title"]
        if clean_text(enrichment.get("seo_description", "")):
            item.seo_description_override = enrichment["seo_description"]

    return audit_entries


def infer_missing_gtins_from_catalog(items: list[SourceItem]) -> list[EnrichmentAuditEntry]:
    token_to_gtins: dict[str, set[str]] = defaultdict(set)
    token_to_sources: dict[str, list[SourceItem]] = defaultdict(list)

    for item in items:
        gtin = valid_gtin(item.gtin)
        if not gtin:
            continue
        for token in item_lookup_tokens(item):
            token_to_gtins[token].add(gtin)
            token_to_sources[token].append(item)

    unique_token_map = {
        token: next(iter(gtins))
        for token, gtins in token_to_gtins.items()
        if len(gtins) == 1
    }

    audit_entries: list[EnrichmentAuditEntry] = []
    for item in items:
        if valid_gtin(item.gtin):
            continue

        matching_tokens = [token for token in item_lookup_tokens(item) if token in unique_token_map]
        unique_gtins = {unique_token_map[token] for token in matching_tokens}
        if len(unique_gtins) != 1 or not matching_tokens:
            continue

        gtin = next(iter(unique_gtins))
        chosen_token = sorted(matching_tokens, key=len, reverse=True)[0]
        source_item = next(
            source
            for source in token_to_sources[chosen_token]
            if valid_gtin(source.gtin) == gtin
        )
        item.gtin = gtin
        audit_entries.append(
            EnrichmentAuditEntry(
                enrichment_type="catalog_cross_reference",
                vendor=item.vendor,
                sku=item.sku,
                vendor_code=clean_text(item.vendor_code or item.sku),
                item_name=item.item_name,
                field="GTIN",
                value=gtin,
                source=source_item.vendor,
                details=(
                    f"Matched on token {chosen_token} from {source_item.vendor} "
                    f"SKU {clean_text(source_item.sku)} ({clean_text(source_item.item_name)})."
                ),
            )
        )

    return audit_entries


def dedupe_same_source(items: list[SourceItem]) -> tuple[list[SourceItem], int]:
    seen: set[tuple[str, ...]] = set()
    kept: list[SourceItem] = []
    skipped = 0
    for item in items:
        key = (
            item.source_file,
            normalize_sku(item.sku),
            valid_gtin(item.gtin),
            normalize_name(item.item_name),
            format_money(item.default_unit_cost),
            format_money(item.price),
        )
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        kept.append(item)
    return kept, skipped


def item_priority_key(item: SourceItem) -> tuple[int, Decimal, int, Decimal, int, int]:
    return (
        1 if item.default_unit_cost is not None else 0,
        item.default_unit_cost or Decimal("-1"),
        1 if item.price is not None else 0,
        item.price or Decimal("-1"),
        1 if valid_gtin(item.gtin) else 0,
        1 if normalize_sku(item.sku) else 0,
    )


def merge_group_items(group_items: list[SourceItem]) -> SourceItem:
    best_item = max(group_items, key=item_priority_key)
    merged_item = replace(best_item, notes=list(best_item.notes))

    available_costs = [item.default_unit_cost for item in group_items if item.default_unit_cost is not None]
    available_prices = [item.price for item in group_items if item.price is not None]
    highest_cost = max(available_costs) if available_costs else None
    prices_at_highest_cost = [
        item.price
        for item in group_items
        if item.default_unit_cost == highest_cost and item.price is not None
    ]

    merged_item.default_unit_cost = highest_cost
    if prices_at_highest_cost:
        merged_item.price = max(prices_at_highest_cost)
    elif merged_item.price is None and available_prices:
        merged_item.price = max(available_prices)

    if not valid_gtin(merged_item.gtin):
        merged_item.gtin = next((item.gtin for item in group_items if valid_gtin(item.gtin)), "")
    if not normalize_sku(merged_item.sku):
        merged_item.sku = next((item.sku for item in group_items if normalize_sku(item.sku)), "")
    if not clean_text(merged_item.vendor_code):
        merged_item.vendor_code = next((item.vendor_code for item in group_items if clean_text(item.vendor_code)), merged_item.sku)
    if not clean_text(merged_item.description):
        merged_item.description = max((item.description for item in group_items), key=len, default="")
    if not clean_text(merged_item.category):
        merged_item.category = next((item.category for item in group_items if clean_text(item.category)), "")
    if not clean_text(merged_item.reporting_category):
        merged_item.reporting_category = next((item.reporting_category for item in group_items if clean_text(item.reporting_category)), "")

    return merged_item


def extract_case_pack(description: str) -> str:
    match = re.search(r"Case pack:\s*([^|]+)", description)
    return clean_text(match.group(1)) if match else ""


def disambiguation_label(item: SourceItem) -> str:
    case_pack = extract_case_pack(item.description)
    if case_pack:
        return f"Case {case_pack}"
    gtin = valid_gtin(item.gtin)
    if gtin:
        return f"GTIN {gtin}"
    sku = clean_text(item.sku)
    if sku:
        return f"SKU {sku}"
    return clean_text(item.source_file)


def merge_duplicate_items(items: list[SourceItem]) -> tuple[list[SourceItem], list[ReviewIssue], int]:
    parent = list(range(len(items)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    def union_groups(groups: dict[str, list[int]]) -> None:
        for indexes in groups.values():
            unique_indexes = sorted(set(indexes))
            if len(unique_indexes) <= 1:
                continue
            first = unique_indexes[0]
            for index in unique_indexes[1:]:
                union(first, index)

    def union_name_groups(groups: dict[str, list[int]]) -> None:
        for indexes in groups.values():
            unique_indexes = sorted(set(indexes))
            if len(unique_indexes) <= 1:
                continue
            for left_pos, left in enumerate(unique_indexes):
                for right in unique_indexes[left_pos + 1 :]:
                    if items[left].vendor != items[right].vendor:
                        union(left, right)

    sku_groups: dict[str, list[int]] = defaultdict(list)
    gtin_groups: dict[str, list[int]] = defaultdict(list)
    name_groups: dict[str, list[int]] = defaultdict(list)

    for index, item in enumerate(items):
        sku_key = normalize_sku(item.sku)
        if sku_key:
            sku_groups[f"{item.vendor}|{sku_key}"].append(index)
        gtin_key = valid_gtin(item.gtin)
        if gtin_key:
            gtin_groups[gtin_key].append(index)
        name_key = normalize_name(item.item_name)
        if name_key and len(name_key) >= 8:
            name_groups[name_key].append(index)

    union_groups(sku_groups)
    union_groups(gtin_groups)
    union_name_groups(name_groups)

    grouped_indexes: dict[int, list[int]] = defaultdict(list)
    for index in range(len(items)):
        grouped_indexes[find(index)].append(index)

    merged_items: list[SourceItem] = []
    merge_issues: list[ReviewIssue] = []
    merged_groups = 0

    for root in sorted(grouped_indexes):
        indexes = grouped_indexes[root]
        group_items = [items[index] for index in indexes]
        if len(group_items) == 1:
            merged_items.append(group_items[0])
            continue

        merged_groups += 1
        merged_item = merge_group_items(group_items)
        highest_cost = merged_item.default_unit_cost

        reason_parts: list[str] = []
        gtin_matches = {valid_gtin(item.gtin) for item in group_items if valid_gtin(item.gtin)}
        name_matches = {normalize_name(item.item_name) for item in group_items if normalize_name(item.item_name)}
        if len(gtin_matches) == 1 and next(iter(gtin_matches), ""):
            reason_parts.append("exact GTIN match")
        if len(name_matches) == 1 and next(iter(name_matches), ""):
            reason_parts.append("exact item-name match")
        vendor_sku_pairs = {(item.vendor, normalize_sku(item.sku)) for item in group_items if normalize_sku(item.sku)}
        if len(vendor_sku_pairs) < len([item for item in group_items if normalize_sku(item.sku)]):
            reason_parts.append("same-vendor SKU duplicate")

        source_list = "; ".join(
            f"{item.vendor} | {item.source_file} | SKU={item.sku or '[generated]'} | Cost={format_money(item.default_unit_cost)}"
            for item in sorted(group_items, key=lambda item: (item.vendor, item.source_file, item.sku, item.item_name))
        )
        reason_text = ", ".join(reason_parts) if reason_parts else "duplicate item match"
        merge_issues.append(
            ReviewIssue(
                issue_type="merged_duplicate",
                vendor=merged_item.vendor,
                source_file=merged_item.source_file,
                item_name=merged_item.item_name,
                sku=merged_item.sku,
                gtin=merged_item.gtin,
                category=merged_item.category,
                default_unit_cost=format_money(merged_item.default_unit_cost),
                price=format_money(merged_item.price),
                details=f"Merged {len(group_items)} source rows by {reason_text}. Kept the highest cost option {format_money(highest_cost)}. Sources: {source_list}",
            )
        )
        merged_items.append(merged_item)

    return merged_items, merge_issues, merged_groups


def resolve_same_vendor_name_collisions(items: list[SourceItem]) -> tuple[list[SourceItem], list[ReviewIssue], int, int]:
    groups: dict[tuple[str, str], list[SourceItem]] = defaultdict(list)
    for item in items:
        groups[(item.vendor, normalize_name(item.item_name))].append(item)

    resolved_items: list[SourceItem] = []
    review_issues: list[ReviewIssue] = []
    merged_groups = 0
    renamed_rows = 0

    for (vendor, name_key), group_items in groups.items():
        if not name_key or len(group_items) == 1:
            resolved_items.extend(group_items)
            continue

        costs = [item.default_unit_cost for item in group_items if item.default_unit_cost is not None]
        safe_to_merge = False
        if costs:
            min_cost = min(costs)
            max_cost = max(costs)
            safe_to_merge = bool(min_cost) and max_cost <= min_cost * Decimal("1.5")
        gtins = {valid_gtin(item.gtin) for item in group_items if valid_gtin(item.gtin)}
        if len(gtins) == 1 and next(iter(gtins), ""):
            safe_to_merge = True

        if safe_to_merge:
            merged_groups += 1
            merged_item = merge_group_items(group_items)
            resolved_items.append(merged_item)
            review_issues.append(
                ReviewIssue(
                    issue_type="merged_same_vendor_duplicate",
                    vendor=merged_item.vendor,
                    source_file=merged_item.source_file,
                    item_name=merged_item.item_name,
                    sku=merged_item.sku,
                    gtin=merged_item.gtin,
                    category=merged_item.category,
                    default_unit_cost=format_money(merged_item.default_unit_cost),
                    price=format_money(merged_item.price),
                    details="Merged same-vendor duplicate names and kept the higher cost option. Sources: "
                    + "; ".join(
                        f"SKU={item.sku or '[generated]'} | Cost={format_money(item.default_unit_cost)}"
                        for item in sorted(group_items, key=lambda item: (item.sku, item.item_name))
                    ),
                )
            )
            continue

        used_labels: set[str] = set()
        for item in group_items:
            renamed_item = replace(item, notes=list(item.notes))
            label = disambiguation_label(renamed_item)
            candidate = label
            suffix = 2
            while candidate in used_labels:
                candidate = f"{label} #{suffix}"
                suffix += 1
            used_labels.add(candidate)
            renamed_item.item_name = f"{clean_text(renamed_item.item_name)} [{candidate}]"
            resolved_items.append(renamed_item)
            renamed_rows += 1

        review_issues.append(
            ReviewIssue(
                issue_type="renamed_same_vendor_duplicate",
                vendor=vendor,
                source_file=group_items[0].source_file,
                item_name=group_items[0].item_name,
                category=group_items[0].category,
                details="Kept separate items with the same vendor/name by renaming them using case pack, GTIN, or SKU.",
            )
        )

    return resolved_items, review_issues, merged_groups, renamed_rows


def generate_unique_skus(items: list[SourceItem]) -> int:
    used: set[str] = set()
    updated = 0

    for item in items:
        sku = clean_code(item.sku)
        if sku and sku not in used:
            item.sku = sku
            used.add(sku)
            continue

        prefix = re.sub(r"[^A-Z0-9]+", "", item.vendor.upper())[:6] or "ITEM"
        base = re.sub(r"[^A-Z0-9]+", "-", clean_text(item.item_name).upper()).strip("-")
        base = re.sub(r"-{2,}", "-", base)
        base = base[:32].strip("-") or "PRODUCT"

        counter = 1
        candidate = f"{prefix}-{base}"
        while candidate in used:
            counter += 1
            suffix = f"-{counter}"
            shortened = base[: max(8, 32 - len(suffix))].rstrip("-") or "PRODUCT"
            candidate = f"{prefix}-{shortened}{suffix}"

        item.sku = candidate
        item.generated_sku = True
        used.add(candidate)
        updated += 1

    return updated


def build_square_row(item: SourceItem, fieldnames: list[str]) -> dict[str, str]:
    row = {field: "" for field in fieldnames}
    customer_facing_name = product_display_name(item)
    description_text = product_description_text(item)
    seo_title = build_seo_title(item)
    seo_description = build_seo_description(item, seo_title)
    row["Token"] = ""
    row["Item Name"] = clean_text(item.item_name)
    row["Customer-facing Name"] = customer_facing_name
    row["Variation Name"] = "Regular"
    row["SKU"] = item.sku
    row["Description"] = description_text
    row["Categories"] = clean_text(item.category)
    row["Reporting Category"] = clean_text(item.reporting_category)
    row["SEO Title"] = seo_title
    row["SEO Description"] = seo_description
    row["GTIN"] = valid_gtin(item.gtin)
    row["Square Online Item Visibility"] = "Hidden"
    row["Item Type"] = "Physical"
    row["Weight (lb)"] = format_weight(item.weight_lb_override)
    row["Social Media Link Title"] = seo_title
    row["Social Media Link Description"] = seo_description
    row["Shipping Enabled"] = "N"
    row["Self-serve Ordering Enabled"] = "N"
    row["Delivery Enabled"] = "N"
    row["Pickup Enabled"] = "N"
    row["Price"] = format_money(item.price)
    row["Archived"] = "N"
    row["Sellable"] = "Y" if item.price is not None else "N"
    row["Contains Alcohol"] = "N"
    row["Stockable"] = "Y"
    row["Skip Detail Screen in POS"] = "N"
    row["Default Unit Cost"] = format_money(item.default_unit_cost)
    row["Default Vendor Name"] = item.vendor
    row["Default Vendor Code"] = clean_text(item.vendor_code or item.sku)
    row["Permalink"] = clean_text(item.permalink_override)
    return row


def assign_unique_permalinks(rows: list[dict[str, str]]) -> int:
    used: dict[str, int] = {}
    assigned = 0

    for row in rows:
        base = slugify(row.get("Permalink") or row.get("Customer-facing Name") or row.get("Item Name") or row.get("SKU"))
        if not base:
            base = slugify(row.get("SKU") or "item")
        candidate = base
        if candidate in used:
            sku_suffix = slugify(row.get("SKU", ""))
            if sku_suffix:
                candidate = f"{base}-{sku_suffix}"
        counter = 2
        while candidate in used:
            candidate = f"{base}-{counter}"
            counter += 1
        used[candidate] = 1
        row["Permalink"] = candidate
        assigned += 1

    return assigned


def write_master_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_review_rows(review_issues: list[ReviewIssue]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    for issue in review_issues:
        rows.append(
            {
                "issue_type": issue.issue_type,
                "vendor": issue.vendor,
                "source_file": issue.source_file,
                "item_name": clean_text(issue.item_name),
                "sku": clean_text(issue.sku),
                "gtin": valid_gtin(issue.gtin),
                "category": clean_text(issue.category),
                "default_unit_cost": issue.default_unit_cost,
                "price": issue.price,
                "details": clean_text(issue.details),
            }
        )

    return rows


def write_review_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "issue_type",
        "vendor",
        "source_file",
        "item_name",
        "sku",
        "gtin",
        "category",
        "default_unit_cost",
        "price",
        "details",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_image_database_rows(
    items: list[SourceItem],
    master_rows: list[dict[str, str]],
) -> tuple[list[str], list[dict[str, str]]]:
    image_columns = [
        "Local Image Relative Path",
        "Local Image Absolute Path",
        "Local Image Match Type",
        "Local Image Source Folder",
        "Has Local Image",
    ]
    fieldnames = list(master_rows[0].keys()) + image_columns if master_rows else image_columns
    rows: list[dict[str, str]] = []
    for item, master_row in zip(items, master_rows):
        row = dict(master_row)
        row["Local Image Relative Path"] = clean_text(item.image_relative_path)
        row["Local Image Absolute Path"] = clean_text(item.image_absolute_path)
        row["Local Image Match Type"] = clean_text(item.image_match_type)
        row["Local Image Source Folder"] = clean_text(item.image_source_folder)
        row["Has Local Image"] = "Y" if clean_text(item.image_relative_path) else "N"
        rows.append(row)
    return fieldnames, rows


def write_image_database_xlsx(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    sheet.freeze_panes = "A2"
    workbook.save(path)


def write_enrichment_audit_csv(path: Path, entries: list[EnrichmentAuditEntry]) -> None:
    fieldnames = [
        "enrichment_type",
        "vendor",
        "sku",
        "vendor_code",
        "item_name",
        "field",
        "value",
        "source",
        "details",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for entry in entries:
            writer.writerow(
                {
                    "enrichment_type": entry.enrichment_type,
                    "vendor": clean_text(entry.vendor),
                    "sku": clean_text(entry.sku),
                    "vendor_code": clean_text(entry.vendor_code),
                    "item_name": clean_text(entry.item_name),
                    "field": clean_text(entry.field),
                    "value": clean_text(entry.value),
                    "source": clean_text(entry.source),
                    "details": clean_text(entry.details),
                }
            )


def write_image_match_audit_csv(path: Path, entries: list[ImageMatchAuditEntry]) -> None:
    fieldnames = [
        "vendor",
        "sku",
        "vendor_code",
        "item_name",
        "image_relative_path",
        "image_absolute_path",
        "match_type",
        "source_folder",
        "details",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for entry in entries:
            writer.writerow(
                {
                    "vendor": clean_text(entry.vendor),
                    "sku": clean_text(entry.sku),
                    "vendor_code": clean_text(entry.vendor_code),
                    "item_name": clean_text(entry.item_name),
                    "image_relative_path": clean_text(entry.image_relative_path),
                    "image_absolute_path": clean_text(entry.image_absolute_path),
                    "match_type": clean_text(entry.match_type),
                    "source_folder": clean_text(entry.source_folder),
                    "details": clean_text(entry.details),
                }
            )


SOURCE_DEFINITIONS = [
    (["*Barrens*Pricelist*.csv"], parse_barrens),
    (["*MPWSR*Price List*.csv"], parse_mpwsr),
    (["*Dealer Pricing*.xlsx"], parse_inseco),
    (["*Price List - Distributors*.xlsx"], parse_jracenstein),
    (["*BE*PriceList*.pdf"], parse_be),
    (["*Trident*Dealer Price Sheet*.pdf"], parse_trident),
    (["*Distributor New Construction*Pricing.pdf"], parse_eaco_new_construction),
    (["*Distr Fleet Distributor Fleet Distributor*.pdf"], parse_eaco_fleet),
]


def summarize(
    counts_by_vendor: Counter[str],
    total_source_items: int,
    review_rows: list[dict[str, str]],
    included_rows: int,
    generated_skus: int,
    skipped_duplicates: int,
    merged_groups: int,
    renamed_rows: int,
    retained_gtins: int,
    verified_gtins_added: int,
    catalog_gtins_added: int,
    missing_gtins: int,
    seo_titles_generated: int,
    permalinks_generated: int,
    website_match_counts: Counter[str],
    website_detail_counts: Counter[str],
    website_notes: list[str],
    image_match_counts: Counter[str],
    image_notes: list[str],
    image_rows_matched: int,
) -> str:
    lines = [
        f"Square master inventory: {MASTER_OUT_PATH}",
        f"Overlap review file: {REVIEW_OUT_PATH}",
        f"Enrichment audit file: {ENRICHMENT_AUDIT_OUT_PATH}",
        f"Source items normalized: {total_source_items}",
        f"Rows included in Square import: {included_rows}",
        f"Rows sent to review: {len(review_rows)}",
        f"Duplicate item groups merged into one inventory row: {merged_groups}",
        f"Rows renamed to avoid duplicate item names: {renamed_rows}",
        f"Generated replacement SKUs: {generated_skus}",
        f"Duplicate rows skipped inside the same source file: {skipped_duplicates}",
        f"GTINs retained from source files: {retained_gtins}",
        f"GTINs added from verified overrides: {verified_gtins_added}",
        f"GTINs added from catalog cross-reference: {catalog_gtins_added}",
        f"Rows still missing GTIN: {missing_gtins}",
        f"SEO titles generated: {seo_titles_generated}",
        f"Permalinks generated: {permalinks_generated}",
        f"Vendor website row matches applied: {sum(website_match_counts.values())}",
        f"Vendor website GTINs applied: {website_detail_counts['gtins']}",
        f"Vendor website descriptions applied: {website_detail_counts['descriptions']}",
        f"Vendor website weights applied: {website_detail_counts['weights']}",
        f"Rows with local product images matched: {image_rows_matched}",
        "Counts by vendor:",
    ]
    for vendor in sorted(counts_by_vendor):
        lines.append(f"  {vendor}: {counts_by_vendor[vendor]}")
    if website_match_counts:
        lines.append("Vendor website matches:")
        for vendor in sorted(website_match_counts):
            lines.append(f"  {vendor}: {website_match_counts[vendor]}")
    lines.append("Notes:")
    lines.append("  - Default Unit Cost uses dealer/distributor pricing when available.")
    lines.append("  - Price uses list/direct/retail pricing when the source file provided it.")
    lines.append("  - Items without a selling price were imported as Stockable=Y and Sellable=N.")
    lines.append("  - GTIN values only populate when they pass a checksum check or come from a verified manual override.")
    lines.append("  - SEO fields and permalinks are generated automatically from the cleaned catalog data.")
    lines.append("  - Local product images are tracked in the internal image database export, not the Square import file.")
    lines.append("  - MPWSR and Barens website enrichments only apply on exact SKU matches or unique exact-title matches.")
    lines.append("  - JRacenstein matches use exact live product or variant code matches from the Storefront catalog, with size/name checks only to break ties. Trident and EacoChem matches use exact base product names plus the existing pack suffix.")
    for note in website_notes:
        lines.append(f"  - {note}")
    for folder in sorted(image_match_counts):
        lines.append(f"  - Local image matches from {folder}: {image_match_counts[folder]}")
    for note in image_notes:
        lines.append(f"  - {note}")
    lines.append("  - EacoChem Price List.pdf was not used because it duplicates the cleaner EacoChem source sheets.")
    return "\n".join(lines)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    square_headers = load_square_headers(resolve_template_path())
    verified_enrichments = load_verified_enrichments(VERIFIED_ENRICHMENT_PATH)

    source_items: list[SourceItem] = []
    parser_issues: list[ReviewIssue] = []

    for patterns, parser in SOURCE_DEFINITIONS:
        path = resolve_latest_source(patterns)
        items, issues = parser(path)
        source_items.extend(items)
        parser_issues.extend(issues)

    source_items, skipped_duplicates = dedupe_same_source(source_items)
    source_items, merge_issues, merged_groups = merge_duplicate_items(source_items)
    source_items, rename_issues, same_vendor_merged_groups, renamed_rows = resolve_same_vendor_name_collisions(source_items)
    retained_gtins = sum(1 for item in source_items if valid_gtin(item.gtin))
    verified_audit_entries = apply_verified_enrichments(source_items, verified_enrichments)
    catalog_audit_entries = infer_missing_gtins_from_catalog(source_items)
    counts_by_vendor = Counter(item.vendor for item in source_items)
    generated_skus = generate_unique_skus(source_items)
    website_audit_entries, website_match_counts, website_detail_counts, website_notes = apply_shopify_vendor_enrichments(source_items)
    direct_audit_entries, direct_match_counts, direct_detail_counts, direct_notes = apply_direct_vendor_enrichments(source_items)
    website_match_counts.update(direct_match_counts)
    website_detail_counts.update(direct_detail_counts)
    website_notes.extend(direct_notes)
    image_audit_entries, image_match_counts, image_notes = apply_local_image_matches(source_items)

    master_rows = [build_square_row(item, square_headers) for item in source_items]
    permalinks_generated = assign_unique_permalinks(master_rows)
    image_database_fieldnames, image_database_rows = build_image_database_rows(source_items, master_rows)
    review_rows = build_review_rows(merge_issues + rename_issues + parser_issues)
    write_enrichment_audit_csv(
        ENRICHMENT_AUDIT_OUT_PATH,
        verified_audit_entries + catalog_audit_entries + website_audit_entries + direct_audit_entries,
    )
    write_image_match_audit_csv(IMAGE_AUDIT_OUT_PATH, image_audit_entries)

    write_master_csv(MASTER_OUT_PATH, square_headers, master_rows)
    write_master_csv(IMAGE_DATABASE_OUT_PATH, image_database_fieldnames, image_database_rows)
    write_image_database_xlsx(IMAGE_DATABASE_XLSX_OUT_PATH, image_database_fieldnames, image_database_rows)
    write_review_csv(REVIEW_OUT_PATH, review_rows)

    summary = summarize(
        counts_by_vendor=counts_by_vendor,
        total_source_items=len(source_items),
        review_rows=review_rows,
        included_rows=len(master_rows),
        generated_skus=generated_skus,
        skipped_duplicates=skipped_duplicates,
        merged_groups=merged_groups + same_vendor_merged_groups,
        renamed_rows=renamed_rows,
        retained_gtins=retained_gtins,
        verified_gtins_added=sum(1 for entry in verified_audit_entries if entry.field == "GTIN"),
        catalog_gtins_added=sum(1 for entry in catalog_audit_entries if entry.field == "GTIN"),
        missing_gtins=sum(1 for row in master_rows if not clean_text(row.get("GTIN", ""))),
        seo_titles_generated=sum(1 for row in master_rows if clean_text(row.get("SEO Title", ""))),
        permalinks_generated=permalinks_generated,
        website_match_counts=website_match_counts,
        website_detail_counts=website_detail_counts,
        website_notes=website_notes,
        image_match_counts=image_match_counts,
        image_notes=image_notes,
        image_rows_matched=len(image_audit_entries),
    )
    SUMMARY_OUT_PATH.write_text(summary, encoding="utf-8")
    print(summary)


if __name__ == "__main__":
    main()
