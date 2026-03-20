from __future__ import annotations

import csv
import math
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
OUTPUT_DIR = BASE_DIR / "outputs"
SQUARE_EXPORT_DIR = INPUT_DIR / "square_exports"
CATALOG_WITH_IMAGES_PATH = OUTPUT_DIR / "inventory_database_with_images.csv"
QUEUE_CSV_PATH = OUTPUT_DIR / "square_image_match_queue.csv"
QUEUE_XLSX_PATH = OUTPUT_DIR / "square_image_match_queue.xlsx"
SUMMARY_PATH = OUTPUT_DIR / "square_image_match_summary.txt"
BATCH_DIR = OUTPUT_DIR / "square_image_match_batches"
BATCH_SIZE = 250


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_key(value: object) -> str:
    return "".join(ch for ch in clean_text(value).upper() if ch.isalnum())


def newest_file(folder: Path, patterns: tuple[str, ...]) -> Path:
    matches: list[Path] = []
    for pattern in patterns:
        matches.extend(path for path in folder.glob(pattern) if path.is_file())
    if not matches:
        raise FileNotFoundError(f"No files found in {folder} for {patterns}.")
    return max(matches, key=lambda path: path.stat().st_mtime)


def read_catalog_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [{key: clean_text(value) for key, value in row.items()} for row in reader]


def read_square_export_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [list(row) for row in csv.reader(handle)]
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["Items"] if "Items" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

    header_index = None
    headers: list[str] = []
    for index, row in enumerate(rows):
        cleaned = [clean_text(value) for value in row]
        if "SKU" in cleaned and "Token" in cleaned:
            header_index = index
            headers = cleaned
            break
    if header_index is None:
        raise ValueError(f"Could not find the item header row in {path.name}.")

    records: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        cleaned = [clean_text(value) for value in row]
        if not any(cleaned):
            continue
        padded = cleaned[: len(headers)] + [""] * max(0, len(headers) - len(cleaned))
        records.append({headers[idx]: padded[idx] for idx in range(len(headers))})
    return records


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Image Queue"
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    workbook.save(path)


def clear_directory(path: Path) -> None:
    if not path.exists():
        return
    for child in path.iterdir():
        if child.is_file():
            child.unlink()
        elif child.is_dir():
            shutil.rmtree(child)


def build_queue_rows(catalog_rows: list[dict[str, str]], square_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[str]]:
    notes: list[str] = []
    square_by_sku: dict[str, dict[str, str]] = {}
    duplicate_square_skus: set[str] = set()
    for row in square_rows:
        sku_key = normalize_key(row.get("SKU", ""))
        if not sku_key:
            continue
        if sku_key in square_by_sku:
            duplicate_square_skus.add(sku_key)
            continue
        square_by_sku[sku_key] = row

    if duplicate_square_skus:
        notes.append(f"Square export had {len(duplicate_square_skus)} duplicate SKU keys; first row was kept for batching.")

    queue_rows: list[dict[str, str]] = []
    skipped_missing_square = 0
    skipped_missing_image = 0
    for catalog_row in catalog_rows:
        if clean_text(catalog_row.get("Has Local Image")) != "Y":
            skipped_missing_image += 1
            continue
        sku = clean_text(catalog_row.get("SKU", ""))
        sku_key = normalize_key(sku)
        if not sku_key or sku_key not in square_by_sku:
            skipped_missing_square += 1
            continue
        square_row = square_by_sku[sku_key]
        image_path = Path(clean_text(catalog_row.get("Local Image Absolute Path", "")))
        queue_rows.append(
            {
                "Batch Number": "",
                "Batch Row": "",
                "Square Token": clean_text(square_row.get("Token", "")),
                "Reference Handle": clean_text(square_row.get("Reference Handle", "")),
                "SKU": sku,
                "Square Item Name": clean_text(square_row.get("Item Name", "")),
                "Square Variation Name": clean_text(square_row.get("Variation Name", "")),
                "Catalog Item Name": clean_text(catalog_row.get("Item Name", "")),
                "Customer-facing Name": clean_text(catalog_row.get("Customer-facing Name", "")),
                "Reporting Category": clean_text(catalog_row.get("Reporting Category", "")),
                "Default Vendor Name": clean_text(catalog_row.get("Default Vendor Name", "")),
                "Local Image Filename": image_path.name,
                "Local Image Relative Path": clean_text(catalog_row.get("Local Image Relative Path", "")),
                "Local Image Absolute Path": clean_text(catalog_row.get("Local Image Absolute Path", "")),
                "Website Image URL": clean_text(catalog_row.get("Website Image URL", "")),
            }
        )

    queue_rows.sort(
        key=lambda row: (
            clean_text(row["Default Vendor Name"]).lower(),
            clean_text(row["Reporting Category"]).lower(),
            clean_text(row["Square Item Name"]).lower(),
            clean_text(row["SKU"]).lower(),
        )
    )
    notes.append(f"Catalog rows skipped because they did not have a local image: {skipped_missing_image}.")
    notes.append(f"Catalog rows with local images skipped because no Square SKU match was found: {skipped_missing_square}.")
    return queue_rows, notes


def write_batch_artifacts(rows: list[dict[str, str]], fieldnames: list[str]) -> int:
    clear_directory(BATCH_DIR)
    BATCH_DIR.mkdir(parents=True, exist_ok=True)
    total_batches = math.ceil(len(rows) / BATCH_SIZE) if rows else 0
    for batch_number in range(1, total_batches + 1):
        start = (batch_number - 1) * BATCH_SIZE
        end = start + BATCH_SIZE
        batch_rows = rows[start:end]
        batch_label = f"batch_{batch_number:03d}"
        batch_folder = BATCH_DIR / batch_label
        image_folder = batch_folder / "images"
        batch_folder.mkdir(parents=True, exist_ok=True)
        image_folder.mkdir(parents=True, exist_ok=True)
        for row_index, row in enumerate(batch_rows, start=1):
            row["Batch Number"] = str(batch_number)
            row["Batch Row"] = str(row_index)
            image_path = Path(row["Local Image Absolute Path"])
            if image_path.is_file():
                shutil.copy2(image_path, image_folder / image_path.name)
        write_csv(batch_folder / f"{batch_label}.csv", batch_rows, fieldnames)
    return total_batches


def write_summary(
    square_export_path: Path,
    queue_rows: list[dict[str, str]],
    total_batches: int,
    notes: list[str],
) -> None:
    lines = [
        f"Square export used: {square_export_path}",
        f"Catalog image source: {CATALOG_WITH_IMAGES_PATH}",
        f"Batch size: {BATCH_SIZE}",
        f"Rows ready for image matching: {len(queue_rows)}",
        f"Batch folders created: {total_batches}",
        f"Master queue CSV: {QUEUE_CSV_PATH}",
        f"Master queue workbook: {QUEUE_XLSX_PATH}",
        f"Batch folder root: {BATCH_DIR}",
        "",
        "Recommended use:",
        "1. Open one batch CSV and work that batch only.",
        "2. In Square, select or search the 250 items from that batch by SKU.",
        "3. Bulk-assign the matching images from the Square image library or use the copied batch image folder if needed.",
        "4. Move to the next batch folder after that group is complete.",
        "",
        "Notes:",
    ]
    lines.extend(f"- {note}" for note in notes)
    SUMMARY_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    if not CATALOG_WITH_IMAGES_PATH.exists():
        raise FileNotFoundError(
            f"Missing {CATALOG_WITH_IMAGES_PATH.name}. Run build_master_inventory.py first."
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    square_export_path = newest_file(SQUARE_EXPORT_DIR, ("*.xlsx", "*.csv"))
    catalog_rows = read_catalog_rows(CATALOG_WITH_IMAGES_PATH)
    square_rows = read_square_export_rows(square_export_path)
    queue_rows, notes = build_queue_rows(catalog_rows, square_rows)

    fieldnames = [
        "Batch Number",
        "Batch Row",
        "Square Token",
        "Reference Handle",
        "SKU",
        "Square Item Name",
        "Square Variation Name",
        "Catalog Item Name",
        "Customer-facing Name",
        "Reporting Category",
        "Default Vendor Name",
        "Local Image Filename",
        "Local Image Relative Path",
        "Local Image Absolute Path",
        "Website Image URL",
    ]
    total_batches = write_batch_artifacts(queue_rows, fieldnames)
    write_csv(QUEUE_CSV_PATH, queue_rows, fieldnames)
    write_xlsx(QUEUE_XLSX_PATH, queue_rows, fieldnames)
    write_summary(square_export_path, queue_rows, total_batches, notes)

    print(f"Square image queue: {QUEUE_CSV_PATH}")
    print(f"Square image queue workbook: {QUEUE_XLSX_PATH}")
    print(f"Square image summary: {SUMMARY_PATH}")
    print(f"Batch folders: {BATCH_DIR}")
    print(f"Rows ready for image matching: {len(queue_rows)}")
    print(f"Batches created: {total_batches}")


if __name__ == "__main__":
    main()
