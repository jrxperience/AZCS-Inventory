from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    Font = None
    get_column_letter = None


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
PRICE_UPDATE_DIR = INPUT_DIR / "price_updates"
PRICING_OVERRIDE_DIR = INPUT_DIR / "pricing_overrides"
OUTPUT_DIR = BASE_DIR / "outputs"

MASTER_PATH = OUTPUT_DIR / "square_master_inventory.csv"
LEGACY_MASTER_PATH = BASE_DIR / "square_master_inventory.csv"
SALES_SIGNALS_PATH = OUTPUT_DIR / "sales_catalog_signals.csv"
PRICING_RECOMMENDATIONS_PATH = OUTPUT_DIR / "pricing_recommendations.csv"
PRICING_RECOMMENDATIONS_XLSX_PATH = OUTPUT_DIR / "pricing_recommendations.xlsx"
SQUARE_STRATEGIC_PRICE_UPDATE_PATH = OUTPUT_DIR / "square_catalog_strategic_price_update.csv"
STRATEGIC_MASTER_PATH = OUTPUT_DIR / "square_master_inventory_strategic_pricing.csv"
SUMMARY_PATH = OUTPUT_DIR / "pricing_strategy_summary.txt"
ISSUES_PATH = OUTPUT_DIR / "pricing_strategy_issues.csv"

PERCENT = Decimal("100")


@dataclass
class PriceTotals:
    current_price: str = ""
    last_price_update_date: str = ""
    last_price_reason: str = ""


@dataclass
class PricingOverride:
    override_price: Decimal | None = None
    target_margin_override: Decimal | None = None
    notes: str = ""


@dataclass
class PricingStrategy:
    cost_band: str
    target_margin: Decimal
    minimum_profit: Decimal
    rounding_profile: str
    strategy_tags: list[str]
    pricing_rule: str
    pricing_notes: str


@dataclass
class SalesSignal:
    sales_lines: int = 0
    distinct_sales_items: int = 0
    quantity_sold: Decimal = Decimal("0")
    net_sales: Decimal = Decimal("0")
    average_realized_unit_price: Decimal | None = None
    match_types: str = ""


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_decimal(value: str, field_name: str, path: Path, row_number: int, issues: list[dict[str, str]]) -> Decimal | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        issues.append(
            {
                "source_file": path.name,
                "row_number": str(row_number),
                "issue_type": "invalid_number",
                "sku": "",
                "details": f"Could not parse {field_name} value '{text}'.",
            }
        )
        return None


def parse_percentage(value: str, field_name: str, path: Path, row_number: int, issues: list[dict[str, str]]) -> Decimal | None:
    text = str(value or "").strip().replace("%", "")
    if not text:
        return None
    try:
        parsed = Decimal(text)
    except InvalidOperation:
        issues.append(
            {
                "source_file": path.name,
                "row_number": str(row_number),
                "issue_type": "invalid_percentage",
                "sku": "",
                "details": f"Could not parse {field_name} value '{text}'.",
            }
        )
        return None
    if parsed > Decimal("1"):
        parsed = parsed / PERCENT
    return parsed


def format_money(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value.quantize(Decimal('0.01'))}"


def format_percent(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{(value * PERCENT).quantize(Decimal('0.1'))}%"


def resolve_master_path() -> Path:
    if MASTER_PATH.exists():
        return MASTER_PATH
    if LEGACY_MASTER_PATH.exists():
        return LEGACY_MASTER_PATH
    raise FileNotFoundError("square_master_inventory.csv was not found in outputs/ or the repo root.")


def parse_date(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def load_price_updates(master_rows: list[dict[str, str]], issues: list[dict[str, str]]) -> dict[str, PriceTotals]:
    totals: dict[str, PriceTotals] = {}
    known_skus = {row.get("SKU", "").strip() for row in master_rows if row.get("SKU", "").strip()}

    for path in sorted(PRICE_UPDATE_DIR.glob("*.csv")):
        rows = read_csv_rows(path)
        for row_number, row in enumerate(rows, start=2):
            sku = str(row.get("SKU", "")).strip()
            new_price = parse_decimal(row.get("New Price", ""), "New Price", path, row_number, issues)
            tx_date = parse_date(row.get("Transaction Date", ""))
            reason = str(row.get("Reason", "")).strip()
            if not sku or new_price is None:
                continue
            if sku not in known_skus:
                issues.append(
                    {
                        "source_file": path.name,
                        "row_number": str(row_number),
                        "issue_type": "unknown_sku",
                        "sku": sku,
                        "details": "Price update row SKU does not exist in the master inventory.",
                    }
                )
                continue
            current = totals.get(sku, PriceTotals())
            replace_current = not current.last_price_update_date or tx_date >= current.last_price_update_date
            if replace_current:
                totals[sku] = PriceTotals(
                    current_price=format_money(new_price),
                    last_price_update_date=tx_date,
                    last_price_reason=reason,
                )
    return totals


def load_pricing_overrides(master_rows: list[dict[str, str]], issues: list[dict[str, str]]) -> dict[str, PricingOverride]:
    overrides: dict[str, PricingOverride] = {}
    known_skus = {row.get("SKU", "").strip() for row in master_rows if row.get("SKU", "").strip()}

    for path in sorted(PRICING_OVERRIDE_DIR.glob("*.csv")):
        rows = read_csv_rows(path)
        for row_number, row in enumerate(rows, start=2):
            sku = str(row.get("SKU", "")).strip()
            if not sku:
                continue
            if sku not in known_skus:
                issues.append(
                    {
                        "source_file": path.name,
                        "row_number": str(row_number),
                        "issue_type": "unknown_sku",
                        "sku": sku,
                        "details": "Pricing override row SKU does not exist in the master inventory.",
                    }
                )
                continue

            override_price = parse_decimal(row.get("Override Price", ""), "Override Price", path, row_number, issues)
            target_margin_override = parse_percentage(
                row.get("Target Margin Override", ""),
                "Target Margin Override",
                path,
                row_number,
                issues,
            )
            notes = str(row.get("Notes", "")).strip()
            if override_price is None and target_margin_override is None and not notes:
                continue
            overrides[sku] = PricingOverride(
                override_price=override_price,
                target_margin_override=target_margin_override,
                notes=notes,
            )

    return overrides


def load_sales_signals(master_rows: list[dict[str, str]], issues: list[dict[str, str]]) -> dict[str, SalesSignal]:
    if not SALES_SIGNALS_PATH.exists():
        return {}

    known_skus = {row.get("SKU", "").strip() for row in master_rows if row.get("SKU", "").strip()}
    signals: dict[str, SalesSignal] = {}
    rows = read_csv_rows(SALES_SIGNALS_PATH)
    for row_number, row in enumerate(rows, start=2):
        sku = str(row.get("Matched Master SKU", "")).strip()
        if not sku:
            continue
        if sku not in known_skus:
            issues.append(
                {
                    "source_file": SALES_SIGNALS_PATH.name,
                    "row_number": str(row_number),
                    "issue_type": "unknown_sku",
                    "sku": sku,
                    "details": "Sales signal row SKU does not exist in the master inventory.",
                }
            )
            continue

        sales_lines = parse_decimal(row.get("Sales Lines", ""), "Sales Lines", SALES_SIGNALS_PATH, row_number, issues) or Decimal("0")
        distinct_sales_items = parse_decimal(
            row.get("Distinct Sales Items", ""),
            "Distinct Sales Items",
            SALES_SIGNALS_PATH,
            row_number,
            issues,
        ) or Decimal("0")
        quantity_sold = parse_decimal(row.get("Quantity Sold", ""), "Quantity Sold", SALES_SIGNALS_PATH, row_number, issues) or Decimal("0")
        net_sales = parse_decimal(row.get("Net Sales", ""), "Net Sales", SALES_SIGNALS_PATH, row_number, issues) or Decimal("0")
        average_realized_unit_price = parse_decimal(
            row.get("Average Realized Unit Price", ""),
            "Average Realized Unit Price",
            SALES_SIGNALS_PATH,
            row_number,
            issues,
        )

        signals[sku] = SalesSignal(
            sales_lines=int(sales_lines),
            distinct_sales_items=int(distinct_sales_items),
            quantity_sold=quantity_sold,
            net_sales=net_sales,
            average_realized_unit_price=average_realized_unit_price,
            match_types=str(row.get("Match Types", "")).strip(),
        )

    return signals


def round_up_to_step(value: Decimal, step: Decimal) -> Decimal:
    units = (value / step).to_integral_value(rounding=ROUND_CEILING)
    return units * step


def round_retail_price(value: Decimal, profile: str) -> Decimal:
    if profile == "ending_.99":
        return round_up_to_step(value + Decimal("0.01"), Decimal("1.00")) - Decimal("0.01")
    if profile == "ending_5_minus_.01":
        return round_up_to_step(value + Decimal("0.01"), Decimal("5.00")) - Decimal("0.01")
    if profile == "ending_10_minus_1":
        return round_up_to_step(value + Decimal("1.00"), Decimal("10.00")) - Decimal("1.00")
    if profile == "ending_50_minus_1":
        return round_up_to_step(value + Decimal("1.00"), Decimal("50.00")) - Decimal("1.00")
    return value.quantize(Decimal("0.01"))


def base_cost_strategy(cost: Decimal) -> tuple[str, Decimal, Decimal, str]:
    if cost < Decimal("5"):
        return "0-4.99", Decimal("0.29"), Decimal("2.00"), "ending_.99"
    if cost < Decimal("10"):
        return "5-9.99", Decimal("0.28"), Decimal("3.00"), "ending_.99"
    if cost < Decimal("20"):
        return "10-19.99", Decimal("0.27"), Decimal("4.50"), "ending_.99"
    if cost < Decimal("50"):
        return "20-49.99", Decimal("0.25"), Decimal("7.00"), "ending_.99"
    if cost < Decimal("100"):
        return "50-99.99", Decimal("0.24"), Decimal("10.00"), "ending_.99"
    if cost < Decimal("250"):
        return "100-249.99", Decimal("0.22"), Decimal("16.00"), "ending_5_minus_.01"
    if cost < Decimal("500"):
        return "250-499.99", Decimal("0.20"), Decimal("30.00"), "ending_10_minus_1"
    if cost < Decimal("1000"):
        return "500-999.99", Decimal("0.17"), Decimal("60.00"), "ending_10_minus_1"
    return "1000+", Decimal("0.14"), Decimal("125.00"), "ending_50_minus_1"


def build_search_blob(row: dict[str, str]) -> str:
    parts = [
        row.get("Item Name", ""),
        row.get("Customer-facing Name", ""),
        row.get("Variation Name", ""),
        row.get("Description", ""),
        row.get("Categories", ""),
        row.get("Reporting Category", ""),
        row.get("Default Vendor Name", ""),
        row.get("Default Vendor Code", ""),
    ]
    return " | ".join(str(part) for part in parts).upper()


def contains_any(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def detect_strategy_tags(row: dict[str, str]) -> list[str]:
    blob = build_search_blob(row)
    tags: list[str] = []

    if contains_any(
        blob,
        (
            "SOAP",
            "CLEANER",
            "DEGREASER",
            "DETERGENT",
            "SURFACTANT",
            "ASSASSIN",
            "RESTORATION",
            "SEALER",
            "OXID",
            "REMOVE",
            "BRIGHTENER",
            "NEUTRAL",
            "STAIN",
            " WASH",
            "RTU",
            "CONCENTRATE",
            "CHEM",
        ),
    ):
        tags.append("chemical")

    if contains_any(blob, ("55 GALLON", "5 GALLON", "PAIL", "DRUM", "KIT", "SET", "PACK", "COMPLETE", "BUNDLE", "CASE")):
        tags.append("bundle")

    if contains_any(
        blob,
        ("PRESSURE WASHER", "SURFACE CLEANER", "PUMP", "ENGINE", "GENERATOR", "TRAILER", "SKID", "COMPRESSOR", "SOFT WASH SYSTEM"),
    ):
        tags.append("equipment")

    if "equipment" not in tags and contains_any(
        blob,
        ("NOZZLE", "QUICK CONNECT", "COUPLER", "PLUG", "O-RING", "FITTING", "ADAPTER", "NIPPLE", "BUSHING", "ELBOW", "TEE", "UNION", "SEAL KIT", "REPAIR KIT", "VALVE KIT"),
    ):
        tags.append("commodity_part")

    if "equipment" not in tags and "chemical" not in tags and contains_any(
        blob,
        ("ETTORE", "SORBO", "MOERMAN", "TUCKER", "SQUEEGEE", "CHANNEL", "SCRAPER", "T-BAR", "MICROFIBER", "APPLICATOR", "WINDOW CLEANING", "POLE", "BRUSH"),
    ):
        tags.append("premium_tool")

    return tags


def build_pricing_strategy(row: dict[str, str], cost: Decimal, override: PricingOverride | None) -> PricingStrategy:
    cost_band, target_margin, minimum_profit, rounding_profile = base_cost_strategy(cost)
    tags = detect_strategy_tags(row)
    notes = [f"Base cost band {cost_band}."]

    if "chemical" in tags:
        target_margin += Decimal("0.07")
        notes.append("Chemical/restoration products carry a stronger margin target.")
    if "premium_tool" in tags:
        target_margin += Decimal("0.06")
        notes.append("Premium cleaning tools support a stronger margin target.")
    if "bundle" in tags:
        target_margin += Decimal("0.02")
        notes.append("Bundle or bulk packaging gets a modest premium.")
    if "commodity_part" in tags:
        target_margin -= Decimal("0.03")
        notes.append("Commodity replacement parts stay more competitive.")
    if "equipment" in tags:
        target_margin -= Decimal("0.04")
        notes.append("Larger equipment uses a thinner margin target.")

    if str(row.get("Default Vendor Name", "")).strip().upper() == "TRIDENT":
        target_margin -= Decimal("0.02")
        notes.append("Trident protective items are kept more competitive.")

    if override and override.target_margin_override is not None:
        target_margin = override.target_margin_override
        notes.append("Manual target margin override applied.")

    target_margin = min(Decimal("0.40"), max(Decimal("0.14"), target_margin))
    pricing_rule = " + ".join([cost_band] + tags) if tags else cost_band

    return PricingStrategy(
        cost_band=cost_band,
        target_margin=target_margin,
        minimum_profit=minimum_profit,
        rounding_profile=rounding_profile,
        strategy_tags=tags,
        pricing_rule=pricing_rule,
        pricing_notes=" ".join(notes),
    )


def compute_price_metrics(cost: Decimal, price: Decimal | None) -> tuple[Decimal | None, Decimal | None]:
    if price is None or price <= Decimal("0") or cost <= Decimal("0"):
        return None, None
    markup = (price - cost) / cost
    margin = (price - cost) / price
    return markup, margin


def classify_sales_demand(signal: SalesSignal | None) -> str:
    if signal is None:
        return ""
    if signal.net_sales >= Decimal("10000") or signal.quantity_sold >= Decimal("50") or signal.sales_lines >= 20:
        return "core"
    if signal.net_sales >= Decimal("2500") or signal.quantity_sold >= Decimal("15") or signal.sales_lines >= 8:
        return "steady"
    if signal.net_sales >= Decimal("500") or signal.quantity_sold >= Decimal("5") or signal.sales_lines >= 3:
        return "emerging"
    return "light"


def sales_realized_margin(cost: Decimal, signal: SalesSignal | None) -> Decimal | None:
    if signal is None or signal.average_realized_unit_price is None or signal.average_realized_unit_price <= Decimal("0") or cost <= Decimal("0"):
        return None
    return (signal.average_realized_unit_price - cost) / signal.average_realized_unit_price


def apply_sales_signal(strategy: PricingStrategy, cost: Decimal, signal: SalesSignal | None) -> tuple[PricingStrategy, str, bool]:
    if signal is None:
        return strategy, "", False

    notes: list[str] = []
    trusted = False
    demand_tier = classify_sales_demand(signal)
    realized_margin = sales_realized_margin(cost, signal)

    if realized_margin is not None and realized_margin >= Decimal("0.10"):
        trusted = True
        margin_adjustment = Decimal("0")
        if demand_tier == "core":
            margin_adjustment = Decimal("0.03")
        elif demand_tier == "steady":
            margin_adjustment = Decimal("0.02")
        elif demand_tier == "emerging":
            margin_adjustment = Decimal("0.01")

        if margin_adjustment:
            strategy.target_margin -= margin_adjustment
            notes.append(f"Sales history marks this as a {demand_tier} seller, so the margin target was tightened.")

    if signal.average_realized_unit_price is not None and realized_margin is not None:
        if realized_margin < Decimal("0"):
            notes.append("Realized sales price trends below cost, so sales data was not used to lower the recommendation.")
        elif realized_margin < Decimal("0.10"):
            notes.append("Sales history exists, but realized margin is too thin to trust for a lower target.")

    strategy.target_margin = min(Decimal("0.40"), max(Decimal("0.12"), strategy.target_margin))
    if demand_tier:
        strategy.pricing_rule = f"{strategy.pricing_rule} + sales_{demand_tier}"
    if notes:
        strategy.pricing_notes = f"{strategy.pricing_notes} {' '.join(notes)}".strip()

    sales_note = ""
    if signal.average_realized_unit_price is not None:
        sales_note = f"{demand_tier or 'none'} seller with average realized price {format_money(signal.average_realized_unit_price)}."

    return strategy, sales_note, trusted


def compute_recommended_price(cost: Decimal, strategy: PricingStrategy) -> Decimal:
    margin_price = cost / (Decimal("1") - strategy.target_margin)
    floor_price = cost + strategy.minimum_profit
    raw_price = max(margin_price, floor_price)
    rounded_price = round_retail_price(raw_price, strategy.rounding_profile)
    minimum_price = (cost + Decimal("0.50")).quantize(Decimal("0.01"))
    if rounded_price <= cost:
        rounded_price = minimum_price
    return rounded_price.quantize(Decimal("0.01"))


def competitive_price_cap(row: dict[str, str]) -> tuple[Decimal | None, str]:
    blob = build_search_blob(row)
    if "HURRICANE CAT 5" in blob and "KIT" in blob and "HALF KIT" not in blob:
        return Decimal("599.99"), "Competitive cap kept Hurricane Cat 5 kits near the proven 599.99 price point."
    return None, ""


def choose_current_price(row: dict[str, str], updates: dict[str, PriceTotals]) -> tuple[Decimal | None, str, str]:
    sku = row.get("SKU", "").strip()
    update = updates.get(sku)
    if update and update.current_price:
        return Decimal(update.current_price), update.last_price_update_date, update.last_price_reason

    base_price_text = str(row.get("Price", "")).strip() or str(row.get("Price AZCS", "")).strip()
    if not base_price_text:
        return None, "", ""
    try:
        return Decimal(base_price_text), "", ""
    except InvalidOperation:
        return None, "", ""


def write_recommendations_xlsx(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> tuple[bool, str]:
    if Workbook is None or get_column_letter is None:
        return False, "openpyxl not available"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing Recommendations"
    sheet.append(fieldnames)
    if Font is not None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}1"

    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])

    width_overrides = {
        "A": 14,
        "B": 44,
        "C": 28,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 18,
        "O": 18,
        "P": 18,
        "Q": 18,
        "R": 20,
        "S": 20,
        "T": 60,
        "U": 60,
        "V": 18,
        "W": 18,
        "X": 16,
        "Y": 14,
        "Z": 16,
        "AA": 18,
        "AB": 18,
        "AC": 18,
        "AD": 24,
        "AE": 20,
        "AF": 18,
    }
    for column, width in width_overrides.items():
        sheet.column_dimensions[column].width = width

    try:
        workbook.save(path)
    except PermissionError:
        return False, f"file locked: {path}"
    return True, str(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PRICING_OVERRIDE_DIR.mkdir(parents=True, exist_ok=True)

    issues: list[dict[str, str]] = []
    master_path = resolve_master_path()
    master_rows = read_csv_rows(master_path)
    if not master_rows:
        raise ValueError("Master inventory is empty.")

    fieldnames = list(master_rows[0].keys())
    price_updates = load_price_updates(master_rows, issues)
    pricing_overrides = load_pricing_overrides(master_rows, issues)
    sales_signals = load_sales_signals(master_rows, issues)

    recommendation_rows: list[dict[str, str]] = []
    strategic_master_rows: list[dict[str, str]] = []

    costed_rows = 0
    current_priced_rows = 0
    fill_missing_count = 0
    raise_low_margin_count = 0
    raise_sales_supported_count = 0
    keep_current_count = 0
    manual_override_count = 0
    preserved_high_margin_count = 0
    rows_with_sales_history = 0
    trusted_sales_signal_count = 0
    sales_anchor_raise_count = 0
    suggested_margin_total = Decimal("0")
    current_margin_total = Decimal("0")

    for row in master_rows:
        updated_row = dict(row)
        sku = row.get("SKU", "").strip()
        cost_text = str(row.get("Default Unit Cost", "")).strip()
        try:
            cost = Decimal(cost_text) if cost_text else None
        except InvalidOperation:
            cost = None

        current_price, last_update_date, last_update_reason = choose_current_price(row, price_updates)
        override = pricing_overrides.get(sku)
        sales_signal = sales_signals.get(sku)
        current_markup, current_margin = compute_price_metrics(cost, current_price) if cost else (None, None)
        suggested_price: Decimal | None = current_price
        suggested_markup: Decimal | None = current_markup
        suggested_margin: Decimal | None = current_margin
        strategy: PricingStrategy | None = None
        pricing_action = "no_cost"
        pricing_notes = ""
        sales_pricing_note = ""
        sales_signal_trusted = False
        sales_demand_tier = classify_sales_demand(sales_signal)
        competitive_note = ""

        if cost is not None and cost > Decimal("0"):
            costed_rows += 1
            strategy = build_pricing_strategy(row, cost, override)
            strategy, sales_pricing_note, sales_signal_trusted = apply_sales_signal(strategy, cost, sales_signal)
            computed_price = compute_recommended_price(cost, strategy)

            if sales_signal is not None:
                rows_with_sales_history += 1
            if sales_signal_trusted:
                trusted_sales_signal_count += 1

            if (
                sales_signal_trusted
                and sales_signal is not None
                and sales_signal.average_realized_unit_price is not None
                and sales_signal.sales_lines >= 3
            ):
                sales_anchor_price = round_retail_price(
                    sales_signal.average_realized_unit_price,
                    strategy.rounding_profile,
                ).quantize(Decimal("0.01"))
                if sales_anchor_price > computed_price * Decimal("1.03"):
                    computed_price = sales_anchor_price
                    sales_anchor_raise_count += 1
                    anchor_note = "Actual realized selling price supports a higher anchor."
                    sales_pricing_note = f"{sales_pricing_note} {anchor_note}".strip()

            price_cap, cap_note = competitive_price_cap(row)
            if price_cap is not None and computed_price > price_cap:
                computed_price = price_cap
                competitive_note = cap_note

            if current_price is not None and current_price > Decimal("0"):
                current_priced_rows += 1
                current_margin_total += current_margin or Decimal("0")

            if override and override.override_price is not None:
                suggested_price = override.override_price.quantize(Decimal("0.01"))
                pricing_action = "manual_override"
                manual_override_count += 1
                pricing_notes = "Manual price override applied."
            elif current_price is None or current_price <= Decimal("0"):
                suggested_price = computed_price
                pricing_action = "fill_missing_price"
                fill_missing_count += 1
                pricing_notes = "Filled missing selling price from the strategic pricing engine."
            else:
                low_margin_floor = max(Decimal("0.10"), strategy.target_margin - Decimal("0.10"))
                high_margin_ceiling = strategy.target_margin + Decimal("0.18")
                if current_price <= cost or (current_margin is not None and current_margin < low_margin_floor):
                    suggested_price = computed_price
                    pricing_action = "raise_low_margin"
                    raise_low_margin_count += 1
                    pricing_notes = "Current price sits below the minimum acceptable margin floor."
                elif sales_signal_trusted and suggested_price is not None and computed_price > current_price * Decimal("1.05"):
                    suggested_price = computed_price
                    pricing_action = "raise_sales_supported"
                    raise_sales_supported_count += 1
                    pricing_notes = "Trusted sales history supports a higher strategic price."
                else:
                    suggested_price = current_price
                    pricing_action = "keep_current"
                    keep_current_count += 1
                    pricing_notes = "Current price is within the acceptable strategic band and was preserved."
                    if (
                        sales_signal_trusted
                        and sales_signal is not None
                        and sales_signal.average_realized_unit_price is not None
                        and suggested_price is not None
                        and sales_signal.average_realized_unit_price > Decimal("0")
                    ):
                        realized_gap = abs(suggested_price - sales_signal.average_realized_unit_price) / sales_signal.average_realized_unit_price
                        if realized_gap <= Decimal("0.06"):
                            sales_alignment_note = "Current price is aligned with recent realized selling history."
                            sales_pricing_note = f"{sales_pricing_note} {sales_alignment_note}".strip()
                    if current_margin is not None and current_margin > high_margin_ceiling:
                        preserved_high_margin_count += 1
                        pricing_notes = "Current price is above the target band and was preserved for review."

            suggested_markup, suggested_margin = compute_price_metrics(cost, suggested_price)
            suggested_margin_total += suggested_margin or Decimal("0")

            updated_row["Price"] = format_money(suggested_price)
            if "Price AZCS" in updated_row:
                updated_row["Price AZCS"] = format_money(suggested_price)
            if suggested_price is not None:
                updated_row["Sellable"] = "Y"

        combined_notes = " ".join(note for note in (pricing_notes, sales_pricing_note, competitive_note) if note).strip()

        recommendation_rows.append(
            {
                "SKU": sku,
                "Item Name": row.get("Item Name", ""),
                "Customer-facing Name": row.get("Customer-facing Name", ""),
                "Default Vendor Name": row.get("Default Vendor Name", ""),
                "Reporting Category": row.get("Reporting Category", ""),
                "Default Vendor Code": row.get("Default Vendor Code", ""),
                "Default Unit Cost": format_money(cost),
                "Current Selling Price": format_money(current_price),
                "Current Gross Margin %": format_percent(current_margin),
                "Current Markup %": format_percent(current_markup),
                "Suggested Selling Price": format_money(suggested_price),
                "Suggested Gross Margin %": format_percent(suggested_margin),
                "Suggested Markup %": format_percent(suggested_markup),
                "Target Gross Margin %": format_percent(strategy.target_margin) if strategy else "",
                "Minimum Dollar Profit": format_money(strategy.minimum_profit) if strategy else "",
                "Cost Band": strategy.cost_band if strategy else "",
                "Strategy Tags": ", ".join(strategy.strategy_tags) if strategy else "",
                "Rounding Profile": strategy.rounding_profile if strategy else "",
                "Pricing Rule": strategy.pricing_rule if strategy else "",
                "Pricing Action": pricing_action,
                "Pricing Notes": combined_notes,
                "Strategy Notes": strategy.pricing_notes if strategy else "",
                "Sales Demand Tier": sales_demand_tier,
                "Sales Lines": str(sales_signal.sales_lines) if sales_signal else "",
                "Sales Quantity Sold": format_money(sales_signal.quantity_sold) if sales_signal else "",
                "Sales Net Sales": format_money(sales_signal.net_sales) if sales_signal else "",
                "Average Realized Unit Price": format_money(sales_signal.average_realized_unit_price) if sales_signal else "",
                "Sales Signal Trusted": "Y" if sales_signal_trusted else "",
                "Sales Match Types": sales_signal.match_types if sales_signal else "",
                "Sales Pricing Note": sales_pricing_note,
                "Last Price Update Date": last_update_date,
                "Last Price Update Reason": last_update_reason,
                "Override Notes": override.notes if override else "",
            }
        )
        strategic_master_rows.append(updated_row)

    recommendation_fieldnames = list(recommendation_rows[0].keys()) if recommendation_rows else []
    write_csv(PRICING_RECOMMENDATIONS_PATH, recommendation_fieldnames, recommendation_rows)
    write_csv(STRATEGIC_MASTER_PATH, fieldnames, strategic_master_rows)
    write_csv(SQUARE_STRATEGIC_PRICE_UPDATE_PATH, fieldnames, strategic_master_rows)
    write_csv(ISSUES_PATH, ["source_file", "row_number", "issue_type", "sku", "details"], issues)
    xlsx_written, xlsx_status = write_recommendations_xlsx(
        PRICING_RECOMMENDATIONS_XLSX_PATH,
        recommendation_rows,
        recommendation_fieldnames,
    )

    average_current_margin = (current_margin_total / Decimal(str(current_priced_rows))) if current_priced_rows else Decimal("0")
    average_suggested_margin = (suggested_margin_total / Decimal(str(costed_rows))) if costed_rows else Decimal("0")

    summary_lines = [
        f"Master inventory source: {master_path}",
        f"Pricing recommendations CSV: {PRICING_RECOMMENDATIONS_PATH}",
        f"Pricing recommendations Excel: {xlsx_status}",
        f"Strategic master inventory: {STRATEGIC_MASTER_PATH}",
        f"Square strategic price update file: {SQUARE_STRATEGIC_PRICE_UPDATE_PATH}",
        f"Issues file: {ISSUES_PATH}",
        f"Price update files processed: {len(list(PRICE_UPDATE_DIR.glob('*.csv')))}",
        f"Pricing override files processed: {len(list(PRICING_OVERRIDE_DIR.glob('*.csv')))}",
        f"Sales signal rows loaded: {len(sales_signals)}",
        f"Rows with usable cost: {costed_rows}",
        f"Rows with current selling price: {current_priced_rows}",
        f"Rows with sales history: {rows_with_sales_history}",
        f"Rows with trusted sales signals: {trusted_sales_signal_count}",
        f"Rows raised by realized sales anchor: {sales_anchor_raise_count}",
        f"Rows with missing price filled: {fill_missing_count}",
        f"Rows raised from low margin: {raise_low_margin_count}",
        f"Rows raised from trusted sales history: {raise_sales_supported_count}",
        f"Rows kept at current price: {keep_current_count}",
        f"Rows using manual overrides: {manual_override_count}",
        f"Rows preserved above target band: {preserved_high_margin_count}",
        f"Average current gross margin on already-priced rows: {format_percent(average_current_margin)}",
        f"Average suggested gross margin across all costed rows: {format_percent(average_suggested_margin)}",
        "Assumption: the pricing engine targets roughly 30% gross margin across the full catalog, not a flat 30% markup.",
        "Rule summary: lower-cost items carry higher markup percentages, large equipment carries lower margin targets, chemicals / premium tools / bundles get higher strategic targets than commodity parts, and trusted sales history can tighten the target or support a higher price anchor.",
    ]
    SUMMARY_PATH.write_text("\n".join(summary_lines), encoding="utf-8")
    print("\n".join(summary_lines))


if __name__ == "__main__":
    main()
